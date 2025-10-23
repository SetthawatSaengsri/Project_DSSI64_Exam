from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.contrib.auth import login, logout,update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.http import require_POST, require_GET ,require_http_methods
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q, Sum
from django.utils import timezone
import time as time_module
from django.db import transaction, IntegrityError
from datetime import datetime
import json, base64, qrcode
from io import BytesIO
from django.http import FileResponse, HttpResponse, HttpResponseRedirect
from django.urls import reverse
from .forms import *
from .models import *
from django.core.cache import cache
from django.conf import settings
from contextlib import contextmanager
import logging

# ========================= หน้าหลักและ Authentication =========================
# แสดงหน้าแรก พร้อมฟอร์มสมัครเจ้าหน้าที่ และรับ POST สมัครสมาชิก
def index_view(request):
    """หน้าแรกของระบบ"""
    if request.method == 'POST':
        form = StaffRegistrationForm(request.POST)  
        if form.is_valid():
            user = form.save()
            messages.success(request, f'สมัครสมาชิกสำเร็จ! ยินดีต้อนรับ {user.get_full_name()}')
            return redirect('index_view')
    else:
        form = StaffRegistrationForm()
    
    return render(request, 'app/index.html', {'form': form})

# ตรวจสอบอีเมล/รหัสผ่าน แล้วพาไป Dashboard ตามบทบาท
def login_user(request):
    """เข้าสู่ระบบ"""
    if request.method == 'POST':
        email = request.POST['email']
        password = request.POST['password']

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            messages.error(request, 'ไม่พบผู้ใช้งานนี้ในระบบ')
            return redirect('index_view')

        if user.check_password(password):
            if not user.is_active:
                messages.error(request, 'บัญชีของคุณยังไม่ได้รับการอนุมัติ')
                return redirect('index_view')

            login(request, user)
            
            # Redirect ตามบทบาท
            if user.is_superuser:
                return redirect('dashboard_admin')
            elif user.is_student:
                return redirect('dashboard_student')
            elif user.is_teacher:
                return redirect('dashboard_teacher')
            elif user.is_staff:
                return redirect('dashboard_staff')
        else:
            messages.error(request, 'รหัสผ่านไม่ถูกต้อง')

    return redirect('index_view')

# ออกจากระบบ แล้วส่งกลับหน้าแรก
@login_required
def logout_user(request):
    """ออกจากระบบ"""
    logout(request)
    return redirect('index_view')

# ========================= Dashboard สำหรับแต่ละบทบาท =========================
# Dashboard ผู้ดูแลระบบ
@staff_member_required
def dashboard_admin(request):
    """Dashboard ผู้ดูแลระบบ"""
    stats = {
        'teachers': TeacherProfile.objects.count(),
        'students': StudentProfile.objects.count(),
        'subjects': ExamSubject.objects.count(),
        'rooms': ExamRoom.objects.count(),
    }
    return render(request, 'app/admin/dashboard_admin.html', {'stats': stats})

# Dashboard เจ้าหน้าที่ แสดงสถิติรวม
@login_required
def dashboard_staff(request):
    """Dashboard เจ้าหน้าที่"""
    if not request.user.is_staff:
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึงส่วนนี้")
    
    # สถิติพื้นฐาน
    stats = {
        'teachers': TeacherProfile.objects.count(),
        'students': StudentProfile.objects.count(),
        'subjects': ExamSubject.objects.count(),
        'attendance_today': Attendance.objects.filter(
            checkin_time__date=timezone.now().date()
        ).count(),
    }
    
    # วิชาสอบที่กำลังจะมาถึง
    upcoming_exams = ExamSubject.objects.filter(
        exam_date__gte=timezone.now().date()
    ).order_by('exam_date', 'start_time')[:5]
    
    return render(request, 'app/staff/dashboard_staff.html', {
        'stats': stats,
        'upcoming_exams': upcoming_exams
    })

@login_required
def dashboard_staff_stats(request):
    """API สำรองสำหรับสถิติแดชบอร์ด - แก้ไขให้ส่งข้อมูลครบ"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    try:
        # ดึงข้อมูลจริงจากฐานข้อมูล
        total_subjects = ExamSubject.objects.count()
        total_students = StudentProfile.objects.count()
        
        # นับการเข้าสอบตามสถานะ
        all_attendance = Attendance.objects.all()
        on_time_count = all_attendance.filter(status='on_time').count()
        late_count = all_attendance.filter(status='late').count()
        absent_count = all_attendance.filter(status='absent').count()
        excused_count = all_attendance.filter(status='excused').count()
        cheating_count = all_attendance.filter(status='cheating').count()
        
        # คำนวณเข้าสอบรวม
        attended_count = on_time_count + late_count
        
        # คำนวณอัตรา
        total_attendance_records = all_attendance.count()
        attendance_rate = (attended_count / total_attendance_records * 100) if total_attendance_records > 0 else 0
        
        # สถิติตามสถานะ
        attendance_stats = {
            'on_time': on_time_count,
            'late': late_count,
            'absent': absent_count,
            'excused': excused_count,
            'cheating': cheating_count,
        }
        
        return JsonResponse({
            'statistics': {
                'subjects': total_subjects,
                'students': total_students,
                'on_time': on_time_count,
                'late': late_count,
                'absent': absent_count,
                'excused': excused_count,
                'cheating': cheating_count,
                'attended': attended_count,
                'rate': round(attendance_rate, 1)
            },
            'chartData': {
                'labels': ['เข้าสอบตรงเวลา', 'เข้าสอบสาย', 'ขาดสอบ', 'ลาป่วย', 'ทุจริต'],
                'datasets': [{
                    'label': 'จำนวนนักเรียน',
                    'data': list(attendance_stats.values()),
                    'backgroundColor': [
                        'rgba(34, 197, 94, 0.8)',   # เขียว - ตรงเวลา
                        'rgba(234, 179, 8, 0.8)',   # เหลือง - สาย
                        'rgba(239, 68, 68, 0.8)',   # แดง - ขาด
                        'rgba(59, 130, 246, 0.8)',  # น้ำเงิน - ลา
                        'rgba(236, 72, 153, 0.8)'   # ชมพู - ทุจริต
                    ],
                    'borderColor': 'rgba(34, 197, 94, 0.8)',
                    'borderWidth': 2,
                    'fill': False
                }]
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}'
        }, status=500)
    
@login_required
def get_academic_years_api(request):
    """API สำหรับดึงรายการปีการศึกษาที่มีข้อมูล"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    try:
        years = ExamSubject.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year')
        return JsonResponse({'years': list(years)})
    except Exception as e:
        # ส่งปีเริ่มต้นหากเกิดข้อผิดพลาด
        return JsonResponse({'years': ['2567', '2566', '2565', '2564']})

@login_required
def exam_statistics_api(request):
    """API สำหรับดึงสถิติการสอบสำหรับกราฟ"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    try:
        # สถิติรายปี
        year_stats = []
        years = ExamSubject.objects.values_list('academic_year', flat=True).distinct().order_by('academic_year')
        for year in years:
            year_subjects = ExamSubject.objects.filter(academic_year=year)
            # คำนวณจำนวนนักเรียนรวมในแต่ละปี
            total_students = sum(s.students.count() for s in year_subjects)
            year_stats.append({
                'year': year,
                'subjects': year_subjects.count(),
                'students': total_students
            })
        
        # สถิติรายเทอม
        term_stats = []
        for term in ['1', '2', '3']:
            term_subjects = ExamSubject.objects.filter(term=term)
            total_students = sum(s.students.count() for s in term_subjects)
            term_stats.append({
                'term': int(term),
                'subjects': term_subjects.count(),
                'students': total_students
            })
        
        # สถิติรายชั้น
        class_stats = []
        classes = StudentProfile.objects.values_list('student_class', flat=True).distinct()
        for class_name in classes:
            if class_name:  # ข้าม class ที่เป็น None
                students_in_class = StudentProfile.objects.filter(student_class=class_name)
                total_attendance = Attendance.objects.filter(
                    student__student_class=class_name
                ).count()
                
                # คำนวณจำนวนการสอบทั้งหมดสำหรับชั้นนี้
                total_possible = students_in_class.count() * ExamSubject.objects.filter(
                    students__student_class=class_name
                ).distinct().count()
                
                attendance_rate = (total_attendance / total_possible * 100) if total_possible > 0 else 0
                
                class_stats.append({
                    'class': class_name,
                    'students': students_in_class.count(),
                    'attendance_rate': round(attendance_rate, 1)
                })
        
        # สถิติรวม
        all_attendance = Attendance.objects.all()
        overall_stats = {
            'on_time': all_attendance.filter(status='on_time').count(),
            'late': all_attendance.filter(status='late').count(),
            'absent': all_attendance.filter(status='absent').count(),
            'excused': all_attendance.filter(status='excused').count(),
        }
        
        return JsonResponse({
            'year_stats': year_stats,
            'term_stats': term_stats,
            'class_stats': class_stats,
            'overall_stats': overall_stats
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}'
        }, status=500)

@login_required  
def dashboard_data_api(request):
    """API หลักสำหรับดึงข้อมูลแดชบอร์ดตามตัวกรอง - แก้ไขให้ส่งข้อมูลครบทุกสถานะ"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    try:
        data_type = request.GET.get('dataType', 'all')
        year = request.GET.get('year', '')
        term = request.GET.get('term', '')
        class_filter = request.GET.get('class', '')
        
        print(f"Dashboard API called with: dataType={data_type}, year={year}, term={term}, class={class_filter}")
        
        # สร้าง queryset สำหรับ subjects ตามตัวกรอง
        subjects_query = ExamSubject.objects.all()
        if year:
            subjects_query = subjects_query.filter(academic_year=year)
        if term:
            subjects_query = subjects_query.filter(term=term)
        if class_filter:
            subjects_query = subjects_query.filter(students__student_class=class_filter).distinct()
        
        # สถิติพื้นฐาน
        total_subjects = subjects_query.count()
        
        # นับจำนวนนักเรียนจากรายวิชาที่ถูกกรอง
        total_students = 0
        for subject in subjects_query:
            if class_filter:
                total_students += subject.students.filter(student_class=class_filter).count()
            else:
                total_students += subject.students.count()
        
        # สถิติการเข้าสอบ - รวมทุกสถานะ
        attendance_query = Attendance.objects.filter(subject__in=subjects_query)
        if class_filter:
            attendance_query = attendance_query.filter(student__student_class=class_filter)
        
        # นับตามสถานะแต่ละประเภท
        on_time_count = attendance_query.filter(status='on_time').count()
        late_count = attendance_query.filter(status='late').count()
        absent_count = attendance_query.filter(status='absent').count()
        excused_count = attendance_query.filter(status='excused').count()
        cheating_count = attendance_query.filter(status='cheating').count()
        
        # คำนวณเข้าสอบรวม (ตรงเวลา + สาย)
        attended_count = on_time_count + late_count
        
        # คำนวณอัตราเข้าสอบ
        total_attendance_records = attendance_query.count()
        attendance_rate = (attended_count / total_attendance_records * 100) if total_attendance_records > 0 else 0
        
        statistics = {
            'subjects': total_subjects,
            'students': total_students,
            'on_time': on_time_count,
            'late': late_count,
            'absent': absent_count,
            'excused': excused_count,
            'cheating': cheating_count,
            'attended': attended_count,  # รวมตรงเวลา + สาย
            'rate': round(attendance_rate, 1)
        }
        
        # สร้างข้อมูลกราฟตามประเภท
        chart_data = None
        
        if data_type == 'year':
            # กราฟตามปีการศึกษา
            years = ExamSubject.objects.values_list('academic_year', flat=True).distinct().order_by('academic_year')
            year_data = []
            year_labels = []
            
            for y in years:
                year_subjects = ExamSubject.objects.filter(academic_year=y)
                if term:
                    year_subjects = year_subjects.filter(term=term)
                if class_filter:
                    year_subjects = year_subjects.filter(students__student_class=class_filter).distinct()
                
                count = year_subjects.count()
                year_data.append(count)
                year_labels.append(f'ปี {y}')
            
            chart_data = {
                'labels': year_labels,
                'datasets': [{
                    'label': 'จำนวนการสอบ',
                    'data': year_data,
                    'backgroundColor': [
                        'rgba(59, 130, 246, 0.8)',
                        'rgba(34, 197, 94, 0.8)', 
                        'rgba(147, 51, 234, 0.8)', 
                        'rgba(234, 179, 8, 0.8)',
                        'rgba(239, 68, 68, 0.8)'
                    ][:len(year_data)],
                    'borderColor': 'rgba(59, 130, 246, 0.8)',
                    'borderWidth': 2,
                    'fill': False
                }]
            }
            
        elif data_type == 'term':
            # กราฟตามเทอม
            terms = ['1', '2', '3']
            term_data = []
            term_labels = []
            
            for t in terms:
                term_subjects = ExamSubject.objects.filter(term=t)
                if year:
                    term_subjects = term_subjects.filter(academic_year=year)
                if class_filter:
                    term_subjects = term_subjects.filter(students__student_class=class_filter).distinct()
                
                count = term_subjects.count()
                term_data.append(count)
                term_labels.append(f'เทอม {t}')
            
            chart_data = {
                'labels': term_labels,
                'datasets': [{
                    'label': 'จำนวนการสอบ',
                    'data': term_data,
                    'backgroundColor': [
                        'rgba(34, 197, 94, 0.8)', 
                        'rgba(59, 130, 246, 0.8)', 
                        'rgba(147, 51, 234, 0.8)'
                    ],
                    'borderColor': 'rgba(34, 197, 94, 0.8)',
                    'borderWidth': 2,
                    'fill': False
                }]
            }
            
        elif data_type == 'class':
            # กราฟตามระดับชั้น
            if class_filter:
                classes = [class_filter]
            else:
                classes = StudentProfile.objects.values_list('student_class', flat=True).distinct()
                
            class_data = []
            class_labels = []
            
            for cls in classes:
                if cls:
                    class_subjects_query = subjects_query.filter(students__student_class=cls).distinct()
                    
                    # หาการเข้าสอบของชั้นนี้
                    class_attendance = Attendance.objects.filter(
                        student__student_class=cls,
                        subject__in=class_subjects_query,
                        status__in=['on_time', 'late', 'excused']
                    ).count()
                    
                    # หาจำนวนการสอบทั้งหมดที่ชั้นนี้ควรเข้า
                    students_in_class = StudentProfile.objects.filter(student_class=cls).count()
                    total_possible_attendance = students_in_class * class_subjects_query.count()
                    
                    rate = (class_attendance / total_possible_attendance * 100) if total_possible_attendance > 0 else 0
                    class_data.append(round(rate, 1))
                    class_labels.append(cls)
            
            chart_data = {
                'labels': class_labels,
                'datasets': [{
                    'label': 'อัตราเข้าสอบ (%)',
                    'data': class_data,
                    'backgroundColor': [
                        'rgba(59, 130, 246, 0.8)', 
                        'rgba(147, 51, 234, 0.8)', 
                        'rgba(34, 197, 94, 0.8)', 
                        'rgba(234, 179, 8, 0.8)', 
                        'rgba(249, 115, 22, 0.8)', 
                        'rgba(236, 72, 153, 0.8)'
                    ][:len(class_data)],
                    'borderColor': 'rgba(59, 130, 246, 0.8)',
                    'borderWidth': 2,
                    'fill': False
                }]
            }
            
        else:  # default: 'all'
            # กราฟภาพรวมการเข้าสอบ - แสดงทุกสถานะ
            attendance_stats = {
                'on_time': on_time_count,
                'late': late_count,
                'absent': absent_count,
                'excused': excused_count,
                'cheating': cheating_count,
            }
            
            chart_data = {
                'labels': ['เข้าสอบตรงเวลา', 'เข้าสอบสาย', 'ขาดสอบ', 'ลาป่วย', 'ทุจริต'],
                'datasets': [{
                    'label': 'จำนวนนักเรียน',
                    'data': list(attendance_stats.values()),
                    'backgroundColor': [
                        'rgba(34, 197, 94, 0.8)',   # เขียว - ตรงเวลา
                        'rgba(234, 179, 8, 0.8)',   # เหลือง - สาย  
                        'rgba(239, 68, 68, 0.8)',   # แดง - ขาด
                        'rgba(59, 130, 246, 0.8)',  # น้ำเงิน - ลา
                        'rgba(236, 72, 153, 0.8)'   # ชมพู - ทุจริต
                    ],
                    'borderColor': 'rgba(34, 197, 94, 0.8)',
                    'borderWidth': 2,
                    'fill': False
                }]
            }
        
        response_data = {
            'success': True,
            'statistics': statistics,
            'chartData': chart_data
        }
        
        print(f"Returning data: {response_data}")
        return JsonResponse(response_data)
        
    except Exception as e:
        print(f"Dashboard API error: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'เกิดข้อผิดพลาด: {str(e)}'
        }, status=500)
    
# Dashboard ครู แสดงสถิติรวม
@login_required
def dashboard_teacher(request):
    """Dashboard ครู"""
    if not request.user.is_teacher:
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึงส่วนนี้")
    
    try:
        teacher_profile = request.user.teacher_profile
    except:
        return render(request, 'app/error.html', {'message': 'ไม่พบข้อมูลครู'})
    
    # วิชาที่ครูคุมสอบ
    my_exams = ExamSubject.objects.filter(
        Q(invigilator=teacher_profile) | Q(secondary_invigilator=teacher_profile)
    ).order_by('exam_date', 'start_time')
    
    stats = {
        'total_exams': my_exams.count(),
        'upcoming_exams': my_exams.filter(exam_date__gte=timezone.now().date()).count(),
        'students_total': StudentProfile.objects.filter(exam_subjects__in=my_exams).distinct().count(),
    }
    
    return render(request, 'app/teacher/dashboard_teacher.html', {
        'stats': stats,
        'my_exams': my_exams[:5]
    })

@login_required
def invigilated_subjects(request):
    """หน้ารายการวิชาที่ครูคุมสอบ - พร้อมสถิติ"""
    if not request.user.is_teacher:
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึงส่วนนี้")
    
    teacher = request.user.teacher_profile
    now = timezone.now()
    
    # Base queryset - วิชาที่ครูคุมสอบ (ทั้งหลักและสำรอง)
    subjects = ExamSubject.objects.filter(
        Q(invigilator=teacher) | Q(secondary_invigilator=teacher)
    ).select_related(
        'room', 'room__building', 
        'invigilator__user', 'secondary_invigilator__user'
    ).prefetch_related('students').order_by('exam_date', 'start_time')
    
    # ฟิลเตอร์
    year_filter = (request.GET.get('year') or '').strip()
    term_filter = (request.GET.get('term') or '').strip()
    
    if year_filter:
        subjects = subjects.filter(academic_year=year_filter)
    if term_filter:
        subjects = subjects.filter(term=term_filter)
    
    # ตัวเลือกสำหรับดรอปดาวน์
    years = ExamSubject.objects.filter(
        Q(invigilator=teacher) | Q(secondary_invigilator=teacher)
    ).values_list('academic_year', flat=True).distinct().order_by('-academic_year')
    
    # คำนวณสถิติ
    upcoming_count = subjects.filter(
        exam_date__gte=now.date()
    ).count()
    
    completed_count = subjects.filter(
        exam_date__lt=now.date()
    ).count()
    
    return render(request, 'app/teacher/invigilated_subjects.html', {
        'subjects': subjects,
        'years': years,
        'year_filter': year_filter,
        'term_filter': term_filter,
        'upcoming_count': upcoming_count,
        'completed_count': completed_count,
    })

@login_required  
def teacher_dashboard_data_api(request):
    """API สำหรับดึงข้อมูลแดชบอร์ดครู - รองรับทุกสถานะ"""
    if not request.user.is_teacher:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    try:
        teacher_profile = request.user.teacher_profile
        year = request.GET.get('year', '')
        term = request.GET.get('term', '')
        subject_id = request.GET.get('subject', '')
        
        print(f"Teacher Dashboard API called with: year={year}, term={term}, subject={subject_id}")
        
        # สร้าง queryset สำหรับ subjects ที่ครูคนนี้คุมสอบ
        subjects_query = ExamSubject.objects.filter(
            models.Q(invigilator=teacher_profile) | 
            models.Q(secondary_invigilator=teacher_profile)
        )
        
        # กรองตามเงื่อนไข
        if year:
            subjects_query = subjects_query.filter(academic_year=year)
        if term:
            subjects_query = subjects_query.filter(term=term)
        if subject_id:
            subjects_query = subjects_query.filter(id=subject_id)
        
        # สถิติพื้นฐาน
        total_subjects = subjects_query.count()
        
        # นับจำนวนนักเรียนทั้งหมด
        total_students = 0
        for subject in subjects_query:
            total_students += subject.students.count()
        
        # สถิติการเข้าสอบ - รวมทุกสถานะ
        attendance_query = Attendance.objects.filter(subject__in=subjects_query)
        
        # นับตามสถานะแต่ละประเภท
        on_time_count = attendance_query.filter(status='on_time').count()
        late_count = attendance_query.filter(status='late').count()
        absent_count = attendance_query.filter(status='absent').count()
        excused_count = attendance_query.filter(status='excused').count()
        cheating_count = attendance_query.filter(status='cheating').count()
        
        # คำนวณเข้าสอบรวม (ตรงเวลา + สาย)
        attended_count = on_time_count + late_count
        
        # คำนวณอัตราเข้าสอบ
        total_attendance_records = attendance_query.count()
        attendance_rate = (attended_count / total_attendance_records * 100) if total_attendance_records > 0 else 0
        
        statistics = {
            'subjects': total_subjects,
            'students': total_students,
            'on_time': on_time_count,
            'late': late_count,
            'absent': absent_count,
            'excused': excused_count,
            'cheating': cheating_count,
            'attended': attended_count,
            'rate': round(attendance_rate, 1)
        }
        
        # สร้างข้อมูลกราฟ - แสดงทุกสถานะ
        attendance_stats = {
            'on_time': on_time_count,
            'late': late_count,
            'absent': absent_count,
            'excused': excused_count,
            'cheating': cheating_count,
        }
        
        chart_data = {
            'labels': ['เข้าสอบตรงเวลา', 'เข้าสอบสาย', 'ขาดสอบ', 'ลาป่วย', 'ทุจริต'],
            'datasets': [{
                'label': 'จำนวนนักเรียน',
                'data': list(attendance_stats.values()),
                'backgroundColor': [
                    'rgba(34, 197, 94, 0.8)',   # เขียว - ตรงเวลา
                    'rgba(234, 179, 8, 0.8)',   # เหลือง - สาย  
                    'rgba(239, 68, 68, 0.8)',   # แดง - ขาด
                    'rgba(59, 130, 246, 0.8)',  # น้ำเงิน - ลา
                    'rgba(236, 72, 153, 0.8)'   # ชมพู - ทุจริต
                ],
                'borderColor': 'rgba(34, 197, 94, 0.8)',
                'borderWidth': 2,
                'fill': False
            }]
        }
        
        response_data = {
            'success': True,
            'statistics': statistics,
            'chartData': chart_data
        }
        
        print(f"Returning teacher data: {response_data}")
        return JsonResponse(response_data)
        
    except Exception as e:
        print(f"Teacher Dashboard API error: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'เกิดข้อผิดพลาด: {str(e)}'
        }, status=500)

@login_required
def get_teacher_subjects_api(request):
    """API สำหรับดึงรายวิชาที่ครูคุมสอบ"""
    if not request.user.is_teacher:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    try:
        teacher_profile = request.user.teacher_profile
        
        # ดึงวิชาที่ครูคุมสอบ (หลักหรือสำรอง)
        subjects = ExamSubject.objects.filter(
            models.Q(invigilator=teacher_profile) | 
            models.Q(secondary_invigilator=teacher_profile)
        ).distinct().order_by('subject_name')
        
        subject_list = []
        for subject in subjects:
            subject_list.append({
                'id': subject.id,
                'name': f"{subject.subject_name} ({subject.subject_code})",
                'code': subject.subject_code,
                'year': subject.academic_year,
                'term': subject.term
            })
        
        return JsonResponse({
            'success': True,
            'subjects': subject_list
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}'
        }, status=500)
    
# Dashboard นักเรียน
@login_required
def dashboard_student(request):
    """Dashboard นักเรียน"""
    if not request.user.is_student:
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึงส่วนนี้")
    
    try:
        student_profile = request.user.student_profile
    except:
        return render(request, 'app/error.html', {'message': 'ไม่พบข้อมูลนักเรียน'})
    
    # วิชาสอบของนักเรียน
    my_exams = ExamSubject.objects.filter(students=student_profile).order_by('exam_date', 'start_time')
    
    # การเข้าสอบ
    my_attendance = Attendance.objects.filter(student=student_profile).order_by('-checkin_time')
    
    return render(request, 'app/student/dashboard_student.html', {
        'my_exams': my_exams,
        'my_attendance': my_attendance[:5]
    })

@login_required
def student_exam_schedule(request):
    """ตารางสอบของนักเรียน"""
    if not request.user.is_student:
        return HttpResponseForbidden()
    
    student = request.user.student_profile
    exams = ExamSubject.objects.filter(students=student).order_by('exam_date', 'start_time')
    
    return render(request, 'app/student/exam_schedule.html', {'exams': exams})

@login_required
def student_exam_history(request):
    """ประวัติการสอบของนักเรียน"""
    if not request.user.is_student:
        return HttpResponseForbidden()
    
    student = request.user.student_profile
    attendance = Attendance.objects.filter(student=student).order_by('-checkin_time')
    
    return render(request, 'app/student/exam_history.html', {'attendance': attendance})

# ========================= จัดการผู้ใช้ (Admin) =========================

@staff_member_required
def manage_users(request):
    """จัดการผู้ใช้งาน"""
    users = User.objects.all().order_by('-date_joined')
    return render(request, 'app/admin/manage_users.html', {'users': users})

@login_required
@require_http_methods(["GET"])
def user_detail_api(request, user_type, user_id):
    """API สำหรับดึงรายละเอียดผู้ใช้"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    try:
        if user_type == 'teacher':
            profile = TeacherProfile.objects.select_related('user').get(id=user_id)
            user_data = {
                'id': profile.id,
                'teacher_id': profile.teacher_id,
                'username': profile.user.username,
                'email': profile.user.email,
                'first_name': profile.user.first_name,
                'last_name': profile.user.last_name,
                'full_name': profile.user.get_full_name(),
                'is_active': profile.user.is_active,
                'date_joined': profile.user.date_joined.strftime('%d/%m/%Y %H:%M'),
                'last_login': profile.user.last_login.strftime('%d/%m/%Y %H:%M') if profile.user.last_login else None,
            }
        elif user_type == 'student':
            profile = StudentProfile.objects.select_related('user').get(id=user_id)
            user_data = {
                'id': profile.id,
                'student_id': profile.student_id,
                'username': profile.user.username,
                'email': profile.user.email,
                'first_name': profile.user.first_name,
                'last_name': profile.user.last_name,
                'full_name': profile.user.get_full_name(),
                'student_class': profile.student_class,
                'student_number': profile.student_number,
                'is_active': profile.user.is_active,
                'date_joined': profile.user.date_joined.strftime('%d/%m/%Y %H:%M'),
                'last_login': profile.user.last_login.strftime('%d/%m/%Y %H:%M') if profile.user.last_login else None,
            }
        else:
            return JsonResponse({'success': False, 'error': 'ประเภทผู้ใช้ไม่ถูกต้อง'}, status=400)
        
        return JsonResponse({'success': True, 'user': user_data})
    
    except ObjectDoesNotExist:
        return JsonResponse({'success': False, 'error': 'ไม่พบข้อมูลผู้ใช้'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required 
@require_http_methods(["POST"])
@csrf_exempt
def update_user_api(request, user_type, user_id):
    """API สำหรับอัปเดตข้อมูลผู้ใช้"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    try:
        # Parse JSON data
        data = json.loads(request.body.decode('utf-8'))
        
        with transaction.atomic():
            if user_type == 'teacher':
                profile = TeacherProfile.objects.select_related('user').get(id=user_id)
                user = profile.user
                
                # อัปเดตข้อมูล User
                user.first_name = data.get('first_name', user.first_name)
                user.last_name = data.get('last_name', user.last_name)
                user.username = data.get('username', user.username)
                user.email = data.get('email', user.email)
                user.is_active = data.get('is_active', 'true').lower() == 'true'
                
                # เปลี่ยนรหัสผ่านถ้ามีการกำหนด
                new_password = data.get('new_password')
                if new_password and new_password.strip():
                    user.set_password(new_password)
                    # อัปเดต session เพื่อไม่ให้ logout
                    update_session_auth_hash(request, user)
                
                user.save()
                
            elif user_type == 'student':
                profile = StudentProfile.objects.select_related('user').get(id=user_id)
                user = profile.user
                
                # อัปเดตข้อมูล User
                user.first_name = data.get('first_name', user.first_name)
                user.last_name = data.get('last_name', user.last_name)
                user.username = data.get('username', user.username)
                user.email = data.get('email', user.email)
                user.is_active = data.get('is_active', 'true').lower() == 'true'
                
                # เปลี่ยนรหัสผ่านถ้ามีการกำหนด
                new_password = data.get('new_password')
                if new_password and new_password.strip():
                    user.set_password(new_password)
                    update_session_auth_hash(request, user)
                
                user.save()
                
                # อัปเดตข้อมูล StudentProfile
                profile.student_class = data.get('student_class', profile.student_class)
                try:
                    profile.student_number = int(data.get('student_number', profile.student_number))
                except (ValueError, TypeError):
                    profile.student_number = profile.student_number
                
                profile.save()
                
            else:
                return JsonResponse({'success': False, 'error': 'ประเภทผู้ใช้ไม่ถูกต้อง'}, status=400)
        
        return JsonResponse({'success': True, 'message': 'อัปเดตข้อมูลเรียบร้อยแล้ว'})
    
    except ObjectDoesNotExist:
        return JsonResponse({'success': False, 'error': 'ไม่พบข้อมูลผู้ใช้'}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'ข้อมูล JSON ไม่ถูกต้อง'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'เกิดข้อผิดพลาด: {str(e)}'}, status=500)


@login_required
def user_list(request):
    """รายชื่อครูและนักเรียนทั้งหมด"""
    if not request.user.is_staff:
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึงส่วนนี้")
    
    # รับพารามิเตอร์การค้นหา
    teacher_search = request.GET.get('teacher_search', '')
    teacher_active_filter = request.GET.get('teacher_active_filter', '')
    student_search = request.GET.get('student_search', '')
    student_class_filter = request.GET.get('student_class_filter', '')
    student_active_filter = request.GET.get('student_active_filter', '')
    
    # สถิติครู (ใช้ int() เพื่อให้แน่ใจว่าไม่มีจุดทศนิยม)
    teacher_stats = {
        'total': int(TeacherProfile.objects.count()),
        'active': int(TeacherProfile.objects.filter(user__is_active=True).count()),
        'inactive': int(TeacherProfile.objects.filter(user__is_active=False).count()),
    }
    
    # สถิตินักเรียน (ใช้ int() เพื่อให้แน่ใจว่าไม่มีจุดทศนิยม)
    student_stats = {
        'total': int(StudentProfile.objects.count()),
        'active': int(StudentProfile.objects.filter(user__is_active=True).count()),
        'inactive': int(StudentProfile.objects.filter(user__is_active=False).count()),
        'classes': int(StudentProfile.objects.values('student_class').distinct().count()),
    }
    
    # รายการระดับชั้น
    classes = StudentProfile.objects.values_list('student_class', flat=True).distinct().order_by('student_class')
    
    # ครูทั้งหมด พร้อมการค้นหาและกรอง
    teachers = TeacherProfile.objects.select_related('user').all()
    
    if teacher_search:
        teachers = teachers.filter(
            Q(teacher_id__icontains=teacher_search) |
            Q(user__first_name__icontains=teacher_search) |
            Q(user__last_name__icontains=teacher_search) |
            Q(user__username__icontains=teacher_search) |
            Q(user__email__icontains=teacher_search)
        )
    
    if teacher_active_filter:
        is_active = teacher_active_filter.lower() == 'true'
        teachers = teachers.filter(user__is_active=is_active)
    
    teachers = teachers.order_by('user__first_name')
    
    # นักเรียนทั้งหมด พร้อมการค้นหาและกรอง
    students = StudentProfile.objects.select_related('user').all()
    
    if student_search:
        students = students.filter(
            Q(student_id__icontains=student_search) |
            Q(user__first_name__icontains=student_search) |
            Q(user__last_name__icontains=student_search) |
            Q(user__username__icontains=student_search) |
            Q(user__email__icontains=student_search) |
            Q(student_class__icontains=student_search)
        )
    
    if student_class_filter:
        students = students.filter(student_class=student_class_filter)
    
    if student_active_filter:
        is_active = student_active_filter.lower() == 'true'
        students = students.filter(user__is_active=is_active)
    
    # Pagination สำหรับนักเรียน
    from django.core.paginator import Paginator
    student_paginator = Paginator(students.order_by('student_class', 'student_number', 'user__first_name'), 50)
    page_number = request.GET.get('page', 1)
    students_page = student_paginator.get_page(page_number)
    
    # ตรวจสอบว่าเป็น AJAX request หรือไม่
    if request.GET.get('ajax') == '1':
        # ส่งข้อมูล JSON สำหรับ AJAX
        students_data = []
        for student in students_page:
            students_data.append({
                'id': student.id,
                'student_id': student.student_id,
                'username': student.user.username,
                'full_name': student.user.get_full_name(),
                'email': student.user.email,
                'student_class': student.student_class,
                'student_number': int(student.student_number),  # ใช้ค่าจริงจากระบบ
                'is_active': student.user.is_active,
                'date_joined': student.user.date_joined.strftime('%d/%m/%Y')
            })
        
        # คำนวณสถิติใหม่หลังจากกรอง (ใช้ int() เพื่อให้แน่ใจว่าไม่มีจุดทศนิยม)
        filtered_student_stats = {
            'total': int(students.count()),
            'active': int(students.filter(user__is_active=True).count()),
            'inactive': int(students.filter(user__is_active=False).count()),
            'classes': int(students.values('student_class').distinct().count()),
        }
        
        return JsonResponse({
            'students': students_data,
            'stats': filtered_student_stats,
            'success': True
        })
    
    return render(request, 'app/staff/user_list.html', {
        'teacher_stats': teacher_stats,
        'student_stats': student_stats,
        'classes': classes,
        'teachers': teachers,
        'students': students_page,
        'teacher_search': teacher_search,
        'teacher_active_filter': teacher_active_filter,
        'student_search': student_search,
        'student_class_filter': student_class_filter,
        'student_active_filter': student_active_filter,
    })
    
# ========================= เพิ่มฟังก์ชันใหม่สำหรับดู user detail =========================

# ใช้สำหรับ: ดึงรายละเอียดผู้ใช้รายบุคคลผ่าน AJAX (teacher/student) เฉพาะ staff
@login_required
def ajax_user_detail(request, user_type, user_id):
    """AJAX endpoint สำหรับดูรายละเอียดผู้ใช้"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    try:
        if user_type == 'teacher':
            teacher = get_object_or_404(TeacherProfile, id=user_id)
            data = {
                'teacher_id': teacher.teacher_id,
                'username': teacher.user.username,
                'full_name': teacher.user.get_full_name(),
                'email': teacher.user.email,
                'is_active': teacher.user.is_active,
                'date_joined': teacher.user.date_joined.strftime('%d/%m/%Y'),
                'success': True
            }
        else:  # student
            student = get_object_or_404(StudentProfile, id=user_id)
            data = {
                'student_id': student.student_id,
                'username': student.user.username,
                'full_name': student.user.get_full_name(),
                'email': student.user.email,
                'student_class': student.student_class,
                'student_number': str(student.student_number),
                'is_active': student.user.is_active,
                'date_joined': student.user.date_joined.strftime('%d/%m/%Y'),
                'success': True
            }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

# ========================= ฟังก์ชัน Import ข้อมูล =========================

# หน้า import ข้อมูลนักเรียน + แสดงผลสำเร็จ/ผิดพลาด
@staff_member_required
def import_students(request):
    """หน้าสำหรับ import ข้อมูลนักเรียน"""
    if request.method == 'POST':
        form = StudentImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # ประมวลผลไฟล์
                student_data = form.process_file()
                overwrite = form.cleaned_data['overwrite_existing']
                
                # นำเข้าข้อมูล
                result = process_student_import(student_data, overwrite)
                
                # แสดงผลลัพธ์
                if result['success_count'] > 0:
                    messages.success(request, f"นำเข้าข้อมูลนักเรียนสำเร็จ {result['success_count']} คน")
                
                if result['error_count'] > 0:
                    messages.warning(request, f"มีข้อผิดพลาด {result['error_count']} รายการ")
                
                if result['updated_count'] > 0:
                    messages.info(request, f"อัปเดตข้อมูลแล้ว {result['updated_count']} คน")
                
                # แสดงรายละเอียดข้อผิดพลาด
                if result['errors']:
                    error_details = "\n".join([f"แถว {err['row']}: {err['message']}" for err in result['errors'][:5]])
                    messages.error(request, f"ตัวอย่างข้อผิดพลาด:\n{error_details}")
                
                return redirect('import_students')
                
            except Exception as e:
                messages.error(request, f"เกิดข้อผิดพลาดในการประมวลผลไฟล์: {str(e)}")
    else:
        form = StudentImportForm()
    
    return render(request, 'app/staff/import_students.html', {
        'form': form,
        'template_url': '/static/templates/student_template.xlsx'
    })

# หน้า import ข้อมูลครู + แสดงผลสำเร็จ/ผิดพลาด
@staff_member_required  
def import_teachers(request):
    """Delegates to import_students to avoid duplicate logic."""
    return import_students(request)

# วนบันทึกนักเรียนจากข้อมูลที่ parse แล้ว (รองรับ overwrite)
def process_student_import(student_data, overwrite=False): 
    """ประมวลผลการ import ข้อมูลนักเรียน"""
    result = {
        'success_count': 0,
        'error_count': 0,
        'updated_count': 0,
        'errors': []
    }
    
    with transaction.atomic():
        for index, data in enumerate(student_data, start=2):  # เริ่มแถว 2 (หัวตารางแถว 1)
            try:
                # ตรวจสอบข้อมูลจำเป็น
                required_fields = ['username', 'password', 'student_id', 'first_name', 'last_name', 'email', 'student_class', 'student_number']
                missing_fields = [field for field in required_fields if not data.get(field)]
                
                if missing_fields:
                    result['errors'].append({
                        'row': index,
                        'message': f"ข้อมูลไม่ครบ: {', '.join(missing_fields)}"
                    })
                    result['error_count'] += 1
                    continue
                
                # ตรวจสอบอีเมลที่ซ้ำ
                if User.objects.filter(email=data['email']).exists():
                    existing_user = User.objects.get(email=data['email'])
                    
                    if overwrite and hasattr(existing_user, 'student_profile'):
                        # อัปเดตข้อมูลที่มีอยู่
                        update_existing_student(existing_user, data)
                        result['updated_count'] += 1
                        continue
                    else:
                        result['errors'].append({
                            'row': index,
                            'message': f"อีเมล {data['email']} มีอยู่ในระบบแล้ว"
                        })
                        result['error_count'] += 1
                        continue
                
                # ตรวจสอบรหัสนักเรียนที่ซ้ำ
                if StudentProfile.objects.filter(student_id=data['student_id']).exists():
                    if not overwrite:
                        result['errors'].append({
                            'row': index,
                            'message': f"รหัสนักเรียน {data['student_id']} มีอยู่ในระบบแล้ว"
                        })
                        result['error_count'] += 1
                        continue
                
                # สร้างผู้ใช้ใหม่
                user = User.objects.create_user(
                    username=data['username'],
                    email=data['email'],
                    first_name=data['first_name'],
                    last_name=data['last_name'],
                    password=data['password'],
                    is_student=True,
                    is_active=True
                )
                
                # สร้างโปรไฟล์นักเรียน
                StudentProfile.objects.create(
                    user=user,
                    student_id=data['student_id'],
                    student_number=data['student_number'],
                    student_class=data['student_class']
                )
                
                result['success_count'] += 1
                
            except IntegrityError as e:
                result['errors'].append({
                    'row': index,
                    'message': f"ข้อมูลซ้ำ: {str(e)}"
                })
                result['error_count'] += 1
                
            except Exception as e:
                result['errors'].append({
                    'row': index,
                    'message': f"ข้อผิดพลาด: {str(e)}"
                })
                result['error_count'] += 1
    
    return result

# วนบันทึกครูจากข้อมูลที่ parse แล้ว (รองรับ overwrite)
def process_teacher_import(teacher_data, overwrite=False):
    """ประมวลผลการ import ข้อมูลครู"""
    result = {
        'success_count': 0,
        'error_count': 0,
        'updated_count': 0,
        'errors': []
    }
    
    with transaction.atomic():
        for index, data in enumerate(teacher_data, start=2):
            try:
                # ตรวจสอบข้อมูลจำเป็น
                required_fields = ['username', 'password', 'teacher_id', 'first_name', 'last_name', 'email']
                missing_fields = [field for field in required_fields if not data.get(field)]
                
                if missing_fields:
                    result['errors'].append({
                        'row': index,
                        'message': f"ข้อมูลไม่ครบ: {', '.join(missing_fields)}"
                    })
                    result['error_count'] += 1
                    continue
                
                # ตรวจสอบอีเมลที่ซ้ำ
                if User.objects.filter(email=data['email']).exists():
                    existing_user = User.objects.get(email=data['email'])
                    
                    if overwrite and hasattr(existing_user, 'teacher_profile'):
                        # อัปเดตข้อมูลที่มีอยู่
                        update_existing_teacher(existing_user, data)
                        result['updated_count'] += 1
                        continue
                    else:
                        result['errors'].append({
                            'row': index,
                            'message': f"อีเมล {data['email']} มีอยู่ในระบบแล้ว"
                        })
                        result['error_count'] += 1
                        continue
                
                # ตรวจสอบรหัสครูที่ซ้ำ
                if TeacherProfile.objects.filter(teacher_id=data['teacher_id']).exists():
                    if not overwrite:
                        result['errors'].append({
                            'row': index,
                            'message': f"รหัสครู {data['teacher_id']} มีอยู่ในระบบแล้ว"
                        })
                        result['error_count'] += 1
                        continue
                
                # สร้างผู้ใช้ใหม่
                user = User.objects.create_user(
                    username=data['username'],
                    email=data['email'],
                    first_name=data['first_name'],
                    last_name=data['last_name'],
                    password=data['password'],
                    is_teacher=True,
                    is_active=True
                )
                
                # สร้างโปรไฟล์ครู
                TeacherProfile.objects.create(
                    user=user,
                    teacher_id=data['teacher_id']
                )
                
                result['success_count'] += 1
                
            except IntegrityError as e:
                result['errors'].append({
                    'row': index,
                    'message': f"ข้อมูลซ้ำ: {str(e)}"
                })
                result['error_count'] += 1
                
            except Exception as e:
                result['errors'].append({
                    'row': index,
                    'message': f"ข้อผิดพลาด: {str(e)}"
                })
                result['error_count'] += 1
    
    return result

#อัปเดตข้อมูลผู้ใช้/โปรไฟล์นักเรียน
def update_existing_student(user, data):
    """อัปเดตข้อมูลนักเรียนที่มีอยู่แล้ว"""
    # อัปเดตข้อมูลผู้ใช้
    user.username = data['username']
    user.first_name = data['first_name']
    user.last_name = data['last_name']
    user.email = data['email']
    user.set_password(data['password'])
    user.save()
    
    # อัปเดตโปรไฟล์นักเรียน
    profile = user.student_profile
    profile.student_id = data['student_id']
    profile.student_number = data['student_number']
    profile.student_class = data['student_class']
    profile.save()

#อัปเดตข้อมูลผู้ใช้/โปรไฟล์ครูเดิม
def update_existing_teacher(user, data):
    """อัปเดตข้อมูลครูที่มีอยู่แล้ว"""
    # อัปเดตข้อมูลผู้ใช้
    user.username = data['username']
    user.first_name = data['first_name']
    user.last_name = data['last_name']
    user.email = data['email']
    user.set_password(data['password'])
    user.save()
    
    # อัปเดตโปรไฟล์ครู
    profile = user.teacher_profile
    profile.teacher_id = data['teacher_id']
    profile.save()

# สร้างไฟล์ Excel เทมเพลต (student/teacher) พร้อมหัวตาราง/ตัวอย่าง/คำแนะนำ
@staff_member_required
def download_template(request, template_type):
    """ดาวน์โหลดไฟล์ template สำหรับ import"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    if template_type == 'student':
        ws.title = "Student Template"
        headers = [
            'username', 'password', 'student_id', 'first_name', 'last_name', 'email', 
            'student_class', 'student_number'
        ]
        sample_data = [
            ['student001', 'password123', 'STD001', 'สมชาย', 'ใจดี', 'somchai@email.com', 'ม.1/1', '1'],
            ['student002', 'password456', 'STD002', 'สมหญิง', 'รักเรียน', 'somying@email.com', 'ม.1/1', '2'],
            ['student003', 'password789', 'STD003', 'วิชัย', 'เก่งมาก', 'wichai@email.com', 'ม.1/2', '1']
        ]
        filename = 'student_import_template.xlsx'
    
    elif template_type == 'teacher':
        ws.title = "Teacher Template"
        headers = [
            'username', 'password', 'teacher_id', 'first_name', 'last_name', 'email'
        ]
        sample_data = [
            ['teacher001', 'teacher123', 'TCH001', 'อาจารย์สมชาย', 'สอนดี', 'teacher1@email.com'],
            ['teacher002', 'teacher456', 'TCH002', 'อาจารย์สมหญิง', 'รู้มาก', 'teacher2@email.com'],
            ['teacher003', 'teacher789', 'TCH003', 'อาจารย์วิชัย', 'เก่งมาก', 'teacher3@email.com']
        ]
        filename = 'teacher_import_template.xlsx'
    
    else:
        return HttpResponse("Template ไม่ถูกต้อง", status=400)
    
    # สร้างหัวตาราง
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # เพิ่มข้อมูลตัวอย่าง
    for row, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # ปรับขนาดคอลัมน์
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # เพิ่มคำแนะนำ
    instruction_row = len(sample_data) + 4
    ws.cell(row=instruction_row, column=1, value="คำแนะนำการใช้งาน:")
    ws.cell(row=instruction_row, column=1).font = Font(bold=True)
    
    instructions = [
        "1. กรอกข้อมูลตามคอลัมน์ที่กำหนด",
        "2. ไม่ควรลบหรือแก้ไขชื่อคอลัมน์",
        "3. ข้อมูลทุกคอลัมน์เป็นข้อมูลจำเป็น",
        "4. อีเมลต้องไม่ซ้ำกัน",
        f"5. {'รหัสนักเรียน' if template_type == 'student' else 'รหัสครู'}ต้องไม่ซ้ำกัน",
        "6. Username ต้องไม่ซ้ำกัน"
    ]
    
    for i, instruction in enumerate(instructions):
        ws.cell(row=instruction_row + i + 1, column=1, value=instruction)
    
    # สร้าง response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

#สร้างไฟล์ Excel ส่งออกข้อมูลผู้ใช้ (student/teacher)
@staff_member_required
def export_users(request, user_type):
    """ส่งออกข้อมูลผู้ใช้เป็น Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    if user_type == 'student':
        ws.title = "Students Data"
        headers = [
            'Username', 'รหัสนักเรียน', 'ชื่อ', 'นามสกุล', 'อีเมล', 
            'ระดับชั้น', 'เลขที่', 'สถานะ', 'วันที่สมัคร'
        ]
        
        students = StudentProfile.objects.select_related('user').all().order_by('student_class', 'student_number')
        data = []
        
        for student in students:
            data.append([
                student.user.username,
                student.student_id,
                student.user.first_name,
                student.user.last_name,
                student.user.email,
                student.student_class,
                student.student_number,
                'ใช้งาน' if student.user.is_active else 'ไม่ใช้งาน',
                student.user.date_joined.strftime('%d/%m/%Y')
            ])
        
        filename = f'students_export_{timezone.now().strftime("%Y%m%d")}.xlsx'
    
    elif user_type == 'teacher':
        ws.title = "Teachers Data"
        headers = [
            'Username', 'รหัสครู', 'ชื่อ', 'นามสกุล', 'อีเมล', 
            'สถานะ', 'วันที่สมัคร'
        ]
        
        teachers = TeacherProfile.objects.select_related('user').all().order_by('user__first_name')
        data = []
        
        for teacher in teachers:
            data.append([
                teacher.user.username,
                teacher.teacher_id,
                teacher.user.first_name,
                teacher.user.last_name,
                teacher.user.email,
                'ใช้งาน' if teacher.user.is_active else 'ไม่ใช้งาน',
                teacher.user.date_joined.strftime('%d/%m/%Y')
            ])
        
        filename = f'teachers_export_{timezone.now().strftime("%Y%m%d")}.xlsx'
    
    else:
        return HttpResponse("ประเภทผู้ใช้ไม่ถูกต้อง", status=400)
    
    # สร้างหัวตาราง
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # เพิ่มข้อมูล
    for row, record in enumerate(data, 2):
        for col, value in enumerate(record, 1):
            ws.cell(row=row, column=col, value=value)
    
    # ปรับขนาดคอลัมน์
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # สร้าง response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

# ========================= จัดการรายวิชาสอบ =========================

# แสดงรายวิชาสอบ + ฟิลเตอร์ class/year + รวมจำนวนนักเรียน
@login_required
def exam_subjects(request):
    """หน้ารายการวิชาสอบ พร้อมฟิลเตอร์ class/year/term และสรุปสถิติ"""
    if not request.user.is_staff:
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึงส่วนนี้")

    # Base queryset
    subjects = (
        ExamSubject.objects
        .select_related('room', 'room__building', 'invigilator__user', 'secondary_invigilator__user')
        .prefetch_related('students')
        .order_by('exam_date', 'start_time')
    )

    # ฟิลเตอร์
    class_filter = (request.GET.get('class') or '').strip()
    year_filter  = (request.GET.get('year')  or '').strip()
    term_filter  = (request.GET.get('term')  or '').strip()

    if class_filter:
        subjects = subjects.filter(students__student_class=class_filter).distinct()
    if year_filter:
        subjects = subjects.filter(academic_year=year_filter)
    if term_filter:
        subjects = subjects.filter(term=term_filter)

    # ตัวเลือกสำหรับดรอปดาวน์
    classes = (
        StudentProfile.objects
        .values_list('student_class', flat=True)
        .distinct()
        .order_by('student_class')
    )
    years = (
        ExamSubject.objects
        .values_list('academic_year', flat=True)
        .distinct()
        .order_by('-academic_year')
    )
    terms = (
        ExamSubject.objects
        .values_list('term', flat=True)
        .distinct()
        .order_by('term')
    )

    # สถิติ
    total_students = sum(s.get_student_count() for s in subjects)
    total_rooms = (
        subjects.exclude(room__isnull=True)
        .values('room').distinct().count()
    )
    total_teachers = (
        subjects.exclude(invigilator__isnull=True)
        .values('invigilator').distinct().count()
    )

    return render(request, 'app/staff/exam_subjects.html', {
        'subjects': subjects,
        'classes': classes,
        'years': years,
        'terms': terms,
        'class_filter': class_filter,
        'year_filter': year_filter,
        'term_filter': term_filter,
        'total_students': total_students,
        'total_rooms': total_rooms,
        'total_teachers': total_teachers,
    })

# เพิ่มวิชาสอบ (เลือกห้อง/ครูแบบ auto/manual) + ตรวจชนเวลา/ความจุ
@login_required
def add_exam_subject(request):
    """เพิ่มรายวิชาสอบ - ปรับปรุงใหม่ให้ทำงานสมบูรณ์"""
    if not request.user.is_staff:
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึงส่วนนี้")
    
    if request.method == 'POST':
        # Debug: ดูข้อมูลที่ส่งมา
        print("POST data:", request.POST)
        
        try:
            with transaction.atomic():
                # สร้าง instance ของ ExamSubject
                subject = ExamSubject()
                
                # กำหนดข้อมูลพื้นฐาน
                subject.subject_name = request.POST.get('subject_name', '').strip()
                subject.subject_code = request.POST.get('subject_code', '').strip()
                subject.academic_year = request.POST.get('academic_year', '').strip()
                subject.term = request.POST.get('term', '')
                subject.exam_date = request.POST.get('exam_date')
                subject.start_time = request.POST.get('start_time')
                subject.end_time = request.POST.get('end_time')
                subject.created_by = request.user
                
                # ตรวจสอบข้อมูลจำเป็น
                if not all([subject.subject_name, subject.subject_code, subject.academic_year, 
                           subject.term, subject.exam_date, subject.start_time, subject.end_time]):
                    messages.error(request, 'กรุณากรอกข้อมูลให้ครบถ้วน')
                    return redirect('add_exam_subject')
                
                # ดึงข้อมูลเพิ่มเติม - แก้ไขการจัดการ checkbox
                student_class = request.POST.get('student_class', '').strip()
                
                # **แก้ไข: จัดการ room assignment type ใหม่**
                # ตรวจสอบว่าเลือก auto หรือไม่จาก checkbox
                room_auto_assign = request.POST.get('room_auto_assign')  # จะมีค่าเมื่อ checkbox ถูกเลือก
                room_type = 'auto' if room_auto_assign else 'manual'
                print(f"Debug: room_auto_assign = {room_auto_assign}, final room_type = {room_type}")
                
                # **แก้ไข: จัดการ teacher assignment type ใหม่**
                teacher_auto_assign = request.POST.get('teacher_auto_assign')  # จะมีค่าเมื่อ checkbox ถูกเลือก
                teacher_type = 'auto' if teacher_auto_assign else 'manual'
                print(f"Debug: teacher_auto_assign = {teacher_auto_assign}, final teacher_type = {teacher_type}")
                
                if not student_class:
                    messages.error(request, 'กรุณาเลือกระดับชั้น')
                    return redirect('add_exam_subject')
                
                # ดึงนักเรียนจากระดับชั้น
                students = StudentProfile.objects.filter(student_class=student_class)
                student_count = students.count()
                
                if student_count == 0:
                    messages.error(request, f'ไม่พบนักเรียนในระดับชั้น {student_class}')
                    return redirect('add_exam_subject')
                
                print(f"Debug: Found {student_count} students in class {student_class}")
                
                # จัดการห้องสอบ
                if room_type == 'auto':
                    print("Debug: Using auto room assignment")
                    # ใช้ระบบจัดห้องแบบเฉลี่ย
                    available_room = find_available_room_balanced(
                        subject.exam_date, 
                        subject.start_time, 
                        subject.end_time, 
                        student_count
                    )
                    if not available_room:
                        messages.error(request, f'ไม่มีห้องว่างสำหรับนักเรียน {student_count} คน')
                        return redirect('add_exam_subject')
                    subject.room = available_room
                    print(f"Debug: Auto assigned room: {available_room.building.name} ห้อง {available_room.name}")
                else:
                    print("Debug: Using manual room assignment")
                    # Manual room selection
                    room_id = request.POST.get('room')
                    if room_id:
                        try:
                            room = ExamRoom.objects.get(id=room_id)
                            
                            # ตรวจสอบความขัดแย้งเวลา
                            room_conflicts = ExamSubject.objects.filter(
                                room=room,
                                exam_date=subject.exam_date,
                                start_time__lt=subject.end_time,
                                end_time__gt=subject.start_time
                            )
                            
                            if room_conflicts.exists():
                                conflict_subject = room_conflicts.first()
                                messages.error(request, f'ห้อง {room.building.name} ห้อง {room.name} มีการใช้งานในช่วงเวลา {conflict_subject.start_time}-{conflict_subject.end_time} แล้ว')
                                return redirect('add_exam_subject')
                            
                            # ตรวจสอบความจุห้อง
                            if room.capacity < student_count:
                                messages.warning(request, f'ห้องจุได้ {room.capacity} คน แต่มีนักเรียน {student_count} คน (ห้องอาจแน่น)')
                            
                            subject.room = room
                            print(f"Debug: Manual assigned room: {room.building.name} ห้อง {room.name}")
                        except ExamRoom.DoesNotExist:
                            messages.error(request, 'ไม่พบห้องที่เลือก')
                            return redirect('add_exam_subject')
                    else:
                        # ถ้าไม่ได้เลือกห้องแต่ไม่ได้เปิด auto ให้แสดง error
                        messages.error(request, 'กรุณาเลือกห้องสอบ หรือเปิดใช้งานการจัดห้องอัตโนมัติ')
                        return redirect('add_exam_subject')
                
                # จัดการครูคุมสอบ
                if teacher_type == 'auto':
                    print("Debug: Using auto teacher assignment")
                    # ใช้ระบบจัดครูอัตโนมัติแบบเฉลี่ย
                    available_teachers = find_available_teachers_balanced(
                        subject.exam_date,
                        subject.start_time,
                        subject.end_time,
                        min_count=2
                    )
                    
                    if len(available_teachers) >= 1:
                        subject.invigilator = available_teachers[0]
                        print(f"Debug: Auto assigned primary teacher: {available_teachers[0].user.get_full_name()}")
                        
                        # ครูสำรอง (ถ้ามี)
                        if len(available_teachers) >= 2:
                            subject.secondary_invigilator = available_teachers[1]
                            print(f"Debug: Auto assigned secondary teacher: {available_teachers[1].user.get_full_name()}")
                    else:
                        messages.error(request, 'ไม่พบครูว่างในช่วงเวลานี้')
                        return redirect('add_exam_subject')
                        
                else:
                    print("Debug: Using manual teacher assignment")
                    # Manual teacher selection
                    invigilator_id = request.POST.get('invigilator')
                    secondary_invigilator_id = request.POST.get('secondary_invigilator')
                    
                    if invigilator_id:
                        try:
                            primary_teacher = TeacherProfile.objects.get(id=invigilator_id)
                            
                            # ตรวจสอบความขัดแย้งของครูหลัก
                            teacher_conflicts = ExamSubject.objects.filter(
                                exam_date=subject.exam_date,
                                start_time__lt=subject.end_time,
                                end_time__gt=subject.start_time
                            ).filter(
                                Q(invigilator=primary_teacher) | 
                                Q(secondary_invigilator=primary_teacher)
                            )
                            
                            if teacher_conflicts.exists():
                                conflict_subject = teacher_conflicts.first()
                                messages.error(request, f'ครู {primary_teacher.user.get_full_name()} มีตารางคุมสอบวิชา "{conflict_subject.subject_name}" ในช่วงเวลา {conflict_subject.start_time}-{conflict_subject.end_time} แล้ว')
                                return redirect('add_exam_subject')
                            
                            subject.invigilator = primary_teacher
                            print(f"Debug: Manual assigned primary teacher: {primary_teacher.user.get_full_name()}")
                        except TeacherProfile.DoesNotExist:
                            messages.error(request, 'ไม่พบครูหลักที่เลือก')
                            return redirect('add_exam_subject')
                    else:
                        # ถ้าไม่ได้เลือกครูแต่ไม่ได้เปิด auto ให้แสดง error
                        messages.error(request, 'กรุณาเลือกครูคุมสอบหลัก หรือเปิดใช้งานการจัดครูอัตโนมัติ')
                        return redirect('add_exam_subject')
                    
                    # ครูสำรอง (ไม่บังคับ)
                    if secondary_invigilator_id:
                        try:
                            secondary_teacher = TeacherProfile.objects.get(id=secondary_invigilator_id)
                            
                            # ตรวจสอบว่าไม่ใช่คนเดียวกัน
                            if secondary_teacher == subject.invigilator:
                                messages.error(request, 'ครูหลักและครูสำรองต้องเป็นคนละคน')
                                return redirect('add_exam_subject')
                            
                            # ตรวจสอบความขัดแย้งของครูสำรอง
                            teacher_conflicts = ExamSubject.objects.filter(
                                exam_date=subject.exam_date,
                                start_time__lt=subject.end_time,
                                end_time__gt=subject.start_time
                            ).filter(
                                Q(invigilator=secondary_teacher) | 
                                Q(secondary_invigilator=secondary_teacher)
                            )
                            
                            if teacher_conflicts.exists():
                                conflict_subject = teacher_conflicts.first()
                                messages.error(request, f'ครูสำรอง {secondary_teacher.user.get_full_name()} มีตารางคุมสอบวิชา "{conflict_subject.subject_name}" ในช่วงเวลา {conflict_subject.start_time}-{conflict_subject.end_time} แล้ว')
                                return redirect('add_exam_subject')
                            
                            subject.secondary_invigilator = secondary_teacher
                            print(f"Debug: Manual assigned secondary teacher: {secondary_teacher.user.get_full_name()}")
                        except TeacherProfile.DoesNotExist:
                            # ไม่ต้องแสดง error ถ้าไม่พบครูสำรอง เพราะไม่บังคับ
                            print("Debug: Secondary teacher not found, but it's optional")
                            pass
                
                # ตรวจสอบความซ้ำซ้อนของรหัสวิชา
                if ExamSubject.objects.filter(
                    subject_code=subject.subject_code,
                    academic_year=subject.academic_year,
                    term=subject.term
                ).exists():
                    messages.error(request, f'รหัสวิชา {subject.subject_code} ในปีการศึกษา {subject.academic_year} เทอม {subject.term} มีอยู่ในระบบแล้ว')
                    return redirect('add_exam_subject')
                
                # บันทึกข้อมูล
                subject.save()
                subject.students.set(students)
                
                # สร้างข้อความสำเร็จ
                success_parts = [
                    f'เพิ่มรายวิชา "{subject.subject_name}" สำเร็จ!',
                    f'นักเรียน: {student_count} คน'
                ]
                
                if subject.room:
                    utilization = (student_count / subject.room.capacity) * 100 if subject.room.capacity > 0 else 0
                    success_parts.append(f'ห้องสอบ: {subject.room.building.name} ห้อง {subject.room.name} (ใช้งาน {utilization:.0f}%)')
                    
                if subject.invigilator:
                    success_parts.append(f'ครูหลัก: {subject.invigilator.user.get_full_name()}')
                    
                if subject.secondary_invigilator:
                    success_parts.append(f'ครูสำรอง: {subject.secondary_invigilator.user.get_full_name()}')
                
                success_message = ' | '.join(success_parts)
                messages.success(request, success_message)
                
                print(f"Debug: Successfully created exam subject: {subject.subject_name}")
                return redirect('exam_subjects')
                
        except Exception as e:
            print("Error in add_exam_subject:", str(e))
            import traceback
            traceback.print_exc()
            messages.error(request, f'เกิดข้อผิดพลาด: {str(e)}')
    
    # GET request - แสดงฟอร์ม
    form = ExamSubjectForm()
    
    # ดึงข้อมูลเพิ่มเติมสำหรับฟอร์ม
    context = {
        'form': form,
        'classes': StudentProfile.objects.values_list('student_class', flat=True).distinct().order_by('student_class'),
        'teachers': TeacherProfile.objects.select_related('user').filter(user__is_active=True).order_by('user__first_name'),
        'buildings': Building.objects.all().order_by('name'),
    }
    
    return render(request, 'app/staff/add_exam_subject.html', context)

# แบบฟอร์มแก้ไขวิชาสอบ + อัปเดตชุดนักเรียนตามชั้นเรียน
@login_required
def edit_exam_subject(request, subject_id):
    subject = get_object_or_404(ExamSubject, id=subject_id)

    if request.method == 'POST':
        form = ExamSubjectForm(request.POST, instance=subject)
        if form.is_valid():
            # ตรวจเวลา
            start_time = form.cleaned_data.get('start_time')
            end_time   = form.cleaned_data.get('end_time')
            if start_time and end_time and start_time >= end_time:
                messages.error(request, 'เวลาเริ่มต้องน้อยกว่าเวลาสิ้นสุด')
                return render(request, 'app/staff/edit_exam_subject.html', {'form': form, 'subject': subject})

            # บันทึกฟิลด์ทั้งหมด (รวมเวลาที่แก้)
            subject = form.save(commit=False)

            # (ออปชัน) ป้องกันแก้ย้อนหลังเสี่ยง: ถ้าไม่ต้องการให้แก้เวลาของการสอบที่จบไปแล้ว
            # if subject.exam_date < now().date():
            #     messages.error(request, 'ไม่อนุญาตให้แก้เวลาของการสอบที่สิ้นสุดแล้ว')
            #     return render(request, 'app/staff/edit_exam_subject.html', {'form': form, 'subject': subject})

            # อัปเดตชุดนักเรียนตามชั้นเรียน (ถ้าแบบฟอร์มมี field นี้)
            student_class = form.cleaned_data.get('student_class')
            subject.save()
            if student_class is not None:
                students = StudentProfile.objects.filter(student_class=student_class)
                subject.students.set(students)

            messages.success(request, 'แก้ไขรายวิชาสำเร็จ!')
            return redirect('exam_subjects')
        else:
            messages.error(request, 'กรุณาตรวจสอบข้อมูลที่กรอก')
    else:
        form = ExamSubjectForm(instance=subject)

    return render(request, 'app/staff/edit_exam_subject.html', {'form': form, 'subject': subject})

@login_required
def delete_exam_subject(request, subject_id):
    """ลบรายวิชาสอบ"""
    if request.method == 'POST':
        subject = get_object_or_404(ExamSubject, id=subject_id)
        subject_name = subject.subject_name
        subject.delete()
        messages.success(request, f'ลบรายวิชา {subject_name} สำเร็จ!')
    return redirect('exam_subjects')

# ลบวิชาสอบแบบ AJAX (ตรวจ ongoing/attendance, อนุญาต superuser ตามเงื่อนไข)
@login_required
@csrf_exempt
def delete_exam_subject_ajax(request, subject_id):
    """ลบรายวิชาสอบผ่าน AJAX - แก้ไขให้ทำงานได้"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง', 'success': False}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed', 'success': False}, status=405)
    
    try:
        subject = get_object_or_404(ExamSubject, id=subject_id)
        
        # ตรวจสอบว่ามีการเช็คชื่อแล้วหรือไม่ 
        attendance_count = Attendance.objects.filter(subject=subject).count()
        
        # ตรวจสอบว่าเป็นการสอบที่กำลังดำเนินการหรือไม่
        now = timezone.now()
        is_ongoing = (subject.exam_date == now.date() and 
                     subject.start_time <= now.time() <= subject.end_time)
        
        # ถ้ามีการเช็คชื่อหรือกำลังสอบ ให้เตือน แต่ยังคงสามารถลบได้ถ้าเป็น superuser
        if attendance_count > 0 or is_ongoing:
            if not request.user.is_superuser:
                message = []
                if attendance_count > 0:
                    message.append(f'มีการบันทึกการเข้าสอบแล้ว ({attendance_count} รายการ)')
                if is_ongoing:
                    message.append('การสอบกำลังดำเนินการอยู่')
                
                return JsonResponse({
                    'error': f'ไม่สามารถลบได้: {", ".join(message)}. ติดต่อผู้ดูแลระบบหากจำเป็น',
                    'success': False
                }, status=400)
        
        subject_name = subject.subject_name
        
        with transaction.atomic():
            # ลบข้อมูล attendance ที่เกี่ยวข้องก่อน (ถ้ามี)
            if attendance_count > 0:
                Attendance.objects.filter(subject=subject).delete()
                print(f"Deleted {attendance_count} attendance records")
            
            # แล้วค่อยลบ subject
            subject.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'ลบรายวิชา "{subject_name}" สำเร็จ' + 
                      (f' (รวมข้อมูลการเข้าสอบ {attendance_count} รายการ)' if attendance_count > 0 else '')
        })
        
    except ExamSubject.DoesNotExist:
        return JsonResponse({
            'error': 'ไม่พบรายวิชาที่ต้องการลบ',
            'success': False
        }, status=404)
        
    except Exception as e:
        print(f"Error in delete_exam_subject_ajax: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)
  
# ลบหลายวิชาพร้อมกัน + รายงานผลสำเร็จ/เตือน/ผิดพลาด
@login_required
@csrf_exempt
def bulk_delete_exam_subjects(request):
    """ลบรายวิชาสอบหลายรายการพร้อมกัน - ปรับปรุงแล้ว"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        subject_ids = data.get('subject_ids', [])
        force_delete = data.get('force_delete', False)  # สำหรับการลบแบบบังคับ
        
        if not subject_ids:
            return JsonResponse({'error': 'ไม่ได้ระบุรายวิชาที่จะลบ'}, status=400)
        
        results = {
            'success_count': 0,
            'error_count': 0,
            'warning_count': 0,
            'errors': [],
            'warnings': []
        }
        
        now = timezone.now()
        
        with transaction.atomic():
            for subject_id in subject_ids:
                try:
                    subject = ExamSubject.objects.get(id=subject_id)
                    
                    # ตรวจสอบข้อจำกัด
                    attendance_count = Attendance.objects.filter(subject=subject).count()
                    is_ongoing = (subject.exam_date == now.date() and 
                                subject.start_time <= now.time() <= subject.end_time)
                    
                    # ถ้ามีข้อจำกัดและไม่ได้บังคับลบ
                    if (attendance_count > 0 or is_ongoing) and not force_delete and not request.user.is_superuser:
                        warning_msg = []
                        if attendance_count > 0:
                            warning_msg.append(f'มีข้อมูลการเข้าสอบ {attendance_count} รายการ')
                        if is_ongoing:
                            warning_msg.append('กำลังสอบอยู่')
                        
                        results['warnings'].append({
                            'subject_id': subject_id,
                            'subject_name': subject.subject_name,
                            'warning': f'ข้ามการลบ: {", ".join(warning_msg)}'
                        })
                        results['warning_count'] += 1
                        continue
                    
                    # ลบข้อมูล attendance ก่อน (ถ้ามี)
                    if attendance_count > 0:
                        Attendance.objects.filter(subject=subject).delete()
                    
                    # ลบ subject
                    subject_name = subject.subject_name
                    subject.delete()
                    results['success_count'] += 1
                    
                except ExamSubject.DoesNotExist:
                    results['errors'].append({
                        'subject_id': subject_id,
                        'error': 'ไม่พบรายวิชานี้ในระบบ'
                    })
                    results['error_count'] += 1
                    
                except Exception as e:
                    results['errors'].append({
                        'subject_id': subject_id,
                        'error': str(e)
                    })
                    results['error_count'] += 1
        
        # สร้างข้อความตอบกลับ
        messages = []
        if results['success_count'] > 0:
            messages.append(f'ลบสำเร็จ {results["success_count"]} รายการ')
        if results['warning_count'] > 0:
            messages.append(f'ข้ามการลบ {results["warning_count"]} รายการ (มีข้อจำกัด)')
        if results['error_count'] > 0:
            messages.append(f'ล้มเหลว {results["error_count"]} รายการ')
        
        message = ', '.join(messages) if messages else 'ไม่มีรายการที่ถูกลบ'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'results': results,
            'can_force_delete': request.user.is_superuser and results['warning_count'] > 0
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}'
        }, status=500)
    
# ส่งออกตารางวิชาสอบเป็น Excel (พร้อมหัวตาราง/สไตล์/หมายเหตุ)
@login_required
def export_exam_subjects(request):
    """Export รายการวิชาสอบเป็น Excel"""
    if not request.user.is_staff:
        return HttpResponseForbidden("ไม่มีสิทธิ์เข้าถึง")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # สร้าง workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "รายการวิชาสอบ"
        
        # กำหนด headers
        headers = [
            'ลำดับ', 'รหัสวิชา', 'ชื่อวิชา', 'ปีการศึกษา', 'เทอม',
            'วันที่สอบ', 'เวลาเริ่ม', 'เวลาสิ้นสุด', 'ระยะเวลา (นาที)',
            'อาคาร', 'ห้องสอบ', 'ความจุห้อง', 'ครูคุมสอบหลัก', 'ครูคุมสอบสำรอง',
            'ระดับชั้น', 'จำนวนนักเรียน', 'สถานะ', 'หมายเหตุ'
        ]
        
        # สร้างสไตล์
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ใส่ headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # ดึงข้อมูลวิชาสอบ
        subjects = ExamSubject.objects.select_related(
            'room__building', 'invigilator__user', 'secondary_invigilator__user'
        ).prefetch_related('students').order_by('exam_date', 'start_time')
        
        # เพิ่มข้อมูล
        for row, subject in enumerate(subjects, 2):
            # ค้นหาระดับชั้นและจำนวนนักเรียน
            student_classes = list(subject.students.values_list('student_class', flat=True).distinct())
            student_class_str = ', '.join(student_classes) if student_classes else '-'
            student_count = subject.get_student_count()
            
            # กำหนดข้อมูลแต่ละคอลัมน์
            row_data = [
                row - 1,  # ลำดับ
                subject.subject_code,
                subject.subject_name,
                subject.academic_year,
                subject.get_term_display(),
                subject.exam_date.strftime('%d/%m/%Y'),
                subject.start_time.strftime('%H:%M'),
                subject.end_time.strftime('%H:%M'),
                subject.get_duration(),
                subject.room.building.name if subject.room and subject.room.building else '-',
                subject.room.name if subject.room else '-',
                subject.room.capacity if subject.room else '-',
                subject.invigilator.user.get_full_name() if subject.invigilator else '-',
                subject.secondary_invigilator.user.get_full_name() if subject.secondary_invigilator else '-',
                student_class_str,
                student_count,
                subject.get_status(),
                get_subject_notes(subject)  # ฟังก์ชันสำหรับสร้างหมายเหตุ
            ]
            
            # ใส่ข้อมูล
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                
                # จัดรูปแบบตามประเภทข้อมูล
                if col in [1, 9, 12, 16]:  # ลำดับ, ระยะเวลา, ความจุ, จำนวนนักเรียน
                    cell.alignment = Alignment(horizontal="center")
                elif col in [6, 7, 8]:  # วันที่, เวลา
                    cell.alignment = Alignment(horizontal="center")
        
        # ปรับความกว้างคอลัมน์
        column_widths = [8, 12, 25, 12, 8, 12, 10, 10, 12, 15, 12, 10, 20, 20, 15, 12, 12, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # เพิ่มข้อมูลสรุป
        summary_row = len(subjects) + 3
        ws.cell(row=summary_row, column=1, value="สรุปข้อมูล").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=1, value=f"จำนวนวิชาทั้งหมด: {len(subjects)} วิชา")
        ws.cell(row=summary_row + 2, column=1, value=f"จำนวนนักเรียนทั้งหมด: {sum(s.get_student_count() for s in subjects)} คน")
        ws.cell(row=summary_row + 3, column=1, value=f"วันที่ Export: {timezone.now().strftime('%d/%m/%Y %H:%M')}")
        
        # สร้าง response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'exam_subjects_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'เกิดข้อผิดพลาดในการ Export: {str(e)}')
        return redirect('exam_subjects')

# สร้างสตริง “หมายเหตุ” ต่อวิชา (ห้องไม่พอ/ยังไม่จัดครู/มีเช็คชื่อแล้ว ฯลฯ)
def get_subject_notes(subject):
    """สร้างหมายเหตุสำหรับวิชาสอบ"""
    notes = []
    
    # ตรวจสอบความสมบูรณ์
    if not subject.room:
        notes.append("ยังไม่ระบุห้องสอบ")
    elif subject.room.capacity < subject.get_student_count():
        notes.append("ห้องสอบจุไม่เพียงพอ")
    
    if not subject.invigilator:
        notes.append("ยังไม่จัดครูคุมสอบหลัก")
    
    if not subject.secondary_invigilator:
        notes.append("ไม่มีครูคุมสอบสำรอง")
    
    # ตรวจสอบสถานะการเช็คชื่อ
    attendance_count = Attendance.objects.filter(subject=subject).count()
    if attendance_count > 0:
        notes.append(f"มีการเช็คชื่อแล้ว {attendance_count} คน")
    
    return "; ".join(notes) if notes else "สมบูรณ์"

# อัปเดตวิชาสอบผ่าน AJAX + ตรวจรหัสซ้ำ/เวลาถูกต้อง/ชนตาราง/ห้อง
@login_required
@csrf_exempt
def edit_exam_subject_ajax(request, subject_id):
    """แก้ไขรายวิชาสอบผ่าน AJAX"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        subject = get_object_or_404(ExamSubject, id=subject_id)
        
        # รับข้อมูลจากฟอร์ม
        subject_name = request.POST.get('subject_name', '').strip()
        subject_code = request.POST.get('subject_code', '').strip()
        academic_year = request.POST.get('academic_year', '').strip()
        term = request.POST.get('term', '').strip()
        exam_date = request.POST.get('exam_date')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        student_class = request.POST.get('student_class', '').strip()
        invigilator_id = request.POST.get('invigilator')
        secondary_invigilator_id = request.POST.get('secondary_invigilator')
        room_id = request.POST.get('room')
        
        # ตรวจสอบข้อมูลจำเป็น
        if not all([subject_name, subject_code, academic_year, term, exam_date, start_time, end_time]):
            return JsonResponse({
                'error': 'กรุณากรอกข้อมูลให้ครบถ้วน'
            }, status=400)
        
        # ตรวจสอบรหัสวิชาซ้ำ (ยกเว้นตัวเอง)
        if ExamSubject.objects.filter(
            subject_code=subject_code,
            academic_year=academic_year,
            term=term
        ).exclude(id=subject_id).exists():
            return JsonResponse({
                'error': f'รหัสวิชา {subject_code} ปี {academic_year} เทอม {term} มีอยู่ในระบบแล้ว'
            }, status=400)
        
        # ตรวจสอบเวลาที่ถูกต้อง
        from datetime import datetime
        try:
            start_dt = datetime.strptime(start_time, '%H:%M').time()
            end_dt = datetime.strptime(end_time, '%H:%M').time()
            if start_dt >= end_dt:
                return JsonResponse({
                    'error': 'เวลาเริ่มต้องน้อยกว่าเวลาสิ้นสุด'
                }, status=400)
        except ValueError:
            return JsonResponse({
                'error': 'รูปแบบเวลาไม่ถูกต้อง'
            }, status=400)
        
        with transaction.atomic():
            # อัปเดตข้อมูลพื้นฐาน
            subject.subject_name = subject_name
            subject.subject_code = subject_code
            subject.academic_year = academic_year
            subject.term = term
            subject.exam_date = datetime.strptime(exam_date, '%Y-%m-%d').date()
            subject.start_time = start_dt
            subject.end_time = end_dt
            
            # อัปเดตนักเรียน
            if student_class:
                students = StudentProfile.objects.filter(student_class=student_class)
                if students.exists():
                    subject.students.set(students)
                else:
                    return JsonResponse({
                        'error': f'ไม่พบนักเรียนในระดับชั้น {student_class}'
                    }, status=400)
            
            # อัปเดตครูคุมสอบ
            if invigilator_id:
                try:
                    invigilator = TeacherProfile.objects.get(id=invigilator_id)
                    
                    # ตรวจสอบความขัดแย้งเวลา
                    conflicts = ExamSubject.objects.filter(
                        exam_date=subject.exam_date,
                        start_time__lt=subject.end_time,
                        end_time__gt=subject.start_time
                    ).filter(
                        Q(invigilator=invigilator) | Q(secondary_invigilator=invigilator)
                    ).exclude(id=subject_id)
                    
                    if conflicts.exists():
                        return JsonResponse({
                            'error': f'ครู {invigilator.user.get_full_name()} มีตารางคุมสอบในช่วงเวลานี้แล้ว'
                        }, status=400)
                    
                    subject.invigilator = invigilator
                except TeacherProfile.DoesNotExist:
                    return JsonResponse({
                        'error': 'ไม่พบครูหลักที่เลือก'
                    }, status=400)
            
            # อัปเดตครูสำรอง
            if secondary_invigilator_id and secondary_invigilator_id != invigilator_id:
                try:
                    secondary_invigilator = TeacherProfile.objects.get(id=secondary_invigilator_id)
                    
                    conflicts = ExamSubject.objects.filter(
                        exam_date=subject.exam_date,
                        start_time__lt=subject.end_time,
                        end_time__gt=subject.start_time
                    ).filter(
                        Q(invigilator=secondary_invigilator) | Q(secondary_invigilator=secondary_invigilator)
                    ).exclude(id=subject_id)
                    
                    if conflicts.exists():
                        return JsonResponse({
                            'error': f'ครูสำรอง {secondary_invigilator.user.get_full_name()} มีตารางคุมสอบในช่วงเวลานี้แล้ว'
                        }, status=400)
                    
                    subject.secondary_invigilator = secondary_invigilator
                except TeacherProfile.DoesNotExist:
                    subject.secondary_invigilator = None
            
            # อัปเดตห้องสอบ
            if room_id:
                try:
                    room = ExamRoom.objects.get(id=room_id)
                    
                    # ตรวจสอบความขัดแย้งห้อง
                    room_conflicts = ExamSubject.objects.filter(
                        room=room,
                        exam_date=subject.exam_date,
                        start_time__lt=subject.end_time,
                        end_time__gt=subject.start_time
                    ).exclude(id=subject_id)
                    
                    if room_conflicts.exists():
                        return JsonResponse({
                            'error': f'ห้อง {room.building.name} ห้อง {room.name} มีการใช้งานในช่วงเวลาดังกล่าวแล้ว'
                        }, status=400)
                    
                    student_count = subject.get_student_count()
                    if room.capacity < student_count:
                        return JsonResponse({
                            'warning': f'ห้องจุได้ {room.capacity} คน แต่มีนักเรียน {student_count} คน',
                            'continue': True
                        }, status=200)
                    
                    subject.room = room
                except ExamRoom.DoesNotExist:
                    return JsonResponse({
                        'error': 'ไม่พบห้องสอบที่เลือก'
                    }, status=400)
            
            subject.save()
        
        return JsonResponse({
            'success': True,
            'message': 'แก้ไขรายวิชาสำเร็จ',
            'subject': {
                'id': subject.id,
                'subject_name': subject.subject_name,
                'subject_code': subject.subject_code,
                'academic_year': subject.academic_year,
                'term': subject.get_term_display(),
                'exam_date': subject.exam_date.strftime('%d/%m/%Y'),
                'start_time': subject.start_time.strftime('%H:%M'),
                'end_time': subject.end_time.strftime('%H:%M'),
                'student_count': subject.get_student_count(),
                'room_name': f"{subject.room.building.name} ห้อง {subject.room.name}" if subject.room else None,
                'teacher_name': subject.invigilator.user.get_full_name() if subject.invigilator else None
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}'
        }, status=500)
    
# ========================= จัดสรรทรัพยากร & ตรวจสอบห้อง/ครู =========================
# ตรวจห้องว่างตามวัน-เวลา-อาคาร (AJAX)
@login_required
def check_room_availability(request):
    
    """AJAX endpoint สำหรับตรวจสอบห้องที่ว่างในช่วงเวลาที่กำหนด"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    date = request.GET.get('date')
    start_time = request.GET.get('start_time') 
    end_time = request.GET.get('end_time')
    building_id = request.GET.get('building_id')
    
    if not all([date, start_time, end_time]):
        return JsonResponse({
            'error': 'ข้อมูลไม่ครบถ้วน',
            'success': False
        }, status=400)
    
    try:
        # หาห้องที่ถูกใช้ในช่วงเวลานั้น
        busy_rooms = ExamSubject.objects.filter(
            exam_date=date,
            start_time__lt=end_time,
            end_time__gt=start_time
        ).values_list('room_id', flat=True)
        
        # หาห้องว่าง
        available_rooms = ExamRoom.objects.exclude(
            id__in=busy_rooms
        ).filter(is_active=True)
        
        if building_id:
            available_rooms = available_rooms.filter(building_id=building_id)
            
        available_rooms = available_rooms.select_related('building').order_by('building__name', 'name')
        
        rooms_data = []
        for room in available_rooms:
            rooms_data.append({
                'id': room.id,
                'name': room.name,
                'building_name': room.building.name,
                'capacity': room.capacity,
                'full_name': f"{room.building.name} ห้อง {room.name}",
                'display_text': f"{room.building.name} ห้อง {room.name} (จุ {room.capacity} คน)",
                'has_projector': getattr(room, 'has_projector', False),
                'has_aircon': getattr(room, 'has_aircon', False),
            })
        
        return JsonResponse({
            'success': True,
            'available_rooms': rooms_data,
            'total_available': len(rooms_data)
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

# เลือก “ห้องไหนก็ได้ที่ว่าง” (ไม่เคร่งความจุ) สำหรับเวลาที่กำหนด
def find_available_room(exam_date, start_time, end_time, min_capacity):
    """หาห้องว่าง - ห้องไหนว่างก็เอา ไม่สนใจความจุ"""
    try:
        from datetime import datetime
        
        # แปลงวันที่และเวลา
        if isinstance(exam_date, str):
            exam_date = datetime.strptime(exam_date, '%Y-%m-%d').date()
        if isinstance(start_time, str):
            start_time = datetime.strptime(start_time, '%H:%M').time()
        if isinstance(end_time, str):
            end_time = datetime.strptime(end_time, '%H:%M').time()
        
        print(f"Debug: หาห้องสำหรับ {min_capacity} คน วันที่ {exam_date} เวลา {start_time}-{end_time}")
        
        # หาห้องที่ถูกใช้งานในช่วงเวลานั้น
        busy_rooms = ExamSubject.objects.filter(
            exam_date=exam_date,
            start_time__lt=end_time,
            end_time__gt=start_time,
            room__isnull=False
        ).values_list('room_id', flat=True)
        
        print(f"Debug: ห้องที่ถูกจองในช่วงเวลานี้ {len(busy_rooms)} ห้อง")
        
        # หาห้องทั้งหมดที่ว่าง - ไม่สนใจความจุ
        available_rooms = ExamRoom.objects.filter(
            is_active=True
        ).exclude(id__in=busy_rooms)
        
        print(f"Debug: ห้องว่างทั้งหมด {available_rooms.count()} ห้อง")
        
        if not available_rooms.exists():
            print("Debug: ไม่มีห้องว่างเลย")
            return None
        
        # เลือกห้องแรกที่เจอ - ไม่สนใจความจุ
        selected_room = available_rooms.first()
        
        print(f"Debug: เลือกห้อง {selected_room.building.name} ห้อง {selected_room.name} (จุ {selected_room.capacity} คน)")
        return selected_room
        
    except Exception as e:
        print(f"Error in find_available_room: {str(e)}")
        return None

# ประเมินความเหมาะสมของห้อง (ความจุ/ชนเวลา/คำแนะนำ) (AJAX)
@login_required
def check_room_suitability(request):
    """AJAX endpoint สำหรับตรวจสอบความเหมาะสมของห้องสอบ"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        room_id = data.get('room_id')
        date = data.get('date')
        start_time = data.get('start_time')
        end_time = data.get('end_time')
        student_count = data.get('student_count', 0)
        
        if not all([room_id, date, start_time, end_time]):
            return JsonResponse({
                'error': 'ข้อมูลไม่ครบถ้วน',
                'success': False
            }, status=400)
        
        room = get_object_or_404(ExamRoom, id=room_id)
        warnings = []
        
        # ตรวจสอบความจุห้อง
        if student_count > room.capacity:
            warnings.append(f'ความจุห้องไม่เพียงพอ: ห้องจุได้ {room.capacity} คน แต่มีนักเรียน {student_count} คน')
        
        # ตรวจสอบความขัดแย้งเวลา
        conflicts = ExamSubject.objects.filter(
            room=room,
            exam_date=date,
            start_time__lt=end_time,
            end_time__gt=start_time
        )
        
        if conflicts.exists():
            conflict_subjects = [f'"{c.subject_name}" ({c.start_time.strftime("%H:%M")}-{c.end_time.strftime("%H:%M")})' 
                               for c in conflicts]
            warnings.append(f'ห้องถูกใช้ในช่วงเวลาดังกล่าว: {", ".join(conflict_subjects)}')
        
        # ตรวจสอบคำแนะนำเพิ่มเติม
        if student_count > 0:
            utilization = (student_count / room.capacity) * 100
            if utilization < 50:
                warnings.append(f'ความจุห้องเหลือเฟือ: ใช้เพียง {utilization:.1f}% ของความจุ')
            elif utilization > 90:
                warnings.append(f'ห้องใกล้เต็ม: ใช้ {utilization:.1f}% ของความจุ')
        
        return JsonResponse({
            'warnings': warnings,
            'room_info': {
                'name': room.name,
                'building': room.building.name,
                'capacity': room.capacity,
                'utilization': (student_count / room.capacity * 100) if student_count > 0 else 0
            },
            'success': True
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

# ========================= แก้ไข/ลบวิชาสอบผ่าน AJAX =========================
    
# ลบแบบบังคับ (เฉพาะ superuser) พร้อมลบ attendance ที่เกี่ยวข้อง
@login_required  
@csrf_exempt
def force_delete_exam_subjects(request):
    """ลบรายวิชาแบบบังคับ - สำหรับ superuser เท่านั้น"""
    if not request.user.is_superuser:
        return JsonResponse({'error': 'ต้องเป็น superuser เท่านั้น'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        subject_ids = data.get('subject_ids', [])
        
        if not subject_ids:
            return JsonResponse({'error': 'ไม่ได้ระบุรายวิชาที่จะลบ'}, status=400)
        
        deleted_count = 0
        deleted_attendance_count = 0
        
        with transaction.atomic():
            for subject_id in subject_ids:
                try:
                    subject = ExamSubject.objects.get(id=subject_id)
                    
                    # นับและลบ attendance ก่อน
                    attendance_count = Attendance.objects.filter(subject=subject).count()
                    if attendance_count > 0:
                        Attendance.objects.filter(subject=subject).delete()
                        deleted_attendance_count += attendance_count
                    
                    # ลบ subject
                    subject.delete()
                    deleted_count += 1
                    
                except ExamSubject.DoesNotExist:
                    continue  # ข้ามถ้าไม่พบ
        
        message = f'ลบรายวิชาแบบบังคับสำเร็จ {deleted_count} รายการ'
        if deleted_attendance_count > 0:
            message += f' (รวมข้อมูลการเข้าสอบ {deleted_attendance_count} รายการ)'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'deleted_count': deleted_count,
            'deleted_attendance_count': deleted_attendance_count
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}'
        }, status=500)

@login_required
@csrf_exempt
def ajax_teachers_add(request):
    """เพิ่มครูใหม่ผ่าน AJAX"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง', 'success': False}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed', 'success': False}, status=405)
    
    try:
        data = json.loads(request.body)
        
        # ข้อมูลที่จำเป็น
        required_fields = ['username', 'teacher_id', 'first_name', 'last_name', 'email', 'password']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({
                    'error': f'กรุณากรอก{field}',
                    'success': False
                }, status=400)
        
        # ตรวจสอบว่าข้อมูลไม่ซ้ำ
        if User.objects.filter(username=data['username']).exists():
            return JsonResponse({
                'error': f'Username {data["username"]} มีอยู่แล้ว',
                'success': False
            }, status=400)
        
        if User.objects.filter(email=data['email']).exists():
            return JsonResponse({
                'error': f'อีเมล {data["email"]} มีอยู่แล้ว',
                'success': False
            }, status=400)
        
        if TeacherProfile.objects.filter(teacher_id=data['teacher_id']).exists():
            return JsonResponse({
                'error': f'รหัสครู {data["teacher_id"]} มีอยู่แล้ว',
                'success': False
            }, status=400)
        
        # สร้างครูใหม่
        with transaction.atomic():
            user = User.objects.create_user(
                username=data['username'],
                email=data['email'],
                first_name=data['first_name'],
                last_name=data['last_name'],
                password=data['password'],
                is_teacher=True,
                is_active=True
            )
            
            teacher = TeacherProfile.objects.create(
                user=user,
                teacher_id=data['teacher_id']
            )
        
        return JsonResponse({
            'success': True,
            'message': f'เพิ่มครู {user.get_full_name()} สำเร็จ!',
            'teacher': {
                'id': teacher.id,
                'teacher_id': teacher.teacher_id,
                'name': user.get_full_name(),
                'email': user.email
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'ข้อมูล JSON ไม่ถูกต้อง', 'success': False}, status=400)
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

# สำหรับดึงรายชื่อครูว่างตอนแก้ไขวิชาสอบ
@login_required
def get_available_teachers_for_edit(request):
    """ดึงรายการครูที่ว่างสำหรับแก้ไข"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    date = request.GET.get('date')
    start_time = request.GET.get('start_time')
    end_time = request.GET.get('end_time')
    exclude_subject_id = request.GET.get('exclude_subject_id')
    
    if not all([date, start_time, end_time]):
        return JsonResponse({
            'error': 'ข้อมูลไม่ครบถ้วน',
            'success': False
        }, status=400)
    
    try:
        # หาครูที่มีตารางในช่วงเวลาดังกล่าว (ยกเว้นวิชาที่กำลังแก้ไข)
        busy_teachers_query = ExamSubject.objects.filter(
            exam_date=date,
            start_time__lt=end_time,
            end_time__gt=start_time
        )
        
        if exclude_subject_id:
            busy_teachers_query = busy_teachers_query.exclude(id=exclude_subject_id)
        
        busy_teachers = busy_teachers_query.values_list('invigilator_id', 'secondary_invigilator_id')
        
        # รวม ID ครูที่ไม่ว่าง
        busy_teacher_ids = set()
        for primary, secondary in busy_teachers:
            if primary:
                busy_teacher_ids.add(primary)
            if secondary:
                busy_teacher_ids.add(secondary)
        
        # หาครูที่ว่าง
        available_teachers = TeacherProfile.objects.exclude(
            id__in=busy_teacher_ids
        ).filter(
            user__is_active=True
        ).select_related('user').order_by('user__first_name')
        
        # จัดรูปแบบข้อมูล
        teachers_data = []
        for teacher in available_teachers:
            teachers_data.append({
                'id': teacher.id,
                'name': teacher.user.get_full_name(),
                'teacher_id': teacher.teacher_id,
                'email': teacher.user.email
            })
        
        return JsonResponse({
            'available_teachers': teachers_data,
            'total_available': len(teachers_data),
            'success': True
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

@login_required
def get_available_rooms_for_edit(request):
    """ดึงรายการห้องที่ว่างสำหรับแก้ไข"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    date = request.GET.get('date')
    start_time = request.GET.get('start_time')
    end_time = request.GET.get('end_time')
    exclude_subject_id = request.GET.get('exclude_subject_id')
    building_id = request.GET.get('building_id')
    
    if not all([date, start_time, end_time]):
        return JsonResponse({
            'error': 'ข้อมูลไม่ครบถ้วน',
            'success': False
        }, status=400)
    
    try:
        # หาห้องที่ถูกจองในช่วงเวลานั้น (ยกเว้นวิชาที่กำลังแก้ไข)
        busy_rooms_query = ExamSubject.objects.filter(
            exam_date=date,
            start_time__lt=end_time,
            end_time__gt=start_time
        )
        
        if exclude_subject_id:
            busy_rooms_query = busy_rooms_query.exclude(id=exclude_subject_id)
        
        busy_rooms = busy_rooms_query.values_list('room_id', flat=True)
        
        # หาห้องว่าง
        available_rooms = ExamRoom.objects.exclude(
            id__in=busy_rooms
        ).filter(is_active=True)
        
        if building_id:
            available_rooms = available_rooms.filter(building_id=building_id)
            
        available_rooms = available_rooms.select_related('building').order_by('building__name', 'name')
        
        rooms_data = []
        for room in available_rooms:
            rooms_data.append({
                'id': room.id,
                'name': room.name,
                'building_name': room.building.name,
                'capacity': room.capacity,
                'full_name': f"{room.building.name} ห้อง {room.name}",
                'display_text': f"{room.building.name} ห้อง {room.name} (จุ {room.capacity} คน)",
            })
        
        return JsonResponse({
            'success': True,
            'available_rooms': rooms_data,
            'total_available': len(rooms_data)
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

@login_required
def get_class_list(request):
    """ดึงรายการระดับชั้นทั้งหมด"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    try:
        classes = StudentProfile.objects.values_list('student_class', flat=True).distinct().order_by('student_class')
        class_data = []
        
        for cls in classes:
            student_count = StudentProfile.objects.filter(student_class=cls).count()
            class_data.append({
                'value': cls,
                'text': f"{cls} ({student_count} คน)",
                'student_count': student_count
            })
        
        return JsonResponse({
            'success': True,
            'classes': class_data
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

# ========================= Import รายวิชา =========================
# เพิ่มรายวิชาโดย excel
@staff_member_required
def import_exam_subjects(request):
    """หน้าสำหรับ import ข้อมูลรายวิชาสอบ"""
    
    if request.method == 'POST':
        # ตรวจสอบว่าเป็น AJAX request หรือไม่
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        form = ExamSubjectImportForm(request.POST, request.FILES)
        
        if form.is_valid():
            try:
                # ประมวลผลไฟล์
                subjects_data = form.process_file()
                overwrite = form.cleaned_data['overwrite_existing']
                auto_assign = form.cleaned_data['auto_assign_resources']
                
                # นำเข้าข้อมูล
                result = process_subjects_import(subjects_data, overwrite, auto_assign, request.user)
                
                # สร้างข้อความผลลัพธ์
                success_msg = f"นำเข้าสำเร็จ {result['success_count']} วิชา"
                if result['error_count'] > 0:
                    success_msg += f" (ผิดพลาด {result['error_count']}, จัดอัตโนมัติไม่ครบ {result.get('partial_assigned_count', 0)})"
                
                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'message': success_msg,
                        'details': {
                            'success_count': result['success_count'],
                            'error_count': result['error_count'],
                            'partial_assigned_count': result.get('partial_assigned_count', 0),
                            'errors': result.get('errors', [])[:5]  # แสดงแค่ 5 error แรก
                        }
                    })
                else:
                    # ไม่ใช่ AJAX ใช้ Django messages แล้ว redirect
                    messages.success(request, success_msg)
                    if result['error_count'] > 0:
                        error_details = "\n".join([f"แถว {err['row']}: {err['message']}" for err in result['errors'][:5]])
                        messages.warning(request, f"รายละเอียดข้อผิดพลาด:\n{error_details}")
                    return redirect('import_exam_subjects')
                    
            except Exception as e:
                error_msg = f"เกิดข้อผิดพลาดในการประมวลผลไฟล์: {str(e)}"
                
                if is_ajax:
                    return JsonResponse({
                        'success': False, 
                        'message': error_msg
                    }, status=400)
                else:
                    messages.error(request, error_msg)
        else:
            # Form validation ผิดพลาด
            if is_ajax:
                # รวบรวม error messages จาก form
                errors = []
                for field, field_errors in form.errors.items():
                    for error in field_errors:
                        errors.append(f"{form[field].label if hasattr(form[field], 'label') else field}: {error}")
                
                return JsonResponse({
                    'success': False,
                    'message': 'ข้อมูลไม่ถูกต้อง',
                    'errors': errors
                }, status=400)
            else:
                messages.error(request, 'กรุณาตรวจสอบข้อมูลที่กรอกและไฟล์ที่อัพโหลด')
    
    else:
        # GET request - แสดงฟอร์ม
        form = ExamSubjectImportForm()
    
    return render(request, 'app/staff/import_subjects.html', {
        'form': form,
        'template_url': reverse('download_subject_template')
    })

def process_subjects_import(subjects_data, overwrite=False, auto_assign=True, user=None):
    """ประมวลผลการ import ข้อมูลรายวิชาสอบ - เวอร์ชันปรับปรุงใหม่"""
    result = {
        'success_count': 0,
        'error_count': 0,
        'partial_assigned_count': 0,
        'capacity_warnings': [],
        'errors': []
    }
    
    with transaction.atomic():
        for index, data in enumerate(subjects_data, start=2):
            try:
                # ตรวจสอบข้อมูลจำเป็น
                required_fields = ['subject_name', 'subject_code', 'academic_year', 
                                 'term', 'exam_date', 'start_time', 'end_time', 'student_class']
                missing_fields = [field for field in required_fields if not data.get(field)]
                
                if missing_fields:
                    result['errors'].append({
                        'row': index,
                        'message': f"ข้อมูลไม่ครบ: {', '.join(missing_fields)}"
                    })
                    result['error_count'] += 1
                    continue
                
                # ตรวจสอบรหัสวิชาซ้ำ
                existing_subject = ExamSubject.objects.filter(
                    subject_code=data['subject_code'],
                    academic_year=data['academic_year'],
                    term=data['term']
                ).first()
                
                if existing_subject and not overwrite:
                    result['errors'].append({
                        'row': index,
                        'message': f"รหัสวิชา {data['subject_code']} ปี {data['academic_year']} เทอม {data['term']} มีอยู่ในระบบแล้ว"
                    })
                    result['error_count'] += 1
                    continue
                
                # ตรวจสอบว่ามีนักเรียนในระดับชั้นหรือไม่
                students = StudentProfile.objects.filter(student_class=data['student_class'])
                if not students.exists():
                    result['errors'].append({
                        'row': index,
                        'message': f"ไม่พบนักเรียนในระดับชั้น {data['student_class']}"
                    })
                    result['error_count'] += 1
                    continue
                
                # หาหรือสร้าง subject
                if existing_subject and overwrite:
                    subject = existing_subject
                    subject.subject_name = data['subject_name']
                    subject.exam_date = data['exam_date']
                    subject.start_time = data['start_time']
                    subject.end_time = data['end_time']
                else:
                    subject = ExamSubject.objects.create(
                        subject_name=data['subject_name'],
                        subject_code=data['subject_code'],
                        academic_year=data['academic_year'],
                        term=data['term'],
                        exam_date=data['exam_date'],
                        start_time=data['start_time'],
                        end_time=data['end_time'],
                        created_by=user
                    )
                
                # เพิ่มนักเรียน
                subject.students.set(students)
                student_count = students.count()
                
                # จัดครูและห้องสอบอัตโนมัติแบบเฉลี่ย
                assignment_success = True
                if auto_assign:
                    assignment_success = auto_assign_resources(subject, student_count)
                    
                    # ตรวจสอบความจุหลังจัดห้อง
                    if subject.room and subject.room.capacity < student_count:
                        result['capacity_warnings'].append({
                            'subject': subject.subject_name,
                            'room': f"{subject.room.building.name} ห้อง {subject.room.name}",
                            'capacity': subject.room.capacity,
                            'students': student_count,
                            'message': 'ห้องจุไม่เพียงพอ'
                        })
                
                if auto_assign and not assignment_success:
                    result['partial_assigned_count'] += 1
                
                subject.save()
                result['success_count'] += 1
                
                # แสดงข้อมูลสำเร็จแบบละเอียด
                if auto_assign:
                    room_info = f" | ห้อง: {subject.room.building.name} ห้อง {subject.room.name}" if subject.room else " | ห้อง: ยังไม่ระบุ"
                    teacher_info = f" | ครู: {subject.invigilator.user.get_full_name()}" if subject.invigilator else " | ครู: ยังไม่ระบุ"
                    print(f"✓ นำเข้า: {subject.subject_name} ({student_count} คน){room_info}{teacher_info}")
                
            except Exception as e:
                result['errors'].append({
                    'row': index,
                    'message': f"ข้อผิดพลาด: {str(e)}"
                })
                result['error_count'] += 1
    
    # แสดงสรุปผลการ import
    print(f"\nสรุปผลการ Import:")
    print(f"  - สำเร็จ: {result['success_count']} วิชา")
    print(f"  - ผิดพลาด: {result['error_count']} วิชา")
    print(f"  - จัดอัตโนมัติไม่ครบ: {result['partial_assigned_count']} วิชา")
    
    return result

def find_available_teachers(exam_date, start_time, end_time, min_count=1):
    """หาครูที่ว่างในช่วงเวลาที่กำหนด - ปรับปรุงใหม่"""
    try:
        from datetime import datetime
        
        # แปลงวันที่และเวลาให้เป็นรูปแบบที่ถูกต้อง
        if isinstance(exam_date, str):
            exam_date = datetime.strptime(exam_date, '%Y-%m-%d').date()
        
        if isinstance(start_time, str):
            start_time = datetime.strptime(start_time, '%H:%M').time()
        
        if isinstance(end_time, str):
            end_time = datetime.strptime(end_time, '%H:%M').time()
        
        # หาครูที่มีตารางในช่วงเวลาดังกล่าว
        busy_teachers = ExamSubject.objects.filter(
            exam_date=exam_date,
            start_time__lt=end_time,
            end_time__gt=start_time
        ).values_list('invigilator_id', 'secondary_invigilator_id')
        
        # รวม ID ครูที่ไม่ว่าง
        busy_teacher_ids = set()
        for primary, secondary in busy_teachers:
            if primary:
                busy_teacher_ids.add(primary)
            if secondary:
                busy_teacher_ids.add(secondary)
        
        # หาครูที่ว่าง
        available_teachers = TeacherProfile.objects.exclude(
            id__in=busy_teacher_ids
        ).filter(
            user__is_active=True
        ).select_related('user').order_by('user__first_name')[:min_count * 2]  # เพิ่ม buffer
        
        return list(available_teachers)
        
    except Exception as e:
        print(f"Error in find_available_teachers: {str(e)}")
        return []

def get_room_assignment_summary(subject):
    """สรุปสถานะการจัดห้องสำหรับรายวิชา"""
    if not subject.room:
        return {
            'status': 'no_room',
            'message': 'ยังไม่ได้จัดห้องสอบ',
            'color': 'red'
        }
    
    student_count = subject.get_student_count()
    room_capacity = subject.room.capacity
    
    if room_capacity >= student_count:
        utilization = (student_count / room_capacity) * 100
        
        if utilization >= 80:
            return {
                'status': 'excellent',
                'message': f'ห้องเหมาะสมมาก ({utilization:.0f}%)',
                'color': 'green'
            }
        elif utilization >= 50:
            return {
                'status': 'good',
                'message': f'ห้องเหมาะสม ({utilization:.0f}%)',
                'color': 'blue'
            }
        elif utilization >= 20:
            return {
                'status': 'spacious',
                'message': f'ห้องกว้างขวาง ({utilization:.0f}%)',
                'color': 'yellow'
            }
        else:
            return {
                'status': 'oversized',
                'message': f'ห้องใหญ่มาก ({utilization:.0f}%)',
                'color': 'orange'
            }
    else:
        shortage = student_count - room_capacity
        return {
            'status': 'insufficient',
            'message': f'ห้องจุไม่พอ (ขาด {shortage} ที่นั่ง)',
            'color': 'red'
        }

def auto_assign_resources(subject, student_count=None):
    """จัดครูและห้องสอบอัตโนมัติแบบเฉลี่ย - กระจายภาระงาน"""
    try:
        success = True
        
        if student_count is None:
            student_count = subject.get_student_count()
        
        print(f"Debug: กำลังจัดทรัพยากรให้ {subject.subject_name} ({student_count} คน)")
        
        # จัดห้องแบบเฉลี่ย
        if not subject.room:
            # ใช้ฟังก์ชันใหม่ที่ปรับปรุงแล้ว
            available_room = find_available_room_balanced(
                subject.exam_date,
                subject.start_time, 
                subject.end_time,
                student_count
            )
            
            if available_room:
                subject.room = available_room
                print(f"✓ จัดห้อง: {available_room.building.name} ห้อง {available_room.name}")
            else:
                print("✗ ไม่พบห้องว่างสำหรับการสอบ")
                success = False
        
        # จัดครูแบบเฉลี่ย
        if not subject.invigilator:
            # ใช้ฟังก์ชันใหม่ที่ปรับปรุงแล้ว
            available_teachers = find_available_teachers_balanced(
                subject.exam_date,
                subject.start_time,
                subject.end_time,
                min_count=2  # ต้องการครูอย่างน้อย 2 คน
            )
            
            if len(available_teachers) >= 1:
                subject.invigilator = available_teachers[0]
                print(f"✓ จัดครูหลัก: {available_teachers[0].user.get_full_name()}")
                
                # จัดครูสำรอง
                if len(available_teachers) >= 2:
                    subject.secondary_invigilator = available_teachers[1]
                    print(f"✓ จัดครูสำรอง: {available_teachers[1].user.get_full_name()}")
            else:
                print("✗ ไม่มีครูว่างในช่วงเวลานี้")
                success = False
        
        subject.save()
        return success
        
    except Exception as e:
        print(f"Error in auto_assign_resources: {str(e)}")
        return False

def find_available_room_balanced(exam_date, start_time, end_time, student_count):
    """หาห้องว่างแบบเฉลี่ย - กระจายการใช้ห้องให้ทุกห้อง"""
    try:
        from datetime import datetime
        from django.db.models import Count
        
        # แปลงวันที่และเวลา
        if isinstance(exam_date, str):
            exam_date = datetime.strptime(exam_date, '%Y-%m-%d').date()
        if isinstance(start_time, str):
            start_time = datetime.strptime(start_time, '%H:%M').time()
        if isinstance(end_time, str):
            end_time = datetime.strptime(end_time, '%H:%M').time()
        
        print(f"Debug: หาห้องแบบเฉลี่ยสำหรับ {student_count} คน วันที่ {exam_date} เวลา {start_time}-{end_time}")
        
        # หาห้องที่ถูกจองในช่วงเวลานั้น
        busy_rooms = ExamSubject.objects.filter(
            exam_date=exam_date,
            start_time__lt=end_time,
            end_time__gt=start_time,
            room__isnull=False
        ).values_list('room_id', flat=True)
        
        # หาห้องทั้งหมดที่ว่างและมีความจุเพียงพอ
        available_rooms = ExamRoom.objects.filter(
            is_active=True,
            capacity__gte=student_count
        ).exclude(id__in=busy_rooms)
        
        # นับจำนวนการใช้ห้องแต่ละห้อง (เฉพาะในปีการศึกษาปัจจุบัน)
        current_year = timezone.now().year + 543  # แปลงเป็น พ.ศ.
        room_usage_count = {}
        
        for room in available_rooms:
            usage_count = ExamSubject.objects.filter(
                room=room,
                academic_year=str(current_year)
            ).count()
            room_usage_count[room.id] = usage_count
        
        # เรียงลำดับห้องตามการใช้งาน (น้อยไปมาก) และความจุ
        sorted_rooms = sorted(available_rooms, key=lambda r: (
            room_usage_count.get(r.id, 0),  # ใช้น้อยสุดก่อน
            abs(r.capacity - student_count),  # ขนาดใกล้เคียงกับจำนวนนักเรียน
        ))
        
        if sorted_rooms:
            selected_room = sorted_rooms[0]
            print(f"Debug: เลือกห้อง {selected_room.building.name} ห้อง {selected_room.name} (ใช้งานไปแล้ว {room_usage_count.get(selected_room.id, 0)} ครั้ง)")
            return selected_room
        
        print("Debug: ไม่มีห้องว่างที่เหมาะสม")
        return None
        
    except Exception as e:
        print(f"Error in find_available_room_balanced: {str(e)}")
        return None

def find_available_teachers_balanced(exam_date, start_time, end_time, min_count=1):
    """หาครูที่ว่างแบบเฉลี่ย - กระจายภาระงานให้ทุกคน"""
    try:
        from datetime import datetime
        from django.db.models import Count, Q
        
        # แปลงวันที่และเวลา
        if isinstance(exam_date, str):
            exam_date = datetime.strptime(exam_date, '%Y-%m-%d').date()
        if isinstance(start_time, str):
            start_time = datetime.strptime(start_time, '%H:%M').time()
        if isinstance(end_time, str):
            end_time = datetime.strptime(end_time, '%H:%M').time()
        
        print(f"Debug: หาครูแบบเฉลี่ยสำหรับ {exam_date} เวลา {start_time}-{end_time}")
        
        # หาครูที่มีตารางในช่วงเวลาดังกล่าว
        busy_teachers = ExamSubject.objects.filter(
            exam_date=exam_date,
            start_time__lt=end_time,
            end_time__gt=start_time
        ).values_list('invigilator_id', 'secondary_invigilator_id')
        
        # รวม ID ครูที่ไม่ว่าง
        busy_teacher_ids = set()
        for primary, secondary in busy_teachers:
            if primary:
                busy_teacher_ids.add(primary)
            if secondary:
                busy_teacher_ids.add(secondary)
        
        # หาครูที่ว่าง
        available_teachers = TeacherProfile.objects.exclude(
            id__in=busy_teacher_ids
        ).filter(
            user__is_active=True
        ).select_related('user')
        
        # นับจำนวนการคุมสอบของแต่ละครู (ในปีการศึกษาปัจจุบัน)
        current_year = timezone.now().year + 543  # แปลงเป็น พ.ศ.
        teacher_workload = {}
        
        for teacher in available_teachers:
            # นับทั้งครูหลักและครูสำรอง
            primary_count = ExamSubject.objects.filter(
                invigilator=teacher,
                academic_year=str(current_year)
            ).count()
            
            secondary_count = ExamSubject.objects.filter(
                secondary_invigilator=teacher,
                academic_year=str(current_year)
            ).count()
            
            # คำนวณภาระงานรวม (ครูสำรองคิดครึ่งงาน)
            total_workload = primary_count + (secondary_count * 0.5)
            teacher_workload[teacher.id] = total_workload
        
        # เรียงลำดับครูตามภาระงาน (น้อยไปมาก)
        sorted_teachers = sorted(available_teachers, key=lambda t: (
            teacher_workload.get(t.id, 0),  # ภาระงานน้อยก่อน
            t.user.first_name  # เรียงตามชื่อถ้าภาระเท่ากัน
        ))
        
        print(f"Debug: พบครูว่าง {len(sorted_teachers)} คน, ต้องการ {min_count} คน")
        for i, teacher in enumerate(sorted_teachers[:5]):  # แสดงแค่ 5 คนแรก
            workload = teacher_workload.get(teacher.id, 0)
            print(f"  {i+1}. {teacher.user.get_full_name()}: ภาระงาน {workload} ครั้ง")
        
        return list(sorted_teachers[:min_count * 3])  # ส่งคืนตามจำนวนที่ต้องการ + buffer
        
    except Exception as e:
        print(f"Error in find_available_teachers_balanced: {str(e)}")
        return []
    
@staff_member_required
@require_GET
def download_subject_template(request):
    """
    ดาวน์โหลดไฟล์ template สำหรับ import รายวิชา
    - ปกติส่ง .xlsx (ต้องมี openpyxl)
    - ถ้าไม่มี openpyxl หรือ ?format=csv จะส่งเป็น .csv
    """
    fmt = (request.GET.get("format") or "xlsx").lower()

    headers = [
        "subject_name", "subject_code", "academic_year", "term",
        "exam_date", "start_time", "end_time", "student_class",
    ]
    sample = [
        ["คณิตศาสตร์", "MATH101", "2567", "1", "2025-03-15", "09:00", "11:00", "ม.1/1"],
        ["ฟิสิกส์",    "PHYS101", "2567", "1", "2025-03-16", "13:00", "15:00", "ม.2/1"],
    ]
    tips = [
        "1. วันที่: YYYY-MM-DD เช่น 2025-03-15",
        "2. เวลา: HH:MM เช่น 09:00",
        "3. ภาคเรียน: 1, 2, หรือ 3",
        "4. รหัสวิชาไม่ควรซ้ำ",
        "5. ระดับชั้นต้องตรงกับในระบบ",
    ]

    # ── CSV ─────────────────────────────────────────────────────────
    if fmt == "csv":
        import csv, io
        buf = io.StringIO()
        w = csv.writer(buf)

        # comment อธิบาย
        w.writerow(["# Exam Subject Import Template"])
        w.writerow(["# ลบแถวที่ขึ้นต้นด้วย # ก่อนนำเข้า"])
        w.writerow([])

        # header + ตัวอย่าง
        w.writerow(headers)
        w.writerows(sample)

        # คำแนะนำ
        w.writerow([])
        w.writerow(["# คำแนะนำ"])
        for t in tips:
            w.writerow([f"# {t}"])

        data = buf.getvalue().encode("utf-8-sig")  # BOM เพื่อให้ Excel อ่านไทยได้ดี
        resp = HttpResponse(data, content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="exam_subject_template_{timezone.now():%Y%m%d}.csv"'
        return resp

    # ── XLSX (ปกติ) ────────────────────────────────────────────────
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception:
        # ไม่มี openpyxl → ส่ง CSV แทน
        return HttpResponseRedirect(reverse("download_subject_template") + "?format=csv")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subject Import Template"

    # header
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # ตัวอย่าง
    for r, row in enumerate(sample, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    # ปรับความกว้าง
    for col in ws.columns:
        max_len = max((len(str(x.value)) if x.value is not None else 0) for x in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # คำแนะนำ
    base = 2 + len(sample) + 2
    ws.cell(row=base, column=1, value="คำแนะนำการใช้งาน:").font = Font(bold=True)
    for i, t in enumerate(tips, 1):
        ws.cell(row=base + i, column=1, value=t)

    # ส่งไฟล์ด้วย FileResponse (เสถียรกว่า wb.save(HttpResponse))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"exam_subject_template_{timezone.now():%Y%m%d}.xlsx"
    return FileResponse(
        buf,
        as_attachment=True,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
# ========================= จัดการห้องสอบ =========================

@login_required
def manage_rooms(request):
    """จัดการห้องสอบ"""
    buildings = Building.objects.all().prefetch_related('rooms')
    return render(request, 'app/staff/manage_rooms.html', {'buildings': buildings})

logger = logging.getLogger(__name__)

@contextmanager
def acquire_lock(lock_key, timeout=10):
    """Context manager สำหรับ distributed lock - ปรับปรุงแล้ว"""
    lock_acquired = False
    lock_value = f"{time_module.time()}_{hash(lock_key)}"  # Unique lock value
    
    try:
        # พยายามได้ lock
        for attempt in range(timeout * 10):  # check every 0.1 second
            if cache.add(lock_key, lock_value, timeout=timeout):
                lock_acquired = True
                logger.debug(f"Lock acquired: {lock_key}")
                break
            time_module.sleep(0.1)
        
        if not lock_acquired:
            logger.warning(f"Failed to acquire lock: {lock_key}")
            raise TimeoutError(f"ไม่สามารถได้ lock '{lock_key}' ได้ภายในเวลาที่กำหนด")
        
        yield
        
    finally:
        if lock_acquired:
            # ตรวจสอบว่า lock ยังเป็นของเราอยู่ก่อนลบ
            current_value = cache.get(lock_key)
            if current_value == lock_value:
                cache.delete(lock_key)
                logger.debug(f"Lock released: {lock_key}")
            else:
                logger.warning(f"Lock value mismatch during release: {lock_key}")

# เพิ่มฟังก์ชันสำหรับทำความสะอาด locks ที่หมดอายุ
def cleanup_expired_locks():
    """ทำความสะอาด locks ที่หมดอายุ - เรียกใช้ใน management command หรือ cron job"""
    try:
        # Django cache จะจัดการ TTL เอง แต่เราสามารถเพิ่ม cleanup logic เพิ่มเติมได้
        logger.info("Lock cleanup completed")
    except Exception as e:
        logger.error(f"Error during lock cleanup: {str(e)}")

@login_required
def add_building(request):
    """เพิ่มอาคาร"""
    if request.method == 'POST':
        form = BuildingForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'เพิ่มอาคารสำเร็จ!')
            return redirect('manage_rooms')
    else:
        form = BuildingForm()
    return render(request, 'app/staff/add_building.html', {'form': form})

# เพิ่ม middleware สำหรับ debug race conditions (ใช้เฉพาะในการ develop)
class RaceConditionDebugMiddleware:
    def __init__(self, get_response):
        self.get_response = get_response

    def __call__(self, request):
        if request.path.startswith('/ajax/rooms/add/'):
            import random
            # จำลอง network delay เพื่อทดสอบ race condition
            if hasattr(request, '_debug_delay'):
                time_module.sleep(random.uniform(0.1, 0.5))
        
        response = self.get_response(request)
        return response
    
@login_required
def add_room(request):
    """เพิ่มห้องสอบ"""
    if request.method == 'POST':
        form = ExamRoomForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'เพิ่มห้องสอบสำเร็จ!')
            return redirect('manage_rooms')
    else:
        form = ExamRoomForm()
    return render(request, 'app/staff/add_room.html', {'form': form})

@login_required
def get_buildings_data(request):
    """AJAX endpoint สำหรับดึงข้อมูลอาคารทั้งหมด"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    try:
        buildings = Building.objects.prefetch_related('rooms').all().order_by('code')
        buildings_data = []
        
        for building in buildings:
            rooms_data = []
            for room in building.rooms.all():
                rooms_data.append({
                    'id': room.id,
                    'name': room.name,
                    'capacity': room.capacity,
                })
            
            buildings_data.append({
                'id': building.id,
                'code': building.code,
                'name': building.name,
                'description': building.description,
                'rooms': rooms_data,
                'room_count': len(rooms_data),
                'total_capacity': sum(room['capacity'] for room in rooms_data)
            })
        
        return JsonResponse({
            'buildings': buildings_data,
            'success': True
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

@login_required
def get_room_statistics(request):
    """AJAX endpoint สำหรับดึงสถิติห้องสอบ"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    try:
        total_buildings = Building.objects.count()
        total_rooms = ExamRoom.objects.count()
        total_capacity = ExamRoom.objects.aggregate(
            total=Sum('capacity')
        )['total'] or 0
        
        # สถิติการใช้งานห้อง (จากการสอบ)
        today = timezone.now().date()
        rooms_in_use_today = ExamSubject.objects.filter(
            exam_date=today
        ).values('room').distinct().count()
        
        return JsonResponse({
            'statistics': {
                'total_buildings': total_buildings,
                'total_rooms': total_rooms,
                'total_capacity': total_capacity,
                'rooms_in_use_today': rooms_in_use_today,
                'average_capacity': round(total_capacity / total_rooms, 1) if total_rooms > 0 else 0
            },
            'success': True
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

@login_required
def export_rooms_data(request):
    """ส่งออกข้อมูลห้องสอบเป็น Excel"""
    if not request.user.is_staff:
        return HttpResponseForbidden("ไม่มีสิทธิ์เข้าถึง")
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rooms Data"
    
    # หัวตาราง
    headers = ['รหัสอาคาร','ชื่ออาคาร','ชื่อห้อง','ความจุ','รายละเอียดอาคาร']
    
    # สร้างหัวตาราง
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # ดึงข้อมูล
    rooms = ExamRoom.objects.select_related('building').all().order_by('building__code', 'name')
    
    # เพิ่มข้อมูล
    for row, room in enumerate(rooms, 2):
        ws.cell(row=row, column=1, value=room.building.code)
        ws.cell(row=row, column=2, value=room.building.name)
        ws.cell(row=row, column=3, value=room.name)
        ws.cell(row=row, column=4, value=room.capacity)
        ws.cell(row=row, column=5, value=room.building.description)
    
    # ปรับขนาดคอลัมน์
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # สร้าง response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'rooms_export_{timezone.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def add_building_ajax(request):
    """เพิ่มอาคารผ่าน AJAX - ปรับปรุงเพื่อป้องกัน race condition"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง', 'success': False}, status=403)
    
    try:
        # ตรวจสอบประเภทข้อมูลที่ส่งมา
        content_type = request.content_type or ''
        
        if 'application/json' in content_type:
            try:
                data = json.loads(request.body)
                code = data.get('code', '').strip().upper()
                name = data.get('name', '').strip()
                description = data.get('description', '').strip()
            except json.JSONDecodeError:
                return JsonResponse({'error': 'รูปแบบ JSON ไม่ถูกต้อง', 'success': False}, status=400)
        else:
            # FormData
            code = (request.POST.get('code') or '').strip().upper()
            name = (request.POST.get('name') or '').strip()
            description = (request.POST.get('description') or '').strip()
        
        # ตรวจสอบข้อมูลที่จำเป็น
        if not code or not name:
            return JsonResponse({
                'error': 'กรุณากรอกรหัสอาคารและชื่ออาคาร',
                'success': False
            }, status=400)
        
        # สร้าง lock key สำหรับ building code
        lock_key = f"building_creation_{code.lower()}"
        
        # ใช้ distributed lock เพื่อป้องกัน race condition
        try:
            with acquire_lock(lock_key, timeout=5):
                # ตรวจสอบรหัสซ้ำภายใน lock
                if Building.objects.filter(code=code).exists():
                    return JsonResponse({
                        'error': f'รหัสอาคาร {code} มีอยู่ในระบบแล้ว',
                        'success': False
                    }, status=400)
                
                # สร้างอาคารใหม่
                with transaction.atomic():
                    building = Building.objects.create(
                        code=code,
                        name=name,
                        description=description or ''
                    )
                
                return JsonResponse({
                    'success': True,
                    'message': f'เพิ่มอาคาร {building.name} สำเร็จ!',
                    'building': {
                        'id': building.id,
                        'code': building.code,
                        'name': building.name,
                        'description': building.description or '',
                        'room_count': 0,
                        'total_capacity': 0,
                        'rooms': []
                    }
                })
                
        except TimeoutError:
            return JsonResponse({
                'error': 'ระบบกำลังประมวลผล กรุณารอสักครู่แล้วลองใหม่',
                'success': False
            }, status=429)
            
    except IntegrityError:
        return JsonResponse({
            'error': f'รหัสอาคาร {code} ถูกสร้างโดยผู้ใช้อื่นแล้วในขณะที่คุณกำลังเพิ่มข้อมูล',
            'success': False
        }, status=400)
        
    except Exception as e:
        print(f"Error in add_building_ajax: {str(e)}")
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def add_room_ajax(request):
    """เพิ่มห้องสอบผ่าน AJAX - แก้ไข race condition"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง', 'success': False}, status=403)
    
    try:
        # ตรวจสอบประเภทข้อมูลที่ส่งมา
        content_type = request.content_type or ''
        
        if 'application/json' in content_type:
            try:
                data = json.loads(request.body)
                building_id = data.get('building_id')
                name = data.get('name', '').strip()
                capacity = data.get('capacity')
            except json.JSONDecodeError:
                return JsonResponse({'error': 'รูปแบบ JSON ไม่ถูกต้อง', 'success': False}, status=400)
        else:
            # FormData
            building_id = request.POST.get('building_id')
            name = (request.POST.get('name') or '').strip()
            capacity = request.POST.get('capacity')
        
        logger.info(f"Room creation request - building_id: {building_id}, name: '{name}', capacity: {capacity}")
        
        # ตรวจสอบข้อมูลที่จำเป็น
        if not building_id or not name or not capacity:
            return JsonResponse({
                'error': 'กรุณากรอกข้อมูลให้ครบถ้วน (อาคาร, ชื่อห้อง, ความจุ)',
                'success': False
            }, status=400)
        
        # ตรวจสอบและแปลงความจุ
        try:
            capacity = int(capacity)
            if capacity <= 0:
                raise ValueError('ความจุต้องมากกว่า 0')
            if capacity > 500:
                raise ValueError('ความจุต้องไม่เกิน 500 คน')
        except (ValueError, TypeError) as e:
            return JsonResponse({
                'error': f'ความจุไม่ถูกต้อง: {str(e)}',
                'success': False
            }, status=400)
        
        # ตรวจสอบอาคาร
        try:
            building = Building.objects.get(id=int(building_id))
            logger.info(f"Found building - {building.name} ({building.code})")
        except (Building.DoesNotExist, ValueError):
            return JsonResponse({
                'error': 'ไม่พบอาคารที่ระบุ',
                'success': False
            }, status=400)
        
        # สร้าง lock key ที่ unique และเข้าใจง่าย
        lock_key = f"room_create_{building_id}_{name.lower().replace(' ', '_')}"
        
        # ใช้ distributed lock เพื่อป้องกัน race condition
        try:
            with acquire_lock(lock_key, timeout=10):
                logger.info(f"Lock acquired for room creation: {lock_key}")
                
                # ใช้ atomic transaction พร้อม select_for_update
                with transaction.atomic():
                    # ตรวจสอบชื่อห้องซ้ำในอาคารเดียวกันพร้อม lock row
                    existing_rooms = ExamRoom.objects.select_for_update().filter(
                        building=building, 
                        name__iexact=name
                    )
                    
                    logger.info(f"Checking for existing rooms with name '{name}' in building '{building.name}'")
                    
                    if existing_rooms.exists():
                        existing_room = existing_rooms.first()
                        logger.warning(f"Room already exists: {existing_room.id}")
                        return JsonResponse({
                            'error': f'ห้อง "{name}" มีอยู่ในอาคาร "{building.name}" แล้ว',
                            'success': False
                        }, status=400)
                    
                    # สร้างห้องใหม่
                    logger.info(f"Creating new room '{name}' in building '{building.name}' with capacity {capacity}")
                    
                    room = ExamRoom.objects.create(
                        building=building,
                        name=name,
                        capacity=capacity,
                        is_active=True
                    )
                    
                    logger.info(f"Successfully created room with ID {room.id}")
                
                return JsonResponse({
                    'success': True,
                    'message': f'เพิ่มห้อง {room.name} ในอาคาร {building.name} สำเร็จ!',
                    'room': {
                        'id': room.id,
                        'name': room.name,
                        'capacity': room.capacity,
                        'is_active': room.is_active,
                        'building': {
                            'id': building.id,
                            'name': building.name,
                            'code': building.code
                        }
                    }
                })
                
        except TimeoutError:
            logger.error(f"Timeout acquiring lock: {lock_key}")
            return JsonResponse({
                'error': 'ระบบกำลังประมวลผล กรุณารอสักครู่แล้วลองใหม่',
                'success': False
            }, status=429)
            
    except IntegrityError as integrity_error:
        logger.error(f"IntegrityError during room creation: {str(integrity_error)}")
        
        # ตรวจสอบสาเหตุของ IntegrityError
        error_message = str(integrity_error).lower()
        
        if 'unique' in error_message or 'duplicate' in error_message:
            return JsonResponse({
                'error': f'ห้อง "{name}" ถูกสร้างโดยผู้ใช้อื่นแล้วในขณะที่คุณกำลังเพิ่มข้อมูล',
                'success': False
            }, status=400)
        else:
            return JsonResponse({
                'error': 'เกิดข้อผิดพลาดในการสร้างห้อง: ข้อมูลไม่ถูกต้อง',
                'success': False
            }, status=400)
            
    except Exception as e:
        logger.error(f"Unexpected error in add_room_ajax: {str(e)}", exc_info=True)
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาดที่ไม่คาดคิด: {str(e)}',
            'success': False
        }, status=500)

@login_required
@csrf_exempt
def edit_building_ajax(request, building_id):
    """แก้ไขอาคารผ่าน AJAX"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง', 'success': False}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed', 'success': False}, status=405)
    
    try:
        building = get_object_or_404(Building, id=building_id)
        
        # ตรวจสอบประเภทข้อมูลที่ส่งมา
        content_type = request.content_type or ''
        
        if 'application/json' in content_type:
            try:
                data = json.loads(request.body)
                code = data.get('code', '').strip().upper()
                name = data.get('name', '').strip()
                description = data.get('description', '').strip()
            except json.JSONDecodeError:
                return JsonResponse({'error': 'Invalid JSON format', 'success': False}, status=400)
        else:
            # FormData
            code = (request.POST.get('code') or '').strip().upper()
            name = (request.POST.get('name') or '').strip()
            description = (request.POST.get('description') or '').strip()
        
        # ตรวจสอบข้อมูลที่จำเป็น
        if not code or not name:
            return JsonResponse({
                'error': 'กรุณากรอกรหัสอาคารและชื่อาคาร',
                'success': False
            }, status=400)
        
        # ตรวจสอบรหัสซ้ำ (ยกเว้นตัวเอง)
        if Building.objects.filter(code=code).exclude(id=building_id).exists():
            return JsonResponse({
                'error': f'รหัสอาคาร {code} มีอยู่ในระบบแล้ว',
                'success': False
            }, status=400)
        
        # อัปเดตข้อมูล
        with transaction.atomic():
            building.code = code
            building.name = name
            building.description = description or ''
            building.save()
        
        return JsonResponse({
            'success': True,
            'message': f'แก้ไขอาคาร {building.name} สำเร็จ!',
            'building': {
                'id': building.id,
                'code': building.code,
                'name': building.name,
                'description': building.description or ''
            }
        })
        
    except Exception as e:
        print(f"Error in edit_building_ajax: {str(e)}")
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def edit_room_ajax(request, room_id):
    """แก้ไขห้องสอบผ่าน AJAX - ปรับปรุงเพื่อป้องกัน race condition"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง', 'success': False}, status=403)
    
    try:
        room = get_object_or_404(ExamRoom, id=room_id)
        
        # ตรวจสอบประเภทข้อมูลที่ส่งมา
        content_type = request.content_type or ''
        
        if 'application/json' in content_type:
            try:
                data = json.loads(request.body)
                building_id = data.get('building_id')
                name = data.get('name', '').strip()
                capacity = data.get('capacity')
            except json.JSONDecodeError:
                return JsonResponse({'error': 'รูปแบบ JSON ไม่ถูกต้อง', 'success': False}, status=400)
        else:
            # FormData
            building_id = request.POST.get('building_id')
            name = (request.POST.get('name') or '').strip()
            capacity = request.POST.get('capacity')
        
        # ตรวจสอบข้อมูลที่จำเป็น
        if not building_id or not name or not capacity:
            return JsonResponse({
                'error': 'กรุณากรอกข้อมูลให้ครบถ้วน',
                'success': False
            }, status=400)
        
        # ตรวจสอบและแปลงความจุ
        try:
            capacity = int(capacity)
            if capacity <= 0:
                raise ValueError('ความจุต้องมากกว่า 0')
            if capacity > 500:
                raise ValueError('ความจุต้องไม่เกิน 500 คน')
        except (ValueError, TypeError) as e:
            return JsonResponse({
                'error': f'ความจุไม่ถูกต้อง: {str(e)}',
                'success': False
            }, status=400)
        
        # ตรวจสอบอาคาร
        try:
            building = Building.objects.get(id=int(building_id))
        except (Building.DoesNotExist, ValueError):
            return JsonResponse({
                'error': 'ไม่พบอาคารที่ระบุ',
                'success': False
            }, status=400)
        
        # สร้าง lock key สำหรับการแก้ไขห้อง
        lock_key = f"room_edit_{building_id}_{name.lower().replace(' ', '_')}_{room_id}"
        
        try:
            with acquire_lock(lock_key, timeout=5):
                # ตรวจสอบชื่อห้องซ้ำในอาคารเดียวกัน (ยกเว้นตัวเอง)
                existing_room = ExamRoom.objects.filter(
                    building=building, 
                    name__iexact=name
                ).exclude(id=room_id).select_for_update()
                
                if existing_room.exists():
                    return JsonResponse({
                        'error': f'ห้อง {name} มีอยู่ในอาคาร {building.name} แล้ว',
                        'success': False
                    }, status=400)
                
                # อัปเดตข้อมูล
                with transaction.atomic():
                    room.building = building
                    room.name = name
                    room.capacity = capacity
                    room.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'แก้ไขห้อง {room.name} ในอาคาร {building.name} สำเร็จ!',
                    'room': {
                        'id': room.id,
                        'name': room.name,
                        'capacity': room.capacity,
                        'building': {
                            'id': building.id,
                            'name': building.name,
                            'code': building.code
                        }
                    }
                })
                
        except TimeoutError:
            return JsonResponse({
                'error': 'ระบบกำลังประมวลผล กรุณารอสักครู่แล้วลองใหม่',
                'success': False
            }, status=429)
            
    except IntegrityError as e:
        # Handle database constraint violations
        error_msg = 'ข้อมูลซ้ำในระบบ'
        if 'UNIQUE constraint failed' in str(e):
            if 'building_id' in str(e) and 'name' in str(e):
                error_msg = f'ห้อง {name} มีอยู่ในอาคารนี้แล้ว'
        
        print(f"IntegrityError in edit_room_ajax: {str(e)}")
        return JsonResponse({
            'error': error_msg,
            'success': False
        }, status=400)
        
    except Exception as e:
        print(f"Error in edit_room_ajax: {str(e)}")
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

@login_required
@csrf_exempt
def delete_building_ajax(request, building_id):
    """ลบอาคารผ่าน AJAX - ปรับปรุงเพื่อป้องกัน concurrent deletion"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง', 'success': False}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed', 'success': False}, status=405)
    
    try:
        # ใช้ select_for_update เพื่อ lock record
        building = get_object_or_404(Building.objects.select_for_update(), id=building_id)
        
        # สร้าง lock key สำหรับการลบอาคาร
        lock_key = f"building_deletion_{building_id}"
        
        try:
            with acquire_lock(lock_key, timeout=5):
                # ตรวจสอบว่ามีห้องที่กำลังถูกใช้สอบหรือไม่
                rooms_in_use = ExamSubject.objects.filter(
                    room__building=building,
                    exam_date__gte=timezone.now().date()
                ).exists()
                
                if rooms_in_use:
                    return JsonResponse({
                        'error': 'ไม่สามารถลบอาคารได้ เนื่องจากมีห้องที่กำลังถูกใช้สำหรับการสอบ',
                        'success': False
                    }, status=400)
                
                building_name = building.name
                
                with transaction.atomic():
                    building.delete()
                
                return JsonResponse({
                    'success': True,
                    'message': f'ลบอาคาร {building_name} สำเร็จ!'
                })
                
        except TimeoutError:
            return JsonResponse({
                'error': 'ระบบกำลังประมวลผล กรุณารอสักครู่แล้วลองใหม่',
                'success': False
            }, status=429)
            
    except Exception as e:
        print(f"Error in delete_building_ajax: {str(e)}")
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)
    
@login_required
@csrf_exempt  
def delete_room_ajax(request, room_id):
    """ลบห้องสอบผ่าน AJAX - ปรับปรุงเพื่อป้องกัน concurrent deletion"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง', 'success': False}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed', 'success': False}, status=405)
    
    try:
        # ใช้ select_for_update เพื่อ lock record
        room = get_object_or_404(ExamRoom.objects.select_for_update(), id=room_id)
        
        # สร้าง lock key สำหรับการลบห้อง
        lock_key = f"room_deletion_{room_id}"
        
        try:
            with acquire_lock(lock_key, timeout=5):
                # ตรวจสอบว่าห้องกำลังถูกใช้สอบหรือไม่
                room_in_use = ExamSubject.objects.filter(
                    room=room,
                    exam_date__gte=timezone.now().date()
                ).exists()
                
                if room_in_use:
                    return JsonResponse({
                        'error': f'ไม่สามารถลบห้อง {room.name} ได้ เนื่องจากกำลังถูกใช้สำหรับการสอบ',
                        'success': False
                    }, status=400)
                
                room_name = f"{room.building.name} ห้อง {room.name}"
                
                with transaction.atomic():
                    room.delete()
                
                return JsonResponse({
                    'success': True,
                    'message': f'ลบห้อง {room_name} สำเร็จ!'
                })
                
        except TimeoutError:
            return JsonResponse({
                'error': 'ระบบกำลังประมวลผล กรุณารอสักครู่แล้วลองใหม่',
                'success': False
            }, status=429)
            
    except Exception as e:
        print(f"Error in delete_room_ajax: {str(e)}")
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

@login_required
def get_exam_subject_detail(request, subject_id):
    """ดึงรายละเอียดวิชาสอบสำหรับแสดงใน Modal"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    try:
        subject = get_object_or_404(
            ExamSubject.objects.select_related('room__building', 'invigilator__user', 'secondary_invigilator__user'),
            id=subject_id
        )
        
        data = {
            'id': subject.id,
            'subject_name': subject.subject_name,
            'subject_code': subject.subject_code,
            'academic_year': subject.academic_year,
            'term': subject.term,
            'term_display': subject.get_term_display(),
            'exam_date': subject.exam_date.strftime('%Y-%m-%d'),
            'exam_date_display': subject.exam_date.strftime('%d/%m/%Y'),
            'start_time': subject.start_time.strftime('%H:%M'),
            'end_time': subject.end_time.strftime('%H:%M'),
            'duration': subject.get_duration(),
            'student_count': subject.get_student_count(),
            'status': subject.get_status(),
            'room': None,
            'invigilator': None,
            'secondary_invigilator': None
        }
        
        if subject.room:
            data['room'] = {
                'id': subject.room.id,
                'name': subject.room.name,
                'building_name': subject.room.building.name if subject.room.building else '',
                'capacity': subject.room.capacity,
                'full_name': f"{subject.room.building.name} ห้อง {subject.room.name}" if subject.room.building else subject.room.name
            }
        
        if subject.invigilator:
            data['invigilator'] = {
                'id': subject.invigilator.id,
                'name': subject.invigilator.user.get_full_name(),
                'teacher_id': subject.invigilator.teacher_id
            }
        
        if subject.secondary_invigilator:
            data['secondary_invigilator'] = {
                'id': subject.secondary_invigilator.id,
                'name': subject.secondary_invigilator.user.get_full_name(),
                'teacher_id': subject.secondary_invigilator.teacher_id
            }
        
        return JsonResponse({
            'success': True,
            'subject': data
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}'
        }, status=500)

@login_required
def get_exam_statistics(request):
    """ดึงสถิติการสอบสำหรับแสดงในหน้า Dashboard"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    try:
        # สถิติพื้นฐาน
        total_subjects = ExamSubject.objects.count()
        total_students = sum(s.get_student_count() for s in ExamSubject.objects.all())
        total_rooms = ExamRoom.objects.filter(is_active=True).count()
        total_teachers = TeacherProfile.objects.filter(user__is_active=True).count()
        
        # สถิติการสอบตามวัน
        today = timezone.now().date()
        exams_today = ExamSubject.objects.filter(exam_date=today).count()
        exams_this_week = ExamSubject.objects.filter(
            exam_date__gte=today,
            exam_date__lt=today + timezone.timedelta(days=7)
        ).count()
        exams_this_month = ExamSubject.objects.filter(
            exam_date__year=today.year,
            exam_date__month=today.month
        ).count()
        
        # สถิติการจัดการ
        subjects_with_rooms = ExamSubject.objects.exclude(room__isnull=True).count()
        subjects_with_teachers = ExamSubject.objects.exclude(invigilator__isnull=True).count()
        complete_subjects = ExamSubject.objects.exclude(
            Q(room__isnull=True) | Q(invigilator__isnull=True)
        ).count()
        
        # สถิติตามปีการศึกษา
        year_stats = []
        years = ExamSubject.objects.values_list('academic_year', flat=True).distinct()
        for year in years:
            year_subjects = ExamSubject.objects.filter(academic_year=year)
            year_stats.append({
                'year': year,
                'subjects': year_subjects.count(),
                'students': sum(s.get_student_count() for s in year_subjects)
            })
        
        # สถิติตามเทอม
        term_stats = []
        for term_num in [1, 2, 3]:
            term_subjects = ExamSubject.objects.filter(term=str(term_num))
            term_stats.append({
                'term': term_num,
                'subjects': term_subjects.count(),
                'students': sum(s.get_student_count() for s in term_subjects)
            })
        
        return JsonResponse({
            'success': True,
            'statistics': {
                'basic': {
                    'total_subjects': total_subjects,
                    'total_students': total_students,
                    'total_rooms': total_rooms,
                    'total_teachers': total_teachers
                },
                'schedule': {
                    'exams_today': exams_today,
                    'exams_this_week': exams_this_week,
                    'exams_this_month': exams_this_month
                },
                'management': {
                    'subjects_with_rooms': subjects_with_rooms,
                    'subjects_with_teachers': subjects_with_teachers,
                    'complete_subjects': complete_subjects,
                    'completion_rate': round((complete_subjects / total_subjects * 100) if total_subjects > 0 else 0, 1)
                },
                'year_stats': year_stats,
                'term_stats': term_stats
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}'
        }, status=500)

# ========================= ระบบ QR Code =========================
@login_required
def generate_qr_code(request, pk):
    """สร้าง QR Code สำหรับเช็คชื่อ - เฉพาะเจ้าหน้าที่และครูผู้คุมสอบเท่านั้น"""
    subject = get_object_or_404(ExamSubject, id=pk)
    
    # ตรวจสอบสิทธิ์ - เฉพาะเจ้าหน้าที่และครูผู้คุมสอบเท่านั้น
    if not (request.user.is_staff or 
            (hasattr(request.user, 'teacher_profile') and 
             (subject.invigilator == request.user.teacher_profile or 
              subject.secondary_invigilator == request.user.teacher_profile))):
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึง QR Code นี้")
    
    # สร้าง QR URL
    qr_url = request.build_absolute_uri(f"/checkin/{pk}/")
    
    # สร้าง QR Code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_url)
    qr.make(fit=True)
    
    # สร้างรูปภาพ QR Code
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    img_base64 = base64.b64encode(buffer.getvalue()).decode()
    
    # ดึงข้อมูลสถิติการเข้าสอบ
    attendance_stats = {
        'total_students': subject.get_student_count(),
        'checked_in': Attendance.objects.filter(subject=subject).count(),
        'teacher_checked_in': 0
    }
    
    if subject.invigilator_checkin:
        attendance_stats['teacher_checked_in'] += 1
    if subject.secondary_invigilator_checkin:
        attendance_stats['teacher_checked_in'] += 1
    
    return render(request, 'app/staff/qr_code.html', {
        'subject': subject,
        'qr_image': img_base64,
        'qr_url': qr_url,
        'today': timezone.now().date(),
        'attendance_stats': attendance_stats,
    })

@csrf_exempt
def checkin_exam(request, pk):
    """เช็คชื่อเข้าสอบผ่าน QR Code - ปรับปรุงใหม่"""
    subject = get_object_or_404(ExamSubject, id=pk)
    
    if request.method == 'GET':
        # แสดงหน้าเช็คชื่อ
        if not request.user.is_authenticated:
            # ถ้ายังไม่ login ให้ไปหน้า login ก่อน
            from django.urls import reverse
            login_url = f"{reverse('index_view')}?next={request.path}"
            return HttpResponseRedirect(login_url)
        
        user_type = None
        profile = None
        can_checkin = False
        seat_number = None  # เพิ่มตัวแปรเลขที่นั่ง
        
        if request.user.is_student:
            try:
                profile = request.user.student_profile
                # ตรวจสอบว่านักเรียนคนนี้ลงทะเบียนในวิชานี้หรือไม่
                if subject.students.filter(id=profile.id).exists():
                    user_type = 'student'
                    can_checkin = True
                    seat_number = profile.student_number  # ใช้หมายเลขนักเรียนเป็นเลขที่นั่ง
            except:
                pass
                
        elif request.user.is_teacher:
            try:
                profile = request.user.teacher_profile
                # ตรวจสอบว่าเป็นครูผู้คุมสอบหรือไม่
                if (subject.invigilator == profile or subject.secondary_invigilator == profile):
                    user_type = 'teacher'
                    can_checkin = True
            except:
                pass
        
        elif request.user.is_staff:
            user_type = 'staff'
            can_checkin = True
        
        if not can_checkin:
            return render(request, 'app/staff/unauthorized_checkin.html', {
                'subject': subject,
                'user': request.user,
                'message': 'คุณไม่มีสิทธิ์เช็คชื่อในวิชานี้'
            })
        
        # ตรวจสอบสถานะการเช็คชื่อ
        already_checked = False
        checkin_time = None
        
        if user_type == 'student':
            attendance = Attendance.objects.filter(student=profile, subject=subject).first()
            if attendance:
                already_checked = True
                checkin_time = attendance.checkin_time
        elif user_type == 'teacher':
            if profile == subject.invigilator and subject.invigilator_checkin:
                already_checked = True
                checkin_time = subject.invigilator_checkin_time
            elif profile == subject.secondary_invigilator and subject.secondary_invigilator_checkin:
                already_checked = True
                checkin_time = subject.secondary_invigilator_checkin_time
        
        return render(request, 'app/staff/checkin_page.html', {
            'subject': subject,
            'user_type': user_type,
            'profile': profile,
            'can_checkin': can_checkin,
            'already_checked': already_checked,
            'checkin_time': checkin_time,
            'seat_number': seat_number,  # เพิ่มเลขที่นั่ง
        })
    
    elif request.method == 'POST':
        # ดำเนินการเช็คชื่อ (ส่วนที่แก้ไข)
        if not request.user.is_authenticated:
            return JsonResponse({'status': 'error', 'message': 'กรุณาเข้าสู่ระบบ'})
        
        current_time = timezone.now()
        
        # ตรวจสอบว่าเป็นเวลาสอบหรือไม่
        exam_datetime_start = timezone.datetime.combine(subject.exam_date, subject.start_time)
        exam_datetime_end = timezone.datetime.combine(subject.exam_date, subject.end_time)
        exam_datetime_start = timezone.make_aware(exam_datetime_start)
        exam_datetime_end = timezone.make_aware(exam_datetime_end)
        
        # กำหนดช่วงเวลา: เริ่มสอบ ถึง สิ้นสุดสอบเท่านั้น
        late_threshold = exam_datetime_start + timezone.timedelta(minutes=30)  # หลังเริ่มสอบ 30 นาที = สาย
        
        # ตรวจสอบช่วงเวลาที่สามารถเช็คชื่อได้
        if current_time < exam_datetime_start:
            return JsonResponse({
                'status': 'error',
                'message': f'ยังไม่ถึงเวลาสอบ\nการสอบจะเริ่มเวลา {exam_datetime_start.strftime("%H:%M")} และสิ้นสุดเวลา {exam_datetime_end.strftime("%H:%M")}'
            })
        
        if current_time > exam_datetime_end:
            return JsonResponse({
                'status': 'error',
                'message': f'หมดเวลาสอบแล้ว\nการสอบสิ้นสุดเวลา {exam_datetime_end.strftime("%H:%M")}'
            })
        
        if request.user.is_student:
            try:
                student = request.user.student_profile
                
                # ตรวจสอบว่าลงทะเบียนในวิชานี้หรือไม่
                if not subject.students.filter(id=student.id).exists():
                    return JsonResponse({'status': 'error', 'message': 'คุณไม่ได้ลงทะเบียนในวิชานี้'})
                
                # ตรวจสอบว่าเช็คชื่อแล้วหรือไม่
                attendance, created = Attendance.objects.get_or_create(
                    student=student, 
                    subject=subject,
                    defaults={
                        'checkin_time': current_time,
                        # กำหนดสถานะตามเวลาที่เช็คชื่อเมื่อเทียบกับเวลาที่การสอบควรจะสิ้นสุดช่วง 30 นาทีแรก
                        'status': 'late' if current_time > late_threshold else 'on_time'
                    }
                )
                
                if not created:
                    return JsonResponse({
                        'status': 'info', 
                        'message': f'คุณได้เช็คชื่อไปแล้วเมื่อ {attendance.checkin_time.strftime("%H:%M")} น.',
                        'checkin_time': attendance.checkin_time.strftime("%H:%M")
                    })
                
                return JsonResponse({
                    'status': 'success',
                    'message': 'เช็คชื่อสำเร็จ!',
                    'time': current_time.strftime('%H:%M'),
                    'status_text': 'ตรงเวลา' if attendance.status == 'on_time' else 'สาย',
                    'timestamp': current_time.isoformat()  # เพิ่ม timestamp แบบเต็ม
                })
                
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'เกิดข้อผิดพลาด: {str(e)}'})
        
        elif request.user.is_teacher:
            try:
                teacher = request.user.teacher_profile
                
                if teacher == subject.invigilator:
                    if subject.invigilator_checkin:
                        return JsonResponse({
                            'status': 'info', 
                            'message': f'คุณได้เช็คชื่อไปแล้วเมื่อ {subject.invigilator_checkin_time.strftime("%H:%M")} น.',
                            'checkin_time': subject.invigilator_checkin_time.strftime("%H:%M")
                        })
                    
                    subject.invigilator_checkin = True
                    subject.invigilator_checkin_time = current_time
                    subject.save()
                    
                    return JsonResponse({
                        'status': 'success', 
                        'message': 'เช็คชื่อครูหลักสำเร็จ!',
                        'time': current_time.strftime('%H:%M')
                    })
                
                elif teacher == subject.secondary_invigilator:
                    if subject.secondary_invigilator_checkin:
                        return JsonResponse({
                            'status': 'info', 
                            'message': f'คุณได้เช็คชื่อไปแล้วเมื่อ {subject.secondary_invigilator_checkin_time.strftime("%H:%M")} น.',
                            'checkin_time': subject.secondary_invigilator_checkin_time.strftime("%H:%M")
                        })
                    
                    subject.secondary_invigilator_checkin = True
                    subject.secondary_invigilator_checkin_time = current_time
                    subject.save()
                    
                    return JsonResponse({
                        'status': 'success', 
                        'message': 'เช็คชื่อครูสำรองสำเร็จ!',
                        'time': current_time.strftime('%H:%M')
                    })
                
                else:
                    return JsonResponse({'status': 'error', 'message': 'คุณไม่ใช่ครูผู้คุมสอบของวิชานี้'})
                    
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'เกิดข้อผิดพลาด: {str(e)}'})
        
        return JsonResponse({'status': 'error', 'message': 'ไม่พบบทบาททีถูกต้อง'})

@csrf_exempt
def get_server_time(request):
    """API สำหรับดึงเวลาจากเซิร์ฟเวอร์เพื่อซิงค์กับ client"""
    if request.method == 'GET':
        current_time = timezone.now()
        return JsonResponse({
            'server_time': current_time.isoformat(),
            'timestamp': current_time.timestamp(),
            'timezone': str(current_time.tzinfo),
            'formatted_time': current_time.strftime('%H:%M:%S'),
            'formatted_date': current_time.strftime('%Y-%m-%d')
        })
    return JsonResponse({'error': 'Method not allowed'}, status=405)

# ========================= รายงานสถานะการสอบบ =========================
@login_required
def exam_attendance(request, pk):
    """ดูสถานะการเข้าสอบ - แก้ไขเพื่อรองรับสถานะใหม่"""
    subject = get_object_or_404(ExamSubject, id=pk)
    
    # ตรวจสอบสิทธิ์
    if not (request.user.is_staff or 
            (hasattr(request.user, 'teacher_profile') and 
             subject.invigilator == request.user.teacher_profile)):
        return HttpResponseForbidden("คุณไม่มีสิทธิ์ดูข้อมูลนี้")
    
    
    students = subject.students.all().order_by('student_class', 'student_number', 'user__first_name')
    attendance_records = Attendance.objects.filter(subject=subject)
    
    # สร้าง attendance_dict ที่ถูกต้อง - ใช้ student.id เป็น key
    attendance_dict = {att.student.id: att for att in attendance_records}
    
    # เพิ่ม attendance_record และ calculated_status attribute ให้กับแต่ละ student object
    for student in students:
        student.attendance_record = attendance_dict.get(student.id)
        student.calculated_status = get_student_status(student, student.attendance_record, subject)
    
    # สถิติการเข้าสอบ - แก้ไขให้นับสถานะที่คำนวณแล้ว
    stats = calculate_exam_stats(students, attendance_records, subject)
    
    # สถิติครู
    teacher_stats = {
        'total_teachers': 0,
        'checked_in': 0
    }
    
    if subject.invigilator:
        teacher_stats['total_teachers'] += 1
        if subject.invigilator_checkin:
            teacher_stats['checked_in'] += 1
    
    if subject.secondary_invigilator:
        teacher_stats['total_teachers'] += 1
        if subject.secondary_invigilator_checkin:
            teacher_stats['checked_in'] += 1
    
    # ตรวจสอบว่าการสอบเสร็จสิ้นแล้วหรือไม่
    now = timezone.now()
    exam_completed = (subject.exam_date < now.date() or 
                     (subject.exam_date == now.date() and subject.end_time < now.time()))
    
    return render(request, 'app/staff/exam_attendance.html', {
        'subject': subject,
        'students': students,
        'attendance_dict': attendance_dict,
        'stats': stats,
        'teacher_stats': teacher_stats,
        'exam_completed': exam_completed,
    })

@login_required
def get_exam_stats(request, subject_id):
    """AJAX endpoint สำหรับดึงสถิติการสอบ"""
    subject = get_object_or_404(ExamSubject, id=subject_id)

    students = subject.students.all()
    attendance_records = Attendance.objects.filter(subject=subject)

    stats = {
        'total_students': students.count(),
        'on_time': attendance_records.filter(status='on_time').count(),
        'late': attendance_records.filter(status='late').count(),
        'absent': attendance_records.filter(status='absent').count(),
        'excused': attendance_records.filter(status='excused').count(),
        'cheating': attendance_records.filter(status='cheating').count(),
    }

    return JsonResponse({'success': True, 'stats': stats})  

@login_required
def export_attendance_report(request, subject_id):
    """ส่งออกรายงานการเข้าสอบเป็น Excel - รายงานสมบูรณ์"""
    if not request.user.is_staff:
        return HttpResponseForbidden("ไม่มีสิทธิ์เข้าถึง")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
        
        subject = get_object_or_404(ExamSubject, id=subject_id)
        students = subject.students.all().order_by('student_class', 'student_number', 'user__first_name')
        attendance_records = Attendance.objects.filter(subject=subject)
        
        # สร้าง workbook และ worksheets
        wb = openpyxl.Workbook()
        
        # ============ Sheet 1: รายละเอียดการเข้าสอบ ============
        ws_detail = wb.active
        ws_detail.title = "รายละเอียดการเข้าสอบ"
        
        # สไตล์
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # หัวข้อรายงาน
        ws_detail.merge_cells('A1:J3')
        title_cell = ws_detail['A1']
        title_cell.value = f"รายงานการเข้าสอบ\n{subject.subject_name} ({subject.subject_code})\nวันที่ {subject.exam_date.strftime('%d/%m/%Y')} เวลา {subject.start_time.strftime('%H:%M')}-{subject.end_time.strftime('%H:%M')}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = center_alignment
        title_cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # ข้อมูลรายวิชา
        info_start_row = 5
        info_data = [
            ['รหัสวิชา:', subject.subject_code],
            ['ชื่อวิชา:', subject.subject_name],
            ['ปีการศึกษา:', subject.academic_year],
            ['เทอม:', subject.get_term_display()],
            ['วันที่สอบ:', subject.exam_date.strftime('%d/%m/%Y')],
            ['เวลาสอบ:', f"{subject.start_time.strftime('%H:%M')} - {subject.end_time.strftime('%H:%M')}"],
            ['ห้องสอบ:', f"{subject.room.building.name} ห้อง {subject.room.name}" if subject.room else "ไม่ระบุ"],
            ['ครูคุมสอบหลัก:', subject.invigilator.user.get_full_name() if subject.invigilator else "ไม่ระบุ"],
            ['ครูคุมสอบสำรอง:', subject.secondary_invigilator.user.get_full_name() if subject.secondary_invigilator else "ไม่ระบุ"],
        ]
        
        for i, (label, value) in enumerate(info_data):
            row = info_start_row + i
            ws_detail[f'A{row}'] = label
            ws_detail[f'B{row}'] = value
            ws_detail[f'A{row}'].font = Font(bold=True)
        
        # สถิติสรุป
        stats_start_row = info_start_row + len(info_data) + 2
        total_students = students.count()
        on_time_count = attendance_records.filter(status='on_time').count()
        late_count = attendance_records.filter(status='late').count()
        absent_count = attendance_records.filter(status='absent').count()
        excused_count = attendance_records.filter(status='excused').count()
        cheating_count = attendance_records.filter(status='cheating').count()
        
        # คำนวณเปอร์เซ็นต์
        attendance_rate = ((on_time_count + late_count) / total_students * 100) if total_students > 0 else 0
        
        ws_detail[f'A{stats_start_row}'] = "สรุปสถิติการเข้าสอบ"
        ws_detail[f'A{stats_start_row}'].font = Font(bold=True, size=14)
        
        stats_data = [
            ['จำนวนนักเรียนทั้งหมด:', total_students],
            ['เข้าสอบตรงเวลา:', f"{on_time_count} คน ({on_time_count/total_students*100:.1f}%)" if total_students > 0 else "0 คน"],
            ['เข้าสอบสาย:', f"{late_count} คน ({late_count/total_students*100:.1f}%)" if total_students > 0 else "0 คน"],
            ['ขาดสอบ:', f"{absent_count} คน ({absent_count/total_students*100:.1f}%)" if total_students > 0 else "0 คน"],
            ['ลาป่วย:', f"{excused_count} คน ({excused_count/total_students*100:.1f}%)" if total_students > 0 else "0 คน"],
            ['ทุจริต:', f"{cheating_count} คน ({cheating_count/total_students*100:.1f}%)" if total_students > 0 else "0 คน"],
            ['อัตราเข้าสอบรวม:', f"{attendance_rate:.1f}%"],
        ]
        
        for i, (label, value) in enumerate(stats_data):
            row = stats_start_row + i + 1
            ws_detail[f'A{row}'] = label
            ws_detail[f'B{row}'] = value
            ws_detail[f'A{row}'].font = Font(bold=True)
        
        # ตารางรายชื่อนักเรียน
        table_start_row = stats_start_row + len(stats_data) + 3
        headers = ['ลำดับ', 'รหัสนักเรียน', 'ชื่อ-นามสกุล', 'ห้อง', 'เลขที่', 'สถานะ', 'เวลาเช็คชื่อ', 'หมายเหตุ', 'แก้ไขโดย', 'เวลาแก้ไข']
        
        # สร้างหัวตาราง
        for col, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=table_start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # สร้าง dictionary สำหรับ attendance
        attendance_dict = {att.student.id: att for att in attendance_records}
        
        # เติมข้อมูลนักเรียน
        for i, student in enumerate(students, 1):
            row = table_start_row + i
            attendance = attendance_dict.get(student.id)
            
            # แปลงสถานะเป็นภาษาไทย
            status_map = {
                'on_time': 'เข้าสอบตรงเวลา',
                'late': 'เข้าสอบสาย',
                'absent': 'ขาดสอบ',
                'excused': 'ลาป่วย',
                'cheating': 'ทุจริต'
            }
            
            status = status_map.get(attendance.status if attendance else 'absent', 'ขาดสอบ')
            checkin_time = attendance.checkin_time.strftime('%H:%M') if attendance and attendance.checkin_time else '-'
            
            # หมายเหตุ
            notes = []
            if attendance and attendance.manually_updated:
                notes.append('แก้ไขด้วยตนเอง')
            if attendance and attendance.status == 'cheating':
                notes.append('พบการทุจริต')
            
            # ข้อมูลแต่ละคอลัมน์
            row_data = [
                i,  # ลำดับ
                student.student_id,
                student.user.get_full_name(),
                student.student_class,
                student.student_number,
                status,
                checkin_time,
                '; '.join(notes) if notes else '-',
                attendance.last_modified_by.get_full_name() if attendance and attendance.last_modified_by else '-',
                attendance.updated_at.strftime('%d/%m/%Y %H:%M') if attendance and attendance.manually_updated else '-'
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws_detail.cell(row=row, column=col, value=value)
                cell.border = border
                
                # จัดตำแหน่ง
                if col in [1, 4, 5]:  # ลำดับ, เลขที่
                    cell.alignment = center_alignment
                
                # สีตามสถานะ
                if col == 6:  # คอลัมน์สถานะ
                    if status == 'เข้าสอบตรงเวลา':
                        cell.fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")
                    elif status == 'เข้าสอบสาย':
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    elif status == 'ขาดสอบ':
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif status == 'ลาป่วย':
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    elif status == 'ทุจริต':
                        cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
        
        # ปรับความกว้างคอลัมน์
        column_widths = [8, 15, 25, 12, 8, 18, 12, 20, 18, 18]
        for col, width in enumerate(column_widths, 1):
            ws_detail.column_dimensions[get_column_letter(col)].width = width
        
        # ============ Sheet 2: สรุปสถิติ ============
        ws_stats = wb.create_sheet(title="สรุปสถิติ")
        
        # หัวข้อ
        ws_stats['A1'] = f"สรุปสถิติการเข้าสอบ - {subject.subject_name}"
        ws_stats['A1'].font = Font(bold=True, size=16)
        ws_stats.merge_cells('A1:C1')
        
        # ตารางสถิติ
        stats_headers = ['สถานะ', 'จำนวน (คน)', 'เปอร์เซ็นต์']
        stats_rows = [
            ['เข้าสอบตรงเวลา', on_time_count, f"{on_time_count/total_students*100:.1f}%" if total_students > 0 else "0%"],
            ['เข้าสอบสาย', late_count, f"{late_count/total_students*100:.1f}%" if total_students > 0 else "0%"],
            ['ขาดสอบ', absent_count, f"{absent_count/total_students*100:.1f}%" if total_students > 0 else "0%"],
            ['ลาป่วย', excused_count, f"{excused_count/total_students*100:.1f}%" if total_students > 0 else "0%"],
            ['ทุจริต', cheating_count, f"{cheating_count/total_students*100:.1f}%" if total_students > 0 else "0%"],
            ['รวม', total_students, '100.0%']
        ]
        
        # หัวตาราง
        for col, header in enumerate(stats_headers, 1):
            cell = ws_stats.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # ข้อมูลตาราง
        for row_idx, row_data in enumerate(stats_rows, 4):
            for col, value in enumerate(row_data, 1):
                cell = ws_stats.cell(row=row_idx, column=col, value=value)
                cell.border = border
                if col == 2:  # คอลัมน์จำนวน
                    cell.alignment = center_alignment
                elif col == 3:  # คอลัมน์เปอร์เซ็นต์
                    cell.alignment = center_alignment
                
                # สีพื้นหลังสำหรับแถวรวม
                if row_idx == len(stats_rows) + 3:
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    cell.font = Font(bold=True)
        
        # สร้างกราฟ
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "สถิติการเข้าสอบ"
        chart.y_axis.title = 'จำนวน (คน)'
        chart.x_axis.title = 'สถานะ'
        
        # ข้อมูลสำหรับกราฟ (ไม่รวมแถวรวม)
        data = Reference(ws_stats, min_col=2, min_row=3, max_row=len(stats_rows) + 2, max_col=2)
        cats = Reference(ws_stats, min_col=1, min_row=4, max_row=len(stats_rows) + 2)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # เพิ่มกราฟใน sheet
        ws_stats.add_chart(chart, "E3")
        
        # ปรับความกว้างคอลัมน์
        ws_stats.column_dimensions['A'].width = 20
        ws_stats.column_dimensions['B'].width = 15
        ws_stats.column_dimensions['C'].width = 15
        
        # ============ Sheet 3: ครูคุมสอบ ============
        ws_teachers = wb.create_sheet(title="ครูคุมสอบ")
        
        ws_teachers['A1'] = f"รายงานครูคุมสอบ - {subject.subject_name}"
        ws_teachers['A1'].font = Font(bold=True, size=16)
        ws_teachers.merge_cells('A1:E1')
        
        teacher_headers = ['ตำแหน่ง', 'รหัสครู', 'ชื่อ-นามสกุล', 'สถานะ', 'เวลาเข้าหน้าที่']
        
        # หัวตาราง
        for col, header in enumerate(teacher_headers, 1):
            cell = ws_teachers.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # ข้อมูลครู
        teacher_data = []
        if subject.invigilator:
            teacher_data.append([
                'ครูหลัก',
                subject.invigilator.teacher_id,
                subject.invigilator.user.get_full_name(),
                'เข้าหน้าที่แล้ว' if subject.invigilator_checkin else 'ยังไม่เข้าหน้าที่',
                subject.invigilator_checkin_time.strftime('%H:%M') if subject.invigilator_checkin and subject.invigilator_checkin_time else '-'
            ])
        
        if subject.secondary_invigilator:
            teacher_data.append([
                'ครูสำรอง',
                subject.secondary_invigilator.teacher_id,
                subject.secondary_invigilator.user.get_full_name(),
                'เข้าหน้าที่แล้ว' if subject.secondary_invigilator_checkin else 'ยังไม่เข้าหน้าที่',
                subject.secondary_invigilator_checkin_time.strftime('%H:%M') if subject.secondary_invigilator_checkin and subject.secondary_invigilator_checkin_time else '-'
            ])
        
        for i, row_data in enumerate(teacher_data, 4):
            for col, value in enumerate(row_data, 1):
                cell = ws_teachers.cell(row=i, column=col, value=value)
                cell.border = border
                if col == 4:  # คอลัมน์สถานะ
                    if 'เข้าหน้าที่แล้ว' in value:
                        cell.fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # ปรับความกว้างคอลัมน์
        teacher_column_widths = [12, 12, 25, 18, 15]
        for col, width in enumerate(teacher_column_widths, 1):
            ws_teachers.column_dimensions[get_column_letter(col)].width = width
        
        # ข้อมูลเพิ่มเติม (footer ในทุก sheet)
        footer_text = f"รายงานสร้างเมื่อ: {timezone.now().strftime('%d/%m/%Y %H:%M')} โดย: {request.user.get_full_name()}"
        
        for ws in [ws_detail, ws_stats, ws_teachers]:
            last_row = ws.max_row + 2
            ws[f'A{last_row}'] = footer_text
            ws[f'A{last_row}'].font = Font(italic=True, size=10, color="666666")
        
        # สร้าง response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f'รายงานการเข้าสอบ_{subject.subject_code}_{subject.exam_date.strftime("%Y%m%d")}_{timezone.now().strftime("%H%M")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'เกิดข้อผิดพลาดในการสร้างรายงาน: {str(e)}')
        return redirect('exam_attendance', pk=subject_id)

@login_required
@csrf_exempt
def teacher_manual_checkin(request):
    """เช็คชื่อครูด้วยตนเอง (รองรับ check และ uncheck)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    data = json.loads(request.body or '{}')
    subject_id = data.get('subject_id')
    teacher_type = data.get('teacher_type')  # 'primary' หรือ 'secondary'
    present = data.get('present', True)      # True = เช็ค, False = ยกเลิก

    subject = get_object_or_404(ExamSubject, id=subject_id)

    # สิทธิ์: staff หรือ ครูคุมสอบที่ตรงกับ teacher_type
    is_owner_teacher = False
    if hasattr(request.user, 'teacher_profile'):
        if teacher_type == 'primary' and subject.invigilator == request.user.teacher_profile:
            is_owner_teacher = True
        if teacher_type == 'secondary' and subject.secondary_invigilator == request.user.teacher_profile:
            is_owner_teacher = True

    if not (request.user.is_staff or is_owner_teacher):
        return JsonResponse({'error': 'ไม่มีสิทธิ์'}, status=403)

    now = timezone.now()
    if teacher_type == 'primary':
        subject.invigilator_checkin = bool(present)
        subject.invigilator_checkin_time = now if present else None
    elif teacher_type == 'secondary':
        subject.secondary_invigilator_checkin = bool(present)
        subject.secondary_invigilator_checkin_time = now if present else None
    else:
        return JsonResponse({'error': 'teacher_type ไม่ถูกต้อง'}, status=400)

    subject.save()
    return JsonResponse({
        'success': True,
        'present': bool(present),
        'time': now.strftime('%H:%M') if present else '-'
    })

@login_required 
@csrf_exempt
def manual_checkin_student(request):
    """เช็คชื่อ/แก้สถานะ นักเรียนแบบ manual (รองรับ real-time)"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body or '{}')
        student_id = data.get('student_id')
        subject_id = data.get('subject_id')
        status = data.get('status', 'on_time')

        student = get_object_or_404(StudentProfile, id=student_id)
        subject = get_object_or_404(ExamSubject, id=subject_id)

        # สิทธิ์: staff หรือ ครูผู้คุมสอบของวิชานี้ (หลัก/สำรอง)
        is_invigilator = (
            hasattr(request.user, 'teacher_profile') and (
                subject.invigilator == request.user.teacher_profile or
                subject.secondary_invigilator == request.user.teacher_profile
            )
        )
        if not (request.user.is_staff or is_invigilator):
            return JsonResponse({'status': 'error', 'message': 'ไม่มีสิทธิ์'}, status=403)

        # นักเรียนต้องลงทะเบียนในวิชานี้
        if not subject.students.filter(id=student.id).exists():
            return JsonResponse({'status': 'error', 'message': 'นักเรียนไม่ได้ลงทะเบียนในวิชานี้'})

        now = timezone.now()
        attendance, created = Attendance.objects.get_or_create(
            student=student, subject=subject,
            defaults={'status': status, 'checkin_time': now}
        )
        if not created:
            attendance.status = status
            attendance.checkin_time = now

        # ถ้า model มี field บันทึกว่าแก้มือและผู้แก้ (โค้ดนี้เช็คแบบป้องกันไว้)
        if hasattr(attendance, 'manually_updated'):
            attendance.manually_updated = True
        if hasattr(attendance, 'updated_by'):
            attendance.updated_by = request.user

        attendance.save()

        return JsonResponse({
            'status': 'success',
            'attendance': {
                'student_id': student.id,
                'status': attendance.status,
                'checkin_time': attendance.checkin_time.strftime('%H:%M'),
            }
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'เกิดข้อผิดพลาด: {e}'}, status=500)

# ========================= ผังที่นั่งสอบบ =========================
@login_required
def exam_seating_data(request, subject_id):
    """AJAX endpoint สำหรับดึงข้อมูลอัปเดต seating chart - real time แก้ไขใหม่"""
    subject = get_object_or_404(ExamSubject, id=subject_id)
    
    # ตรวจสอบสิทธิ์
    if not (request.user.is_staff or 
            (hasattr(request.user, 'teacher_profile') and 
             (subject.invigilator == request.user.teacher_profile or 
              subject.secondary_invigilator == request.user.teacher_profile))):
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    try:
        students = subject.students.all().order_by('student_class', 'student_number', 'user__first_name')
        attendance_records = Attendance.objects.filter(subject=subject)
        
        # จัดรูปแบบข้อมูล attendance พร้อมสถานะที่คำนวณแล้ว
        attendance_data = {}
        for student in students:
            attendance = attendance_records.filter(student=student).first()
            status = get_student_status(student, attendance, subject)
            
            attendance_data[str(student.id)] = {
                'status': status,
                'checkin_time': attendance.checkin_time.strftime('%H:%M') if attendance and attendance.checkin_time else None,
                'created_at': attendance.checkin_time.isoformat() if attendance and attendance.checkin_time else None,
                'manually_updated': getattr(attendance, 'manually_updated', False) if attendance else False
            }
        
        # คำนวณสถิติ - ใช้ฟังก์ชันใหม่
        stats = calculate_exam_stats_new(students, attendance_records, subject)
        
        return JsonResponse({
            'success': True,
            'attendance': attendance_data,
            'stats': stats,
            'last_updated': timezone.now().isoformat()
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

@login_required
def exam_seating_view(request, pk):
    """หน้าแสดงการจัดที่นั่งแบบโรงหนัง - แก้ไขเพื่อจัดเรียงตามเลขที่และสถานะใหม่"""
    subject = get_object_or_404(ExamSubject, id=pk)
    
    # ตรวจสอบสิทธิ์
    if not (request.user.is_staff or 
            (hasattr(request.user, 'teacher_profile') and 
             (subject.invigilator == request.user.teacher_profile or 
              subject.secondary_invigilator == request.user.teacher_profile))):
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึงส่วนนี้")
    
    # เรียงตามห้องแล้วตามเลขที่
    students = subject.students.all().order_by('student_class', 'student_number', 'user__first_name')
    attendance_records = Attendance.objects.filter(subject=subject)
    attendance_dict = {att.student.id: att for att in attendance_records}
    
    # จัดเรียงที่นั่งตามเลขที่นักเรียน - แก้ไขใหม่
    seating_chart = None
    if subject.room and students.exists():
        # เรียงลำดับนักเรียนตามเลขที่ (student_number)
        sorted_students = sorted(students, key=lambda s: int(s.student_number) if s.student_number and s.student_number.isdigit() else 999)
        
        # สมมติแถวละ 6 ที่นั่ง (ปรับได้ตามความต้องการ)
        seats_per_row = 6
        students_count = len(sorted_students)
        rows = max(1, (students_count + seats_per_row - 1) // seats_per_row)
        
        # สร้าง seating chart โดยเรียงตามลำดับเลขที่
        seating_chart = []
        
        for row in range(rows):
            row_seats = []
            for seat in range(seats_per_row):
                seat_index = row * seats_per_row + seat
                if seat_index < len(sorted_students):
                    student = sorted_students[seat_index]
                    attendance = attendance_dict.get(student.id)
                    
                    # กำหนดสถานะตามเงื่อนไขเวลาใหม่
                    status = get_student_status(student, attendance, subject)
                    
                    row_seats.append({
                        'student': student,
                        'attendance': attendance,
                        'status': status,  # สถานะที่คำนวณแล้ว
                        'seat_number': int(student.student_number) if student.student_number and student.student_number.isdigit() else seat_index + 1
                    })
                else:
                    # ที่นั่งว่าง
                    row_seats.append(None)
            
            if row_seats:
                seating_chart.append(row_seats)
    
    # สถิติการเข้าสอบ - ใช้ฟังก์ชันใหม่
    stats = calculate_exam_stats(students, attendance_records, subject)
    
    # ตรวจสอบว่าการสอบเสร็จสิ้นแล้วหรือไม่
    now = timezone.now()
    exam_completed = (subject.exam_date < now.date() or 
                     (subject.exam_date == now.date() and subject.end_time < now.time()))
    
    return render(request, 'app/staff/exam_seating.html', {
        'subject': subject,
        'students': students,
        'attendance_dict': attendance_dict,
        'seating_chart': seating_chart,
        'stats': stats,
        'exam_completed': exam_completed,
    })

def calculate_exam_stats(students, attendance_records, subject):
    """คำนวณสถิติการสอบแบบใหม่ - รองรับสถานะที่คำนวณตามเวลา"""
    total_students = students.count()
    
    # นับตามสถานะจริงที่คำนวณแล้ว
    status_counts = {
        'on_time': 0,
        'late': 0,
        'absent': 0,
        'excused': 0,
        'cheating': 0,
        'not_checked': 0,
        'not_started': 0
    }
    
    # นับสถานะของแต่ละนักเรียน
    for student in students:
        attendance = attendance_records.filter(student=student).first()
        status = get_student_status(student, attendance, subject)
        
        if status in status_counts:
            status_counts[status] += 1
        else:
            # กรณีสถานะไม่รู้จัก ให้ถือเป็น not_checked
            status_counts['not_checked'] += 1
    
    # คำนวณเปอร์เซ็นต์
    stats = {
        'total': total_students,
        'on_time': status_counts['on_time'],
        'late': status_counts['late'],
        'absent': status_counts['absent'],
        'excused': status_counts['excused'], 
        'cheating': status_counts['cheating'],
        'not_checked': status_counts['not_checked'],
        'not_started': status_counts['not_started'],
        'checked_in': status_counts['on_time'] + status_counts['late'] + status_counts['excused'] + status_counts['cheating'],
        'present': status_counts['on_time'] + status_counts['late'],  # เข้าสอบรวม
    }
    
    # คำนวณเปอร์เซ็นต์
    if total_students > 0:
        stats.update({
            'on_time_percentage': (status_counts['on_time'] / total_students) * 100,
            'late_percentage': (status_counts['late'] / total_students) * 100,
            'absent_percentage': (status_counts['absent'] / total_students) * 100,
            'excused_percentage': (status_counts['excused'] / total_students) * 100,
            'cheating_percentage': (status_counts['cheating'] / total_students) * 100,
            'not_checked_percentage': (status_counts['not_checked'] / total_students) * 100,
            'attendance_rate': ((status_counts['on_time'] + status_counts['late']) / total_students) * 100
        })
    else:
        stats.update({
            'on_time_percentage': 0,
            'late_percentage': 0,
            'absent_percentage': 0,
            'excused_percentage': 0,
            'cheating_percentage': 0,
            'not_checked_percentage': 0,
            'attendance_rate': 0
        })
    
    return stats

def get_student_status(student, attendance, subject):
    """คำนวณสถานะของนักเรียนตามเงื่อนไขเวลา - แก้ไขใหม่"""
    now = timezone.now()
    
    # สร้าง datetime objects สำหรับเปรียบเทียบ
    exam_datetime_start = timezone.datetime.combine(subject.exam_date, subject.start_time)
    exam_datetime_end = timezone.datetime.combine(subject.exam_date, subject.end_time)
    exam_datetime_start = timezone.make_aware(exam_datetime_start)
    exam_datetime_end = timezone.make_aware(exam_datetime_end)
    
    # ถ้าไม่มี attendance record
    if not attendance:
        # ตรวจสอบเวลา
        if now < exam_datetime_start:
            return 'not_started'  # การสอบยังไม่เริ่ม
        elif now > exam_datetime_end:
            return 'absent'  # หมดเวลาสอบแล้ว = ขาดสอบทันที
        else:
            return 'not_checked'  # อยู่ในช่วงเวลาสอบแต่ยังไม่เช็คชื่อ
    
    # ถ้ามี attendance record ให้ใช้สถานะที่บันทึกไว้
    return attendance.status

# ========================= AJAX Endpoints =========================
@login_required
@user_passes_test(lambda u: u.is_staff)
@require_POST
def ajax_buildings_add(request):
    """
    เพิ่ม (หรืออัปเดตชื่อ/รายละเอียด ถ้ามี code เดิม) อาคารผ่าน AJAX
    รับค่าแบบ form-data: code, name, description (optional)
    คืนค่า JSON: {success, message, building}
    """
    code = (request.POST.get('code') or '').strip().upper()
    name = (request.POST.get('name') or '').strip()
    description = (request.POST.get('description') or '').strip()

    if not code or not name:
        return JsonResponse(
            {"success": False, "error": "กรุณากรอกรหัสอาคารและชื่ออาคาร"},
            status=400
        )

    try:
        with transaction.atomic():
            obj, created = Building.objects.get_or_create(
                code=code,
                defaults={"name": name, "description": description}
            )
            if not created:
                # อัปเดตเฉพาะฟิลด์ที่อนุญาต
                obj.name = name
                obj.description = description
                obj.save(update_fields=["name", "description", "updated_at"])

        return JsonResponse({
            "success": True,
            "message": "เพิ่มอาคารสำเร็จ" if created else "อัปเดตอาคารสำเร็จ",
            "building": {
                "id": obj.id,
                "code": obj.code,
                "name": obj.name,
                "description": obj.description or "",
            }
        })

    except IntegrityError as e:
        # เผื่อชน unique code (แม้เราใช้ get_or_create แล้ว โอกาสน้อย)
        return JsonResponse(
            {"success": False, "error": "รหัสอาคารซ้ำในระบบ"},
            status=400
        )
    except Exception as e:
        # กันพังเป็น 500 โดยส่งข้อความอ่านง่ายกลับไป
        return JsonResponse(
            {"success": False, "error": f"เกิดข้อผิดพลาด: {str(e)}"},
            status=500
        )

@login_required
@user_passes_test(lambda u: u.is_staff)
@require_GET
def ajax_buildings(request):
    """
    คืนรายการอาคาร + ห้อง (ตรงกับที่ JS ฝั่งหน้า manage_rooms เรียกใช้)
    คืน JSON: {success, buildings: [{id, code, name, description, room_count, total_capacity, rooms: [...] }]}
    """
    items = []
    buildings = Building.objects.all().order_by('code').prefetch_related('rooms')

    for b in buildings:
        rooms_payload = [
            {"id": r.id, "name": r.name, "capacity": r.capacity}
            for r in b.rooms.all().order_by('name')
        ]
        items.append({
            "id": b.id,
            "code": b.code,
            "name": b.name,
            "description": b.description or "",
            "room_count": len(rooms_payload),
            "total_capacity": sum(r["capacity"] for r in rooms_payload) if rooms_payload else 0,
            "rooms": rooms_payload,
        })

    return JsonResponse({"success": True, "buildings": items})

@login_required
def get_rooms_by_building(request):
    """ดึงรายการห้องตามอาคาร - ปรับปรุงใหม่ให้ส่งข้อมูลเพิ่มเติม"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง', 'success': False}, status=403)
    
    building_id = request.GET.get('building_id')
    
    # Debug logging
    print(f"get_rooms_by_building called with building_id: {building_id}")
    
    if not building_id:
        return JsonResponse({
            'rooms': [],
            'success': True,
            'message': 'ไม่ได้ระบุ building_id'
        })
    
    try:
        # ตรวจสอบว่าอาคารมีอยู่จริง
        try:
            building = Building.objects.get(id=building_id)
            print(f"Found building: {building.name}")
        except Building.DoesNotExist:
            return JsonResponse({
                'error': 'ไม่พบอาคารที่ระบุ',
                'success': False,
                'rooms': []
            }, status=404)
        
        # ดึงห้องที่อยู่ในอาคารนี้
        rooms = ExamRoom.objects.filter(
            building_id=building_id, 
            is_active=True
        ).select_related('building').order_by('name')
        
        print(f"Found {rooms.count()} rooms in building {building.name}")
        
        room_data = []
        for room in rooms:
            room_data.append({
                'id': room.id,
                'name': room.name,
                'capacity': room.capacity,
                'full_name': f"{room.building.name} ห้อง {room.name}",
                'building_name': room.building.name,
                'display_text': f"ห้อง {room.name} (จุ {room.capacity} คน)",
                'has_projector': getattr(room, 'has_projector', False),
                'has_aircon': getattr(room, 'has_aircon', False),
            })
            print(f"  - Room: {room.name} (capacity: {room.capacity})")
        
        return JsonResponse({
            'rooms': room_data,
            'success': True,
            'building_name': building.name,
            'total_rooms': len(room_data),
            'message': f'พบห้องทั้งหมด {len(room_data)} ห้อง' if len(room_data) > 0 else 'ไม่มีห้องในอาคารนี้'
        })
        
    except Exception as e:
        print(f"Error in get_rooms_by_building: {str(e)}")
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False,
            'rooms': []
        }, status=500)


@csrf_exempt
@login_required
def manual_checkin(request):
    """เช็คชื่อด้วยตนเอง (สำหรับเจ้าหน้าที่)"""
    if request.method == 'POST' and request.user.is_staff:
        data = json.loads(request.body)
        student_id = data.get('student_id')
        subject_id = data.get('subject_id')
        status = data.get('status')
        
        student = get_object_or_404(StudentProfile, id=student_id)
        subject = get_object_or_404(ExamSubject, id=subject_id)
        
        attendance, created = Attendance.objects.get_or_create(
            student=student, 
            subject=subject,
            defaults={'status': status, 'checkin_time': timezone.now()}
        )
        
        if not created:
            attendance.status = status
            attendance.checkin_time = timezone.now()
            attendance.save()
        
        return JsonResponse({'status': 'success'})
    
@login_required
def ajax_search_users(request):
    """AJAX endpoint สำหรับการค้นหาผู้ใช้แบบ real-time"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    user_type = request.GET.get('type', 'teacher')  # teacher or student
    search_term = request.GET.get('search', '')
    
    if user_type == 'teacher':
        teachers = TeacherProfile.objects.select_related('user').filter(
            Q(teacher_id__icontains=search_term) |
            Q(user__first_name__icontains=search_term) |
            Q(user__last_name__icontains=search_term) |
            Q(user__email__icontains=search_term)
        )[:10]  # จำกัดผลลัพธ์ 10 รายการ
        
        results = [{
            'id': teacher.teacher_id,
            'name': teacher.user.get_full_name(),
            'email': teacher.user.email,
            'username': teacher.user.username,
            'is_active': teacher.user.is_active,
            'date_joined': teacher.user.date_joined.strftime('%d/%m/%Y')
        } for teacher in teachers]
        
    else:  # student
        students = StudentProfile.objects.select_related('user').filter(
            Q(student_id__icontains=search_term) |
            Q(user__first_name__icontains=search_term) |
            Q(user__last_name__icontains=search_term) |
            Q(user__email__icontains=search_term) |
            Q(student_class__icontains=search_term)
        )[:10]
        
        results = [{
            'id': student.student_id,
            'name': student.user.get_full_name(),
            'email': student.user.email,
            'username': student.user.username,
            'student_class': student.student_class,
            'student_number': student.student_number,
            'is_active': student.user.is_active,
            'date_joined': student.user.date_joined.strftime('%d/%m/%Y')
        } for student in students]
    
    return JsonResponse({'results': results})

@login_required
def get_class_students_count(request):
    """AJAX endpoint สำหรับดึงจำนวนนักเรียนในระดับชั้น"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    student_class = request.GET.get('class', '')
    
    if student_class:
        try:
            count = StudentProfile.objects.filter(student_class=student_class).count()
            return JsonResponse({
                'count': count,
                'class': student_class,
                'success': True
            })
        except Exception as e:
            return JsonResponse({
                'error': f'เกิดข้อผิดพลาด: {str(e)}',
                'success': False
            }, status=500)
    
    return JsonResponse({
        'error': 'ไม่ได้ระบุระดับชั้น',
        'success': False
    }, status=400)

@login_required
def check_teacher_conflicts(request):
    """AJAX endpoint สำหรับตรวจสอบความขัดแย้งของครู"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    date = request.GET.get('date')
    start_time = request.GET.get('start_time')
    end_time = request.GET.get('end_time')
    invigilator_id = request.GET.get('invigilator')
    secondary_invigilator_id = request.GET.get('secondary_invigilator')
    
    if not all([date, start_time, end_time]):
        return JsonResponse({
            'error': 'ข้อมูลไม่ครบถ้วน',
            'success': False
        }, status=400)
    
    try:
        conflicts = []
        
        # ตรวจสอบครูหลัก
        if invigilator_id:
            teacher_conflicts = ExamSubject.objects.filter(
                exam_date=date,
                start_time__lt=end_time,
                end_time__gt=start_time
            ).filter(
                Q(invigilator_id=invigilator_id) | 
                Q(secondary_invigilator_id=invigilator_id)
            ).select_related('invigilator__user')
            
            for conflict in teacher_conflicts:
                teacher_name = conflict.invigilator.user.get_full_name() if conflict.invigilator else 'ไม่ระบุ'
                conflicts.append(
                    f"ครู {teacher_name} มีตารางคุมสอบวิชา {conflict.subject_name} "
                    f"ในช่วงเวลา {conflict.start_time.strftime('%H:%M')} - {conflict.end_time.strftime('%H:%M')}"
                )
        
        # ตรวจสอบครูสำรอง
        if secondary_invigilator_id and secondary_invigilator_id != invigilator_id:
            teacher_conflicts = ExamSubject.objects.filter(
                exam_date=date,
                start_time__lt=end_time,
                end_time__gt=start_time
            ).filter(
                Q(invigilator_id=secondary_invigilator_id) | 
                Q(secondary_invigilator_id=secondary_invigilator_id)
            ).select_related('secondary_invigilator__user')
            
            for conflict in teacher_conflicts:
                teacher_name = conflict.secondary_invigilator.user.get_full_name() if conflict.secondary_invigilator else 'ไม่ระบุ'
                conflicts.append(
                    f"ครูสำรอง {teacher_name} มีตารางคุมสอบวิชา {conflict.subject_name} "
                    f"ในช่วงเวลา {conflict.start_time.strftime('%H:%M')} - {conflict.end_time.strftime('%H:%M')}"
                )
        
        return JsonResponse({
            'conflicts': conflicts,
            'has_conflicts': len(conflicts) > 0,
            'success': True
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

@login_required
def get_available_teachers(request):
    """AJAX endpoint สำหรับค้นหาครูที่ว่างในช่วงเวลา"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    date = request.GET.get('date')
    start_time = request.GET.get('start_time')
    end_time = request.GET.get('end_time')
    
    if not all([date, start_time, end_time]):
        return JsonResponse({
            'error': 'ข้อมูลไม่ครบถ้วน',
            'success': False
        }, status=400)
    
    try:
        # ค้นหาครูที่มีตารางในช่วงเวลานั้น
        busy_teachers = ExamSubject.objects.filter(
            exam_date=date,
            start_time__lt=end_time,
            end_time__gt=start_time
        ).values_list('invigilator_id', 'secondary_invigilator_id')
        
        # รวม ID ครูที่ไม่ว่าง
        busy_teacher_ids = set()
        for primary, secondary in busy_teachers:
            if primary:
                busy_teacher_ids.add(primary)
            if secondary:
                busy_teacher_ids.add(secondary)
        
        # ค้นหาครูที่ว่าง
        available_teachers = TeacherProfile.objects.exclude(
            id__in=busy_teacher_ids
        ).filter(
            user__is_active=True
        ).select_related('user').order_by('user__first_name')
        
        # จัดรูปแบบข้อมูล
        teachers_data = []
        for teacher in available_teachers:
            teachers_data.append({
                'id': teacher.id,
                'name': teacher.user.get_full_name(),
                'teacher_id': teacher.teacher_id,
                'email': teacher.user.email
            })
        
        return JsonResponse({
            'available_teachers': teachers_data,
            'total_available': len(teachers_data),
            'success': True
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

@login_required 
def auto_assign_room(request):
    """AJAX endpoint สำหรับจัดห้องอัตโนมัติ"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        date = data.get('date')
        start_time = data.get('start_time')
        end_time = data.get('end_time')
        student_count = data.get('student_count', 0)
        
        if not all([date, start_time, end_time]):
            return JsonResponse({
                'error': 'ข้อมูลไม่ครบถ้วน',
                'success': False
            }, status=400)
        
        # ค้นหาห้องที่เหมาะสมที่สุด
        available_room = find_available_room(date, start_time, end_time, student_count)
        
        if available_room:
            return JsonResponse({
                'room': {
                    'id': available_room.id,
                    'name': available_room.name,
                    'building': available_room.building.name,
                    'capacity': available_room.capacity,
                    'has_projector': available_room.has_projector,
                    'has_aircon': available_room.has_aircon
                },
                'success': True
            })
        else:
            return JsonResponse({
                'error': f'ไม่พบห้องว่างที่เหมาะสมสำหรับนักเรียน {student_count} คน',
                'success': False
            }, status=404)
            
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

@login_required
def auto_assign_teachers(request):
    """AJAX endpoint สำหรับจัดครูอัตโนมัติ"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        date = data.get('date')
        start_time = data.get('start_time')
        end_time = data.get('end_time')
        
        if not all([date, start_time, end_time]):
            return JsonResponse({
                'error': 'ข้อมูลไม่ครบถ้วน',
                'success': False
            }, status=400)
        
        # ค้นหาครูที่ว่าง
        available_teachers_response = get_available_teachers(request)
        if available_teachers_response.status_code != 200:
            return available_teachers_response
        
        available_teachers_data = json.loads(available_teachers_response.content)
        available_teachers = available_teachers_data.get('available_teachers', [])
        
        if len(available_teachers) < 1:
            return JsonResponse({
                'error': 'ไม่มีครูว่างในช่วงเวลานี้',
                'success': False
            }, status=404)
        
        # เลือกครูหลัก (คนแรก)
        primary_teacher = available_teachers[0]
        
        # เลือกครูสำรอง (คนที่สอง ถ้ามี)
        secondary_teacher = available_teachers[1] if len(available_teachers) > 1 else None
        
        return JsonResponse({
            'primary_teacher': primary_teacher,
            'secondary_teacher': secondary_teacher,
            'total_available': len(available_teachers),
            'success': True
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)

@login_required
def bulk_attendance_update(request, pk):  
    """อัพเดทการเข้าสอบแบบกลุ่ม"""
    subject = get_object_or_404(ExamSubject, id=pk)  # เปลี่ยนจาก subject_id เป็น pk
    
    # ตรวจสอบสิทธิ์
    if not (request.user.is_staff or 
            (hasattr(request.user, 'teacher_profile') and 
             subject.invigilator == request.user.teacher_profile)):
        return HttpResponseForbidden("คุณไม่มีสิทธิ์จัดการข้อมูลนี้")
    
    # ... rests of the code remain the same

@login_required
def cheating_reports(request):
    """รายการรายงานทุจริตทั้งหมด"""
    if not request.user.is_staff:
        return HttpResponseForbidden("คุณไม่มีสิทธิ์เข้าถึงส่วนนี้")
    
    reports = CheatingReport.objects.select_related(
        'attendance__student__user',
        'attendance__subject',
        'reported_by'
    ).order_by('-created_at')
    
    # ค้นหาและกรอง
    search = request.GET.get('search', '')
    cheating_type = request.GET.get('type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if search:
        reports = reports.filter(
            Q(attendance__student__user__first_name__icontains=search) |
            Q(attendance__student__user__last_name__icontains=search) |
            Q(attendance__student__student_id__icontains=search) |
            Q(attendance__subject__subject_name__icontains=search)
        )
    
    if cheating_type:
        reports = reports.filter(cheating_type=cheating_type)
    
    if date_from:
        reports = reports.filter(created_at__date__gte=date_from)
    
    if date_to:
        reports = reports.filter(created_at__date__lte=date_to)
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(reports, 20)
    page_number = request.GET.get('page')
    reports_page = paginator.get_page(page_number)
    
    return render(request, 'app/staff/cheating_reports.html', {
        'reports': reports_page,
        'search': search,
        'cheating_type': cheating_type,
        'date_from': date_from,
        'date_to': date_to,
        'cheating_types': CheatingReport.CHEATING_TYPE_CHOICES,
    })

def get_subject_status(subject):
    """ตรวจสอบสถานะของรายวิชาสอบ - เพิ่มการตรวจสอบความจุ"""
    student_count = subject.get_student_count()
    
    status = {
        'has_room': subject.room is not None,
        'has_teacher': subject.invigilator is not None,
        'has_secondary_teacher': subject.secondary_invigilator is not None,
        'has_students': student_count > 0,
        'room_capacity_sufficient': True,  # เพิ่มตัวแปรนี้
        'is_complete': False,
        'warnings': []
    }
    
    # ตรวจสอบความจุห้อง
    if status['has_room'] and subject.room.capacity < student_count:
        status['room_capacity_sufficient'] = False
        status['has_room'] = False  # ถือว่าไม่มีห้องถ้าจุไม่พอ
        status['warnings'].append(f'ห้องจุได้ {subject.room.capacity} คน แต่มีนักเรียน {student_count} คน')
    
    # ตรวจสอบความสมบูรณ์
    status['is_complete'] = (
        status['has_room'] and 
        status['has_teacher'] and 
        status['has_students']
    )
    
    # คำเตือนต่างๆ
    if not status['has_room']:
        if subject.room and not status['room_capacity_sufficient']:
            status['warnings'].append('ยังไม่ระบุห้อง (ห้องปัจจุบันจุไม่พอ)')
        else:
            status['warnings'].append('ยังไม่ระบุห้อง')
    
    if not status['has_teacher']:
        status['warnings'].append('ยังไม่มีครูคุมสอบ')
    
    if not status['has_secondary_teacher']:
        status['warnings'].append('ไม่มีครูสำรอง')
    
    if not status['has_students']:
        status['warnings'].append('ไม่มีนักเรียนลงทะเบียน')
    
    return status

@login_required
@csrf_exempt
def assign_room_manual(request, subject_id):
    """จัดห้องสอบแบบ manual"""
    if request.method == 'POST':
        subject = get_object_or_404(ExamSubject, id=subject_id)
        
        if not request.user.is_staff:
            return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
        
        try:
            data = json.loads(request.body)
            room_id = data.get('room_id')
            
            if room_id:
                room = ExamRoom.objects.get(id=room_id)
                
                # ตรวจสอบความพร้อมใช้งาน
                conflicts = ExamSubject.objects.filter(
                    exam_date=subject.exam_date,
                    start_time__lt=subject.end_time,
                    end_time__gt=subject.start_time,
                    room=room
                ).exclude(id=subject.id)
                
                if conflicts.exists():
                    return JsonResponse({
                        'error': 'ห้องนี้มีการใช้งานในช่วงเวลาดังกล่าวแล้ว'
                    }, status=400)
                
                if room.capacity < subject.get_student_count():
                    return JsonResponse({
                        'error': f'ห้องจุได้ {room.capacity} คน แต่มีนักเรียน {subject.get_student_count()} คน'
                    }, status=400)
                
                subject.room = room
                subject.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'จัดห้องสอบ {room.building.name} ห้อง {room.name} สำเร็จ',
                    'room_info': {
                        'name': room.name,
                        'building': room.building.name,
                        'capacity': room.capacity
                    }
                })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@login_required
@csrf_exempt 
def assign_teachers_manual(request, subject_id):
    """จัดครูคุมสอบแบบ manual"""
    if request.method == 'POST':
        subject = get_object_or_404(ExamSubject, id=subject_id)
        
        if not request.user.is_staff:
            return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
        
        try:
            data = json.loads(request.body)
            primary_teacher_id = data.get('primary_teacher_id')
            secondary_teacher_id = data.get('secondary_teacher_id')
            
            # ตรวจสอบครูหลัก
            if primary_teacher_id:
                teacher = TeacherProfile.objects.get(id=primary_teacher_id)
                
                # ตรวจสอบความขัดแย้ง
                conflicts = ExamSubject.objects.filter(
                    exam_date=subject.exam_date,
                    start_time__lt=subject.end_time,
                    end_time__gt=subject.start_time
                ).filter(
                    Q(invigilator=teacher) | Q(secondary_invigilator=teacher)
                ).exclude(id=subject.id)
                
                if conflicts.exists():
                    return JsonResponse({
                        'error': f'ครู {teacher.user.get_full_name()} มีตารางคุมสอบในช่วงเวลานี้แล้ว'
                    }, status=400)
                
                subject.invigilator = teacher
            
            # ตรวจสอบครูสำรอง
            if secondary_teacher_id and secondary_teacher_id != primary_teacher_id:
                teacher = TeacherProfile.objects.get(id=secondary_teacher_id)
                
                conflicts = ExamSubject.objects.filter(
                    exam_date=subject.exam_date,
                    start_time__lt=subject.end_time,
                    end_time__gt=subject.start_time
                ).filter(
                    Q(invigilator=teacher) | Q(secondary_invigilator=teacher)
                ).exclude(id=subject.id)
                
                if conflicts.exists():
                    return JsonResponse({
                        'error': f'ครูสำรอง {teacher.user.get_full_name()} มีตารางคุมสอบในช่วงเวลานี้แล้ว'
                    }, status=400)
                
                subject.secondary_invigilator = teacher
            
            subject.save()
            
            return JsonResponse({
                'success': True,
                'message': 'จัดครูคุมสอบสำเร็จ',
                'teacher_info': {
                    'primary': subject.invigilator.user.get_full_name() if subject.invigilator else None,
                    'secondary': subject.secondary_invigilator.user.get_full_name() if subject.secondary_invigilator else None
                }
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@login_required
def get_available_resources_for_manual_assignment(request):
    """AJAX endpoint สำหรับดึงครูและห้องที่ว่างสำหรับการจัดแบบ manual"""
    date = request.GET.get('date')
    start_time = request.GET.get('start_time')
    end_time = request.GET.get('end_time')
    student_count = int(request.GET.get('student_count', 0))
    
    if not all([date, start_time, end_time]):
        return JsonResponse({'error': 'ข้อมูลไม่ครบ'}, status=400)
    
    # หาห้องที่ว่าง
    busy_rooms = ExamSubject.objects.filter(
        exam_date=date,
        start_time__lt=end_time,
        end_time__gt=start_time
    ).values_list('room_id', flat=True)
    
    available_rooms = ExamRoom.objects.exclude(
        id__in=busy_rooms
    ).filter(
        capacity__gte=student_count,
        is_active=True
    ).select_related('building').order_by('building__name', 'name')
    
    # หาครูที่ว่าง
    busy_teachers = ExamSubject.objects.filter(
        exam_date=date,
        start_time__lt=end_time,
        end_time__gt=start_time
    ).values_list('invigilator_id', 'secondary_invigilator_id')
    
    busy_teacher_ids = set()
    for primary, secondary in busy_teachers:
        if primary:
            busy_teacher_ids.add(primary)
        if secondary:
            busy_teacher_ids.add(secondary)
    
    available_teachers = TeacherProfile.objects.exclude(
        id__in=busy_teacher_ids
    ).filter(
        user__is_active=True
    ).select_related('user').order_by('user__first_name')
    
    # จัดรูปแบบข้อมูล
    rooms_data = []
    for room in available_rooms:
        rooms_data.append({
            'id': room.id,
            'name': room.name,
            'building': room.building.name,
            'capacity': room.capacity,
            'full_name': f"{room.building.name} ห้อง {room.name}",
            'display_text': f"{room.building.name} ห้อง {room.name} (จุ {room.capacity} คน)"
        })
    
    teachers_data = []
    for teacher in available_teachers:
        teachers_data.append({
            'id': teacher.id,
            'name': teacher.user.get_full_name(),
            'teacher_id': teacher.teacher_id,
            'display_text': f"{teacher.user.get_full_name()} ({teacher.teacher_id})"
        })
    
    return JsonResponse({
        'available_rooms': rooms_data,
        'available_teachers': teachers_data,
        'success': True
    })

@login_required
@csrf_exempt
def bulk_auto_assign(request):
    """จัดครูและห้องสอบอัตโนมัติแบบ bulk"""
    if request.method == 'POST':
        if not request.user.is_staff:
            return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง'}, status=403)
        
        try:
            data = json.loads(request.body)
            subject_ids = data.get('subject_ids', [])
            assignment_type = data.get('type', 'both')  # both, teachers, rooms
            
            results = {
                'success_count': 0,
                'error_count': 0,
                'errors': []
            }
            
            for subject_id in subject_ids:
                try:
                    subject = ExamSubject.objects.get(id=subject_id)
                    
                    if assignment_type in ['both', 'rooms'] and not subject.room:
                        # จัดห้อง
                        available_room = find_available_room(
                            subject.exam_date,
                            subject.start_time,
                            subject.end_time,
                            subject.get_student_count()
                        )
                        if available_room:
                            subject.room = available_room
                    
                    if assignment_type in ['both', 'teachers'] and not subject.invigilator:
                        # จัดครู
                        available_teachers = find_available_teachers(
                            subject.exam_date,
                            subject.start_time,
                            subject.end_time
                        )
                        if len(available_teachers) >= 1:
                            subject.invigilator = available_teachers[0]
                        if len(available_teachers) >= 2:
                            subject.secondary_invigilator = available_teachers[1]
                    
                    subject.save()
                    results['success_count'] += 1
                    
                except Exception as e:
                    results['errors'].append({
                        'subject_id': subject_id,
                        'error': str(e)
                    })
                    results['error_count'] += 1
            
            return JsonResponse({
                'success': True,
                'message': f'จัดการสำเร็จ {results["success_count"]} วิชา',
                'results': results
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@login_required
@csrf_exempt
def ajax_teachers_add(request):
    """เพิ่มครูใหม่ผ่าน AJAX"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'ไม่มีสิทธิ์เข้าถึง', 'success': False}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed', 'success': False}, status=405)
    
    try:
        # ตรวจสอบประเภทข้อมูลที่ส่งมา
        content_type = request.content_type or ''
        
        if 'application/json' in content_type:
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({'error': 'Invalid JSON format', 'success': False}, status=400)
        else:
            # FormData - แปลงเป็น dict
            data = {
                'username': request.POST.get('username', '').strip(),
                'teacher_id': request.POST.get('teacher_id', '').strip(),
                'first_name': request.POST.get('first_name', '').strip(),
                'last_name': request.POST.get('last_name', '').strip(),
                'email': request.POST.get('email', '').strip(),
                'password': request.POST.get('password', '').strip()
            }
        
        # ข้อมูลที่จำเป็น
        required_fields = ['username', 'teacher_id', 'first_name', 'last_name', 'email', 'password']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({
                    'error': f'กรุณากรอก{field}',
                    'success': False
                }, status=400)
        
        # ตรวจสอบว่าข้อมูลไม่ซ้ำ
        if User.objects.filter(username=data['username']).exists():
            return JsonResponse({
                'error': f'Username {data["username"]} มีอยู่แล้ว',
                'success': False
            }, status=400)
        
        if User.objects.filter(email=data['email']).exists():
            return JsonResponse({
                'error': f'อีเมล {data["email"]} มีอยู่แล้ว',
                'success': False
            }, status=400)
        
        if TeacherProfile.objects.filter(teacher_id=data['teacher_id']).exists():
            return JsonResponse({
                'error': f'รหัสครู {data["teacher_id"]} มีอยู่แล้ว',
                'success': False
            }, status=400)
        
        # สร้างครูใหม่
        with transaction.atomic():
            user = User.objects.create_user(
                username=data['username'],
                email=data['email'],
                first_name=data['first_name'],
                last_name=data['last_name'],
                password=data['password'],
                is_teacher=True,
                is_active=True
            )
            
            teacher = TeacherProfile.objects.create(
                user=user,
                teacher_id=data['teacher_id']
            )
        
        return JsonResponse({
            'success': True,
            'message': f'เพิ่มครู {user.get_full_name()} สำเร็จ!',
            'teacher': {
                'id': teacher.id,
                'teacher_id': teacher.teacher_id,
                'name': user.get_full_name(),
                'email': user.email,
                'username': user.username
            }
        })
        
    except Exception as e:
        print(f"Error in ajax_teachers_add: {str(e)}")
        return JsonResponse({
            'error': f'เกิดข้อผิดพลาด: {str(e)}',
            'success': False
        }, status=500)